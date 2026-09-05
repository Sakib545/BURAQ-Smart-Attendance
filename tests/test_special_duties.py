"""Salary/duty integration tests, including real HTTP authorization and exports."""
import base64
import io
import json
import os
from datetime import datetime, timedelta
import pytest
from fastapi.testclient import TestClient
from itsdangerous import TimestampSigner
from openpyxl import load_workbook
from app import main, special_duties
from app.database import get_db, init_db
from app.payroll_ops import bulk_finalize, finalize_blockers

MONTH='2026-07'


def client(role='super_admin'):
    c=TestClient(main.app)
    if role:
        session={'role':role,'user_name':'Duty Reviewer'}
        session.update({'admin':True} if role=='super_admin' else {'hr_id':987654})
        c.cookies.set('session',TimestampSigner(os.environ['SESSION_SECRET']).sign(base64.b64encode(json.dumps(session).encode())).decode())
    return c


@pytest.fixture
def employee():
    init_db()
    with get_db() as c:
        c.execute("INSERT INTO employees(staff_id,name,shift,is_active,fixed_salary,join_date) VALUES('SPECIAL-TEST','Special Test','morning',?,31000,'2026-07-01')",(True,))
        eid=c.execute("SELECT id FROM employees WHERE staff_id='SPECIAL-TEST'").fetchone()['id']
        for day in range(7):
            c.execute('INSERT INTO duty_schedules(employee_id,weekday,start_time,end_time,break_minutes,is_active) VALUES(?,?,?,?,?,?)',(eid,day,'09:00','17:00',0,True))
        for day in range(1,32):
            date=f'2026-07-{day:02}'
            c.execute('INSERT INTO attendance(employee_id,work_date,check_in,check_out,status) VALUES(?,?,?,?,?)',(eid,date,date+' 09:00:00',date+' 17:00:00','present'))
    yield eid
    with get_db() as c:
        c.execute('DELETE FROM payroll_change_logs WHERE payroll_id IN (SELECT id FROM payroll_records WHERE employee_id=?)',(eid,))
        for table in ['special_duties','payroll_records','attendance','duty_schedules','custom_duties','leave_requests']:
            c.execute(f'DELETE FROM {table} WHERE employee_id=?',(eid,))
        c.execute('DELETE FROM employees WHERE id=?',(eid,))


def add(eid,kind='night',start='18:00',end='20:00',amount='300',date='2026-07-02'):
    return special_duties.add_completed(eid,date,kind,start,end,amount,'Extra work completed','Reviewer')


def save(eid,c=None):
    c=c or client()
    r=c.post('/payroll',data={'employee_id':eid,'salary_month':MONTH,'fixed_salary':31000},follow_redirects=False)
    assert r.status_code==303,r.text
    with get_db() as db:
        return dict(db.execute('SELECT * FROM payroll_records WHERE employee_id=?',(eid,)).fetchone())


def test_regular_friday_and_night_stay_in_basic(employee):
    r=main._calculate_employee_payroll(employee,MONTH,31000,0)
    assert r['earned_basic_salary']==31000 and r['total_allowance']==0
    with get_db() as c:
        c.execute("UPDATE duty_schedules SET start_time='22:00',end_time='06:00' WHERE employee_id=?",(employee,))
    r=main._calculate_employee_payroll(employee,MONTH,31000,0)
    assert r['earned_basic_salary']==31000 and r['total_allowance']==0
    with pytest.raises(ValueError,match='regular duty'):
        add(employee,start='01:00',end='03:00')  # previous night's regular duty


@pytest.mark.parametrize('kind',['night','friday','eid','other'])
def test_dated_extra_added_once_and_does_not_change_basic(employee,kind):
    add(employee,kind=kind)
    r=main._calculate_employee_payroll(employee,MONTH,31000,0)
    assert r['salary_divisor']==31 and r['worked']==31
    assert r['earned_basic_salary']==31000 and r['net_salary']==31300
    assert len(r['special_duty_records'])==1
    # Caller-supplied monthly allowances cannot double the dated payment.
    r=main._calculate_employee_payroll(employee,MONTH,31000,0,night_allowance=300)
    assert r['net_salary']==31300


def test_overlap_duplicate_cancel_and_replacement(employee):
    with pytest.raises(ValueError,match='regular duty'):
        add(employee,start='10:00',end='12:00')
    sid=add(employee)
    with pytest.raises(ValueError,match='Duplicate'):
        add(employee,start='21:00',end='22:00')
    with pytest.raises(ValueError,match='overlapping'):
        add(employee,kind='other',start='19:00',end='21:00')
    special_duties.cancel(sid,'Wrong duty amount','Admin')
    add(employee,amount='450')
    r=main._calculate_employee_payroll(employee,MONTH,31000,0)
    assert r['net_salary']==31450
    with get_db() as c:
        old=c.execute('SELECT * FROM special_duties WHERE id=?',(sid,)).fetchone()
        assert old['payment_amount']==300 and old['cancel_reason']=='Wrong duty amount'


def test_roster_change_blocks_special_and_finalize(employee):
    add(employee)
    row=save(employee)
    with get_db() as c:
        c.execute("UPDATE duty_schedules SET end_time='20:00' WHERE employee_id=?",(employee,))
    blockers=finalize_blockers(row)
    assert any('overlaps regular' in s for s in blockers)
    assert main._calculate_employee_payroll(employee,MONTH,31000,0)['total_allowance']==0


def test_stale_draft_must_be_resaved_before_single_or_bulk_finalize(employee):
    row=save(employee)
    add(employee)
    c=client()
    response=c.post(f"/payroll/{row['id']}/status",data={'month':MONTH,'status':'finalized'})
    assert response.status_code==409
    assert any('Special duty changed' in x for x in finalize_blockers(row))
    bulk=bulk_finalize(MONTH,'Reviewer')
    assert not any(r['id']==row['id'] for r in bulk['ready'])
    row=save(employee,c)
    assert not finalize_blockers(row)
    response=c.post(f"/payroll/{row['id']}/status",data={'month':MONTH,'status':'finalized'},follow_redirects=False)
    assert response.status_code==303,response.text
    with pytest.raises(ValueError,match='locked'):
        add(employee,kind='other',start='21:00',end='22:00')
    with get_db() as db:
        sid=db.execute('SELECT id FROM special_duties WHERE employee_id=?',(employee,)).fetchone()['id']
    with pytest.raises(ValueError,match='locked'):
        special_duties.cancel(sid,'Correct payment','Admin')


def test_preview_sheet_payslip_and_excel_agree(employee):
    add(employee,kind='other',amount='456.78')
    c=client(); row=save(employee,c)
    preview=c.get('/payroll/preview',params={'employee_id':employee,'month':MONTH,'fixed_salary':31000})
    assert preview.status_code==200
    assert preview.json()['net_salary']==31456.78
    assert '31,456.78' in c.get('/payroll?month='+MONTH).text
    workbook=load_workbook(io.BytesIO(c.get('/payroll/export.xlsx?month='+MONTH).content))
    sheet=workbook['Salary Sheet']
    data=next(r for r in sheet.iter_rows(min_row=5,values_only=True) if r[1]=='SPECIAL-TEST')
    assert data[23]==31456.78 and data[26]==456.78
    extra=workbook['Special Duty']
    extra_row=next(r for r in extra.iter_rows(min_row=2,values_only=True) if r[0]=='SPECIAL-TEST')
    assert extra_row[5]==456.78
    for url in ['/payroll/export.pdf?month='+MONTH,f"/payroll/{row['id']}/payslip.pdf"]:
        response=c.get(url)
        assert response.status_code==200 and response.content.startswith(b'%PDF'),response.text[:200]
    page=c.get('/payroll/special-duties?month='+MONTH)
    assert page.status_code==200 and '456.78' in page.text


@pytest.mark.parametrize('role,status',[(None,401),('viewer',403),('hr_officer',403)])
def test_special_write_permission(employee,role,status):
    r=client(role).post('/payroll/special-duties',data={'employee_id':employee,'duty_date':'2026-07-02','duty_type':'night','start_time':'18:00','end_time':'20:00','payment_amount':300,'note':'Completed work','completed':'yes','month':MONTH})
    assert r.status_code==status


def test_http_create_cancel_and_manual_allowance_rejected(employee):
    c=client('admin')
    r=c.post('/payroll/special-duties',data={'employee_id':employee,'duty_date':'2026-07-02','duty_type':'night','start_time':'18:00','end_time':'20:00','payment_amount':300,'note':'Completed work','completed':'yes','month':MONTH},follow_redirects=False)
    assert r.status_code==303 and 'error=' in r.headers['location'] and r.headers['location'].endswith('error=')
    r=c.post('/payroll',data={'employee_id':employee,'salary_month':MONTH,'fixed_salary':31000,'night_allowance':300})
    assert r.status_code==400
    with get_db() as db: sid=db.execute('SELECT id FROM special_duties WHERE employee_id=?',(employee,)).fetchone()['id']
    r=c.post(f'/payroll/special-duties/{sid}/cancel',data={'reason':'Wrong date entered','month':MONTH},follow_redirects=False)
    assert r.status_code==303
    assert main._calculate_employee_payroll(employee,MONTH,31000,0)['net_salary']==31000


def test_joining_date_and_future_special_rejected(employee):
    with get_db() as c: c.execute("UPDATE employees SET join_date='2026-07-11' WHERE id=?",(employee,))
    r=main._calculate_employee_payroll(employee,MONTH,31000,0)
    assert r['scheduled']==21 and r['absent']==0 and r['net_salary']==21000
    with pytest.raises(ValueError,match='joining'):
        add(employee)
    with pytest.raises(ValueError,match='completed'):
        add(employee,date='2099-01-01')


def test_schema_migration_is_repeatable_and_keeps_records(employee):
    sid=add(employee)
    init_db()
    with get_db() as c:
        assert c.execute('SELECT payment_amount FROM special_duties WHERE id=?',(sid,)).fetchone()['payment_amount']==300
