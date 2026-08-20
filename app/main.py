import csv
import asyncio
import hashlib
import hmac
import io
import json
import logging
import os
import re
import secrets
import tempfile
import time
import uuid
from datetime import datetime, timedelta
from pathlib import Path
from zoneinfo import ZoneInfo
from html import escape
from urllib.parse import quote_plus

from fastapi import FastAPI, BackgroundTasks, Form, HTTPException, Query, Request
from fastapi.responses import HTMLResponse, JSONResponse, PlainTextResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from starlette.middleware.sessions import SessionMiddleware

from app import ui
from app.config import settings
from app.database import database_kind, database_ok, database_warning, get_db, init_db
from app.runtime import configured, get_setting, set_setting, import_environment_defaults, get_stored_setting, restore_stored_setting
from app.employee_seed import import_employees
from app.whatsapp import handle, send_approval_flow, send_document_bytes, send_selfie_review_result, send_text
from app.location_links import verify_location_token
from app.reminders import reminder_worker
from app.time_format import format_time_12h
from app.payroll import PayrollInput, adjustment_reason_required, calculate_payroll
from app.backups import backup_status, create_full_backup, inspect_backup, payroll_backup_worker, read_backup, restore_full_backup, upload_offsite
from app.services import approve_pending_attendance, phones_match, receive_location, state
from app.shift_rules import (
    CUTOFF_KEY, FIRST_END_KEY, FIRST_START_KEY, GRACE_KEY,
    SECOND_END_KEY, SECOND_START_KEY, get_shift_rules, save_shift_rules, shift_window,
)

logging.basicConfig(level=getattr(logging, settings.log_level.upper(), logging.INFO))
logger = logging.getLogger(__name__)
APP_VERSION = "9.24.0"

app = FastAPI(title=settings.app_name, version=APP_VERSION, docs_url=None, redoc_url=None)
app.add_middleware(SessionMiddleware, secret_key=os.getenv("SESSION_SECRET", secrets.token_urlsafe(32)), https_only=settings.environment == "production", same_site="lax")

STATIC_DIR = Path(__file__).resolve().parent.parent / "static"
app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

@app.middleware("http")
async def request_context(request: Request, call_next):
    request_id = request.headers.get("x-request-id") or uuid.uuid4().hex[:16]
    started = time.perf_counter()
    try:
        response = await call_next(request)
    except Exception:
        logger.exception("Unhandled request error request_id=%s method=%s path=%s", request_id, request.method, request.url.path)
        return JSONResponse({"detail": "Internal server error", "request_id": request_id}, status_code=500)
    duration_ms = round((time.perf_counter() - started) * 1000, 1)
    response.headers["X-Request-ID"] = request_id
    response.headers["X-Response-Time-ms"] = str(duration_ms)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(), microphone=(), geolocation=(self)"
    logger.info("request_id=%s method=%s path=%s status=%s duration_ms=%s", request_id, request.method, request.url.path, response.status_code, duration_ms)
    return response

def nav_badges(request: Request) -> dict:
    """Counts shown next to sidebar items. Never allowed to break a page."""
    if not has_permission(request, "approvals_view"):
        return {}
    try:
        with get_db() as c:
            waiting = c.execute("SELECT COUNT(*) c FROM attendance_fingerprints WHERE review_status='pending'").fetchone()["c"]
        return {"duplicates": int(waiting or 0)}
    except Exception:
        logger.warning("nav badge count failed", exc_info=True)
        return {}


def layout(title: str, body: str, request: Request | None = None, active: str = ""):
    """Wrap route markup in the application shell."""
    chrome = request is not None and logged_in(request)
    nav_groups: list = []
    user_name = role_label = today_line = ""
    if chrome:
        nav_groups = ui.build_nav(lambda flag: has_permission(request, flag), nav_badges(request))
        user_name = str(request.session.get("user_name", "Admin"))
        role_label = str(request.session.get("role", "super_admin")).replace("_", " ").title()
        today_line = datetime.now(ZoneInfo(settings.timezone)).strftime("%a %d %b, %I:%M %p")
    return HTMLResponse(ui.render_page(
        title=title,
        body=body,
        chrome=chrome,
        nav_groups=nav_groups,
        active=active,
        user_name=user_name,
        role_label=role_label,
        today_line=today_line,
    ))


def hash_password(password: str, salt: str | None = None):
    salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 180000).hex()
    return f"{salt}${digest}"

def verify_password(password: str, stored: str):
    try:
        salt, _ = stored.split("$", 1)
        return hmac.compare_digest(hash_password(password, salt), stored)
    except Exception:
        return False

PERMISSION_CATALOG = {
    "dashboard_view": ("Dashboard", "Dashboard summary দেখবে"),
    "employees_view": ("Employees: View", "Employee list দেখবে"),
    "employees_add": ("Employees: Add", "নতুন employee যোগ করবে"),
    "employees_edit": ("Employees: Edit", "Employee তথ্য পরিবর্তন করবে"),
    "employees_delete": ("Employees: Delete", "Employee মুছতে পারবে"),
    "performance_view": ("Performance: View", "Performance review দেখবে"),
    "performance_manage": ("Performance: Create", "Performance review তৈরি করবে"),
    "face_reset": ("Face AI: Reset", "Employee face profile reset করবে"),
    "face_security_view": ("Face Security: View", "Face verification, rejection ও spoof attempt দেখবে"),
    "approvals_view": ("Approvals: View", "Pending registration দেখবে"),
    "approvals_manage": ("Approvals: Approve/Reject", "Registration approve/reject করবে"),
    "reports_view": ("Reports: View", "Attendance report দেখবে"),
    "reports_export": ("Reports: Export", "CSV/PDF/Excel export করবে"),
    "payroll_view": ("Payroll: View", "Private salary records দেখবে"),
    "payroll_manage": ("Payroll: Manage", "Salary, overtime, bonus ও payment status পরিবর্তন করবে"),
    "payroll_export": ("Payroll: Export", "Private payslip ও monthly Excel/PDF export করবে"),
    "duty_view": ("Duty: View", "Employee duty roster ও reminder status দেখবে"),
    "duty_manage": ("Duty: Manage", "Duty schedule তৈরি ও পরিবর্তন করবে"),
    "leave_view": ("Leave: View", "Leave এবং correction request দেখবে"),
    "leave_manage": ("Leave: Approve/Reject", "Leave ও correction approve/reject করবে"),
    "attendance_edit": ("Attendance: Correct", "Attendance correction request করবে"),
    "shift_manage": ("Shift Management", "Shift manage করবে"),
    "department_manage": ("Department Management", "Department manage করবে"),
    "audit_view": ("Audit Log: View", "Activity log দেখবে"),
    "settings_view": ("Settings: View", "সাধারণ settings page দেখবে"),
    "whatsapp_settings": ("WhatsApp Settings", "Token, Phone ID ও Webhook দেখবে/পরিবর্তন করবে"),
    "user_accounts_view": ("User Accounts: View", "Admin/HR account list দেখবে"),
    "user_accounts_manage": ("User Accounts: Manage", "Admin/HR account create, disable বা delete করবে"),
}
DEFAULT_ROLE_PERMISSIONS = {
    "admin": {"dashboard_view","employees_view","employees_add","employees_edit","performance_view","performance_manage","face_reset","face_security_view","approvals_view","approvals_manage","reports_view","reports_export","payroll_view","payroll_manage","payroll_export","duty_view","duty_manage","leave_view","leave_manage","attendance_edit","shift_manage","department_manage","audit_view"},
    "hr_manager": {"dashboard_view","employees_view","employees_add","employees_edit","performance_view","performance_manage","face_reset","face_security_view","approvals_view","approvals_manage","reports_view","reports_export","payroll_view","payroll_manage","payroll_export","duty_view","duty_manage","leave_view","leave_manage","attendance_edit","shift_manage","department_manage","audit_view"},
    "hr_executive": {"dashboard_view","employees_view","employees_add","employees_edit","performance_view","performance_manage","approvals_view","approvals_manage","reports_view","reports_export","leave_view","leave_manage","attendance_edit"},
    "hr_officer": {"dashboard_view","employees_view","performance_view","reports_view","leave_view"},
    "viewer": {"dashboard_view","reports_view"},
}

def logged_in(request: Request): return bool(request.session.get("admin") or request.session.get("hr_id"))
def require_login(request: Request):
    if not logged_in(request): raise HTTPException(401, "Login required")

def current_permissions(request: Request):
    require_login(request)
    cached = getattr(request.state, "permission_cache", None)
    if cached is not None:
        return cached
    if request.session.get("role") == "super_admin" and request.session.get("admin"):
        allowed = set(PERMISSION_CATALOG) | {"*"}
        request.state.permission_cache = allowed
        return allowed
    account_id = request.session.get("hr_id")
    if not account_id:
        return set()
    with get_db() as c:
        rows = c.execute("SELECT permission FROM account_permissions WHERE account_id=?", (account_id,)).fetchall()
        raw = {r["permission"] for r in rows}
        if "__configured__" in raw:
            allowed = {p for p in raw if p in PERMISSION_CATALOG}
            request.state.permission_cache = allowed
            return allowed
        explicit = {p for p in raw if p in PERMISSION_CATALOG}
        if explicit:
            request.state.permission_cache = explicit
            return explicit
        role = request.session.get("role", "viewer")
        allowed = set(DEFAULT_ROLE_PERMISSIONS.get(role, set()))
        request.state.permission_cache = allowed
        return allowed

def has_permission(request: Request, permission: str):
    if not logged_in(request):
        return False
    allowed = current_permissions(request)
    return "*" in allowed or permission in allowed

def require_permission(request: Request, permission: str):
    require_login(request)
    if not has_permission(request, permission):
        raise HTTPException(403, "Permission denied")

def require_super_admin(request: Request):
    require_login(request)
    if request.session.get("role") != "super_admin" or not request.session.get("admin"):
        raise HTTPException(403, "Super Admin access required")

def audit(request: Request, action: str, target_type: str = "", target_id: str = "", details: str = "", db=None):
    """Write an audit event without allowing audit logging to break the main action.

    When the caller already has an open write transaction, pass that connection as
    ``db``. Opening a second SQLite write transaction used to wait for 30 seconds
    and then raise ``database is locked`` during HR login.
    """
    actor_type = "admin" if request.session.get("admin") else "hr"
    actor_id = str(request.session.get("hr_id", "admin"))
    actor_name = str(request.session.get("user_name", "Admin"))
    ip = request.client.host if request.client else ""
    params = (actor_type, actor_id, actor_name, action, target_type, target_id, details, ip)
    sql = "INSERT INTO audit_logs(actor_type,actor_id,actor_name,action,target_type,target_id,details,ip_address) VALUES(?,?,?,?,?,?,?,?)"
    try:
        if db is not None:
            db.execute(sql, params)
        else:
            with get_db() as c:
                c.execute(sql, params)
    except Exception:
        logger.exception("Audit logging failed: action=%s target=%s:%s", action, target_type, target_id)

def base_url(request: Request): return str(request.base_url).rstrip("/")

def admin_setup_hash() -> str:
    """Read setup state without converting a database outage into first-time setup."""
    try:
        with get_db() as c:
            row=c.execute("SELECT value FROM system_settings WHERE key=?",("admin_password_hash",)).fetchone()
        return str(row["value"]) if row and row["value"] else ""
    except Exception as exc:
        logger.exception("Could not read persistent Admin setup state")
        raise HTTPException(503,"Database temporarily unavailable. Admin setup was not reset; please retry shortly.") from exc

def admin_setup_completed() -> bool:
    try:
        with get_db() as c:
            row=c.execute("SELECT value FROM system_settings WHERE key=?",("admin_setup_completed",)).fetchone()
        return bool(row and str(row["value"]) == "1")
    except Exception as exc:
        logger.exception("Could not read persistent Admin setup marker")
        raise HTTPException(503,"Database temporarily unavailable. Please retry shortly.") from exc

@app.on_event("startup")
def startup():
    issues = settings.production_issues()
    if issues:
        raise RuntimeError("Production configuration invalid: " + "; ".join(issues))
    for warning in settings.production_warnings():
        logger.warning("Optional configuration warning: %s", warning)
    init_db()
    import_environment_defaults()
    if not get_setting("admin_email"):
        set_setting("admin_email", os.getenv("SUPER_ADMIN_EMAIL", "admin@buraq.com").strip().lower())
    if not get_setting("admin_name"):
        set_setting("admin_name", os.getenv("SUPER_ADMIN_NAME", "Super Admin").strip())
    # Upgrade existing installations to the permanent one-time setup marker.
    if get_setting("admin_password_hash") and not get_setting("admin_setup_completed"):
        set_setting("admin_setup_completed","1")
    imported = import_employees()
    logger.info("BURAQ v%s started database=%s employees_synced=%s", APP_VERSION, database_kind(), imported)

@app.on_event("startup")
async def start_reminders():
    app.state.reminder_task=asyncio.create_task(reminder_worker())
    app.state.payroll_backup_task=asyncio.create_task(payroll_backup_worker())

@app.on_event("shutdown")
async def stop_reminders():
    for name in ("reminder_task","payroll_backup_task"):
        task=getattr(app.state,name,None)
        if task:
            task.cancel()
            try: await task
            except asyncio.CancelledError: pass

@app.get("/health")
def health():
    return {"status": "ok", "service": settings.app_name, "version": APP_VERSION}


@app.get("/ready")
def ready():
    db_ok = database_ok()
    configured_ok = configured()
    setup_ok=False
    if db_ok:
        try: setup_ok=bool(admin_setup_hash()) and admin_setup_completed()
        except HTTPException: setup_ok=False
    payload = {
        "status": "ready" if db_ok else "not_ready",
        "database": database_kind(),
        "database_ok": db_ok,
        "whatsapp_configured": configured_ok,
        "admin_setup_complete": setup_ok,
        "version": APP_VERSION,
    }
    return JSONResponse(payload, status_code=200 if db_ok else 503)

@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    admin_hash=admin_setup_hash(); completed=admin_setup_completed()
    if completed and not admin_hash:
        raise HTTPException(503,"Admin setup is protected but credentials are unavailable. Restore the latest backup.")
    if not admin_hash:
        return RedirectResponse("/setup", 302)
    if not logged_in(request): return RedirectResponse("/login", 302)
    return RedirectResponse("/dashboard", 302)

@app.get("/setup", response_class=HTMLResponse)
def setup_page(request: Request):
    admin_hash=admin_setup_hash(); completed=admin_setup_completed()
    if completed and not admin_hash:
        raise HTTPException(503,"Admin setup is protected. Restore the latest backup instead of creating a new Admin.")
    if admin_hash:
        return RedirectResponse("/dashboard" if logged_in(request) else "/login", 302)
    cfg_note = "<div class='notice'>Railway Variables থেকে WhatsApp configuration পাওয়া গেছে। শুধু Admin password তৈরি করুন।</div>" if configured() else "<div class='notice' style='background:#fef3c7;color:#92400e'>WhatsApp credentials পরে Dashboard → Settings থেকে যোগ করতে পারবেন।</div>"
    body=f"<div class='login'><div class='card'><div class='title'>BURAQ Smart Attendance</div><p class='sub'>প্রথমবারের নিরাপদ Admin setup</p>{cfg_note}<form method='post'><label>Super Admin email</label><input type='email' name='email' value='admin@buraq.com' required><label>নতুন Admin password</label><input type='password' name='password' minlength='8' required><label>Confirm password</label><input type='password' name='confirm_password' minlength='8' required><button class='btn' type='submit'>Create Admin & Open Dashboard</button></form><p class='sub'>এটি শুধু একবারই করতে হবে। পরে Settings থেকে email/password পরিবর্তন করা যাবে।</p></div></div>"
    return layout("Initial Setup", body)

@app.post("/setup")
def save_setup(request: Request, email: str = Form(...), password: str = Form(...), confirm_password: str = Form(...)):
    if admin_setup_hash() or admin_setup_completed():
        raise HTTPException(403)
    if password != confirm_password or len(password) < 8:
        raise HTTPException(400, "Passwords do not match or are too short")
    values={"admin_email":email.strip().lower(),"admin_name":"Super Admin","admin_password_hash":hash_password(password),"admin_setup_completed":"1"}
    with get_db() as c:
        for key,value in values.items():
            c.execute("INSERT INTO system_settings(key,value,updated_at) VALUES(?,?,CURRENT_TIMESTAMP) ON CONFLICT(key) DO UPDATE SET value=excluded.value,updated_at=CURRENT_TIMESTAMP",(key,value))
    request.session["admin"] = True
    request.session["role"] = "super_admin"
    request.session["user_name"] = get_setting("admin_name", "Super Admin")
    return RedirectResponse("/dashboard", 303)

@app.get("/login", response_class=HTMLResponse)
def login_page(error: str = ""):
    msg = "<div class='notice' style='background:#fee2e2;color:#991b1b'>Email অথবা Password সঠিক নয়।</div>" if error else ""
    body = f"""<div class='login'><div class='card'><div class='title'>BURAQ Smart Attendance</div><p class='sub'>Super Admin, Admin এবং HR-এর জন্য একটি নিরাপদ login</p>{msg}<form method='post'><label>Email</label><input type='email' name='email' placeholder='name@buraq.com' autocomplete='username' required><label>Password</label><input type='password' name='password' placeholder='Password' autocomplete='current-password' required><button class='btn' type='submit'>Sign In</button></form></div></div>"""
    return layout("Unified Login", body)

@app.post("/login")
def login(request: Request, password: str = Form(...), email: str = Form(...)):
    normalized_email = email.strip().lower()
    admin_email = get_setting("admin_email", "admin@buraq.com").strip().lower()
    admin_hash = get_setting("admin_password_hash")

    if normalized_email == admin_email and verify_password(password, admin_hash):
        request.session.clear()
        request.session["admin"] = True
        request.session["role"] = "super_admin"
        request.session["user_name"] = get_setting("admin_name", "Super Admin")
        audit(request, "login", "user_account", "super_admin", "super_admin login")
        return RedirectResponse("/dashboard", 303)

    with get_db() as c:
        row = c.execute(
            "SELECT * FROM hr_accounts WHERE LOWER(email)=LOWER(?) AND is_active=?",
            (normalized_email, True),
        ).fetchone()
        if row and verify_password(password, row["password_hash"]):
            request.session.clear()
            request.session["hr_id"] = row["id"]
            request.session["role"] = row["role"]
            request.session["user_name"] = row["name"]
            c.execute("UPDATE hr_accounts SET last_login_at=CURRENT_TIMESTAMP WHERE id=?", (row["id"],))
            # Use the same transaction to avoid SQLite's nested-write lock.
            audit(request, "login", "user_account", str(row["id"]), f"{row['role']} login", db=c)
            return RedirectResponse("/dashboard", 303)

    return RedirectResponse("/login?error=1", 303)

@app.get("/logout")
def logout(request: Request):
    request.session.clear(); return RedirectResponse("/login", 302)

@app.get("/my-account", response_class=HTMLResponse)
def my_account_page(request: Request, saved: str = "", error: str = ""):
    require_login(request)
    notice = ""
    if saved:
        notice = "<div class='notice'>Account information updated successfully.</div>"
    elif error == "password":
        notice = "<div class='notice' style='background:#fee2e2;color:#991b1b'>Current password is incorrect.</div>"
    elif error == "mismatch":
        notice = "<div class='notice' style='background:#fee2e2;color:#991b1b'>New passwords do not match or are shorter than 8 characters.</div>"
    elif error == "email":
        notice = "<div class='notice' style='background:#fee2e2;color:#991b1b'>This email is already being used by another account.</div>"
    elif error:
        notice = "<div class='notice' style='background:#fee2e2;color:#991b1b'>Account could not be updated.</div>"

    if request.session.get("role") == "super_admin" and request.session.get("admin"):
        name = get_setting("admin_name", "Super Admin")
        email = get_setting("admin_email", "admin@buraq.com")
        role_label = "Super Admin"
    else:
        account_id = request.session.get("hr_id")
        with get_db() as c:
            row = c.execute("SELECT name,email,role FROM hr_accounts WHERE id=?", (account_id,)).fetchone()
        if not row:
            request.session.clear()
            return RedirectResponse("/login", 303)
        name, email = row["name"], row["email"]
        role_label = row["role"].replace("_", " ").title()

    body = f"""{notice}
    <div class='hero'><div><div class='eyebrow'>Account</div><h2>My Account</h2></div><span class='pill'>{escape(role_label)}</span></div>
    <div class='two'>
      <div class='card'>
        <h2>Profile Information</h2>
        <form method='post' action='/my-account/profile'>
          <label>Full Name</label><input name='name' value='{escape(str(name))}' required>
          <label>Email Address</label><input type='email' name='email' value='{escape(str(email))}' autocomplete='email' required>
          <label>Current Password</label><input type='password' name='current_password' autocomplete='current-password' required>
          <button class='btn'>Save Profile</button>
        </form>
      </div>
      <div class='card'>
        <h2>Change Password</h2>
        <form method='post' action='/my-account/password'>
          <label>Current Password</label><input type='password' name='current_password' autocomplete='current-password' required>
          <label>New Password</label><input type='password' name='new_password' minlength='8' autocomplete='new-password' required>
          <label>Confirm New Password</label><input type='password' name='confirm_password' minlength='8' autocomplete='new-password' required>
          <button class='btn'>Change Password</button>
        </form>
      </div>
    </div>"""
    return layout("My Account", body, request, "account")

@app.post("/my-account/profile")
def update_my_account_profile(request: Request, name: str = Form(...), email: str = Form(...), current_password: str = Form(...)):
    require_login(request)
    normalized_email = email.strip().lower()
    clean_name = name.strip()
    if not clean_name or not normalized_email:
        return RedirectResponse("/my-account?error=1", 303)

    if request.session.get("role") == "super_admin" and request.session.get("admin"):
        if not verify_password(current_password, admin_setup_hash()):
            return RedirectResponse("/my-account?error=password", 303)
        with get_db() as c:
            duplicate = c.execute("SELECT id FROM hr_accounts WHERE LOWER(email)=LOWER(?)", (normalized_email,)).fetchone()
        if duplicate:
            return RedirectResponse("/my-account?error=email", 303)
        set_setting("admin_name", clean_name)
        set_setting("admin_email", normalized_email)
        request.session["user_name"] = clean_name
        audit(request, "profile_update", "user_account", "super_admin", "Super Admin name/email changed")
    else:
        account_id = request.session.get("hr_id")
        with get_db() as c:
            row = c.execute("SELECT password_hash FROM hr_accounts WHERE id=?", (account_id,)).fetchone()
            if not row or not verify_password(current_password, row["password_hash"]):
                return RedirectResponse("/my-account?error=password", 303)
            admin_email = get_setting("admin_email", "admin@buraq.com").strip().lower()
            duplicate = c.execute("SELECT id FROM hr_accounts WHERE LOWER(email)=LOWER(?) AND id<>?", (normalized_email, account_id)).fetchone()
            if normalized_email == admin_email or duplicate:
                return RedirectResponse("/my-account?error=email", 303)
            c.execute("UPDATE hr_accounts SET name=?,email=?,updated_at=CURRENT_TIMESTAMP WHERE id=?", (clean_name, normalized_email, account_id))
            audit(request, "profile_update", "user_account", str(account_id), "HR/Admin name/email changed", db=c)
        request.session["user_name"] = clean_name
    return RedirectResponse("/my-account?saved=profile", 303)

@app.post("/my-account/password")
def update_my_account_password(request: Request, current_password: str = Form(...), new_password: str = Form(...), confirm_password: str = Form(...)):
    require_login(request)
    if len(new_password) < 8 or new_password != confirm_password:
        return RedirectResponse("/my-account?error=mismatch", 303)

    if request.session.get("role") == "super_admin" and request.session.get("admin"):
        if not verify_password(current_password, admin_setup_hash()):
            return RedirectResponse("/my-account?error=password", 303)
        set_setting("admin_password_hash", hash_password(new_password))
        audit(request, "password_change", "user_account", "super_admin", "Super Admin password changed")
    else:
        account_id = request.session.get("hr_id")
        with get_db() as c:
            row = c.execute("SELECT password_hash FROM hr_accounts WHERE id=?", (account_id,)).fetchone()
            if not row or not verify_password(current_password, row["password_hash"]):
                return RedirectResponse("/my-account?error=password", 303)
            c.execute("UPDATE hr_accounts SET password_hash=?,updated_at=CURRENT_TIMESTAMP WHERE id=?", (hash_password(new_password), account_id))
            audit(request, "password_change", "user_account", str(account_id), "HR/Admin password changed", db=c)
    return RedirectResponse("/my-account?saved=password", 303)

# --- small helpers: keep the route body readable and kill divide-by-zero ---

def _pct(part, whole) -> float:
    """Percentage clamped to 0-100. Returns 0 when the denominator is 0."""
    if not whole:
        return 0.0
    return round(min(100.0, max(0.0, float(part) / float(whole) * 100)), 1)


def _hours_minutes(minutes) -> str:
    """480 -> '8h 00m'.  45 -> '45m'.  Never shows raw minute counts > 59."""
    total = max(0, int(minutes or 0))
    if total < 60:
        return f"{total}m"
    return f"{total // 60}h {total % 60:02d}m"


def _kpi_card(icon_name: str, tone: str, label: str, value: str,
              foot: str, pct: float | None = None, href: str = "") -> str:
    """One KPI tile. Every tile has a bar track, so the cards never
    end up different heights when one of them has no percentage."""
    width = 0.0 if pct is None else pct
    empty = " is-empty" if pct is None else ""
    inner = (
        f"<div class='kpi-row'>"
        f"<span class='kpi-symbol kpi-{tone}'>{ui.icon(icon_name)}</span>"
        f"<div class='kpi-text'>"
        f"<div class='metric-label'>{escape(label)}</div>"
        f"<div class='metric'>{escape(value)}</div>"
        f"</div></div>"
        f"<div class='metric-foot'>{foot}</div>"
        f"<div class='mini-line{empty}'><span class='mini-{tone}' style='width:{width}%'></span></div>"
    )
    if href:
        return (f"<a class='card dashboard-kpi dashboard-kpi-link' href='{href}'>"
                f"{inner}<span class='kpi-go' aria-hidden='true'></span></a>")
    return f"<div class='card dashboard-kpi'>{inner}</div>"


@app.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request):
    require_permission(request, "dashboard_view")
    now = datetime.now(ZoneInfo(settings.timezone))
    today = now.date().isoformat()
    month = now.strftime("%Y-%m")
    week_days = [(now.date() - timedelta(days=i)) for i in range(6, -1, -1)]

    with get_db() as c:
        # ------------------------------------------------------------------
        # ONE query for the whole workforce picture.
        #
        # Everything is scoped to active employees and uses the *same*
        # leave-beats-attendance precedence as the Live Attendance table
        # below, so present + on_leave + absent == total, always.
        # ------------------------------------------------------------------
        snap = c.execute("""
            SELECT COUNT(*) total,
                   SUM(CASE WHEN e.registration_status='approved' THEN 1 ELSE 0 END) registered,
                   SUM(CASE WHEN l.employee_id IS NOT NULL THEN 1 ELSE 0 END) on_leave,
                   SUM(CASE WHEN l.employee_id IS NULL AND a.check_in IS NOT NULL THEN 1 ELSE 0 END) present,
                   SUM(CASE WHEN l.employee_id IS NULL AND a.check_in IS NULL THEN 1 ELSE 0 END) absent,
                   SUM(CASE WHEN a.check_out IS NOT NULL THEN 1 ELSE 0 END) checked_out,
                   SUM(CASE WHEN a.late_minutes>0 THEN 1 ELSE 0 END) late,
                   COALESCE(SUM(a.overtime_minutes),0) overtime
            FROM employees e
            LEFT JOIN attendance a ON a.employee_id=e.id AND a.work_date=?
            LEFT JOIN (SELECT DISTINCT employee_id FROM leave_requests
                       WHERE status='approved' AND start_date<=? AND end_date>=?) l
                   ON l.employee_id=e.id
            WHERE e.is_active
        """, (today, today, today)).fetchone()

        employees = int(snap["total"] or 0)
        registered = int(snap["registered"] or 0)
        present = int(snap["present"] or 0)
        absent = int(snap["absent"] or 0)
        on_leave = int(snap["on_leave"] or 0)
        checked_out = int(snap["checked_out"] or 0)
        late = int(snap["late"] or 0)
        overtime = int(snap["overtime"] or 0)

        pending_registration = int(c.execute(
            "SELECT COUNT(*) c FROM pending_registrations WHERE status='pending'").fetchone()["c"] or 0)
        pending_leave = int(c.execute(
            "SELECT COUNT(*) c FROM leave_requests WHERE status='pending'").fetchone()["c"] or 0)
        pending_correction = int(c.execute(
            "SELECT COUNT(*) c FROM attendance_corrections WHERE status='pending'").fetchone()["c"] or 0)
        pending_selfie = int(c.execute(
            "SELECT COUNT(*) c FROM attendance_fingerprints WHERE review_status='pending'").fetchone()["c"] or 0)

        # Real payroll figure instead of the hard-coded 0.
        payroll_pending = int(c.execute("""
            SELECT COUNT(*) c FROM employees e
            WHERE e.is_active AND NOT EXISTS (
                SELECT 1 FROM payroll_records p
                WHERE p.employee_id=e.id AND p.salary_month=?)
        """, (month,)).fetchone()["c"] or 0)

        # 7-day trend, active employees only, matching the KPI scope.
        week_counts = c.execute("""
            SELECT a.work_date, COUNT(*) c
            FROM attendance a JOIN employees e ON e.id=a.employee_id
            WHERE a.work_date>=? AND a.work_date<=? AND a.check_in IS NOT NULL AND e.is_active
            GROUP BY a.work_date
        """, (week_days[0].isoformat(), week_days[-1].isoformat())).fetchall()

        live_rows = c.execute("""
            SELECT e.name, e.staff_id, a.check_in, a.check_out, a.late_minutes,
                   a.attendance_shift,
                   CASE WHEN l.employee_id IS NOT NULL THEN 'leave'
                        WHEN a.check_in IS NOT NULL THEN 'present'
                        ELSE 'absent' END status
            FROM employees e
            LEFT JOIN attendance a ON a.employee_id=e.id AND a.work_date=?
            LEFT JOIN (SELECT DISTINCT employee_id FROM leave_requests
                       WHERE status='approved' AND start_date<=? AND end_date>=?) l
                   ON l.employee_id=e.id
            WHERE e.is_active
            ORDER BY CASE WHEN a.check_in IS NOT NULL THEN 0 ELSE 1 END,
                     a.check_in DESC, e.name
            LIMIT 6
        """, (today, today, today)).fetchall()

    face_today = face_security_summary(now)

    present_pct = _pct(present, employees)
    absent_pct = _pct(absent, employees)
    leave_pct = _pct(on_leave, employees)
    late_pct = _pct(late, employees)
    unregistered = max(employees - registered, 0)

    # --- 7-day bar chart -------------------------------------------------
    by_day = {str(r["work_date"]): int(r["c"]) for r in week_counts}
    weekly = [(d, by_day.get(d.isoformat(), 0)) for d in week_days]
    max_week = max([v for _, v in weekly] + [1])
    week_total = sum(v for _, v in weekly)
    week_avg = round(week_total / 7, 1)
    chart = "".join(
        f"<div class='bar-wrap'>"
        f"<div class='bar{' is-today' if d == now.date() else ''}' "
        f"style='height:{max(4.0, v / max_week * 100):.1f}%' "
        f"title='{d.strftime('%d %b')}: {v} present'>"
        f"<span class='bar-value'>{v}</span></div>"
        f"<div class='bar-label'>{d.strftime('%a')}</div></div>"
        for d, v in weekly
    )

    # --- live attendance rows --------------------------------------------
    rows = []
    for r in live_rows:
        status = str(r["status"])
        cls = {"present": "status-present", "leave": "status-leave"}.get(status, "status-absent")
        late_min = int(r["late_minutes"] or 0)
        note = f" <span class='tag late-tag'>{late_min}m late</span>" if status == "present" and late_min > 0 else ""
        shift_note = ""
        if status == "present":
            shift_note = "Second Shift" if r["attendance_shift"] == "second" else "First Shift"
        rows.append(
            f"<tr><td><div class='kpi-row'>"
            f"<span class='avatar'>{escape(ui.initials_of(r['name']))}</span>"
            f"<span><b>{escape(str(r['name']))}</b>"
            f"<div class='sub'>{escape(str(r['staff_id']))}{' · ' + shift_note if shift_note else ''}</div></span></div></td>"
            f"<td><span class='status-badge {cls}'>{escape(status.title())}</span>{note}</td>"
            f"<td class='num'>{escape(format_time_12h(r['check_in']) or '—')}</td>"
            f"<td class='num'>{escape(format_time_12h(r['check_out']) or '—')}</td></tr>"
        )
    live_table = "".join(rows) or (
        "<tr><td colspan='4'><div class='empty-cell'>No active employees yet. "
        "Add your first employee to start tracking attendance.</div></td></tr>")

    # --- header ----------------------------------------------------------
    name = escape(str(request.session.get("user_name", "Admin")))
    role = escape(str(request.session.get("role", "super_admin")).replace("_", " ").title())
    search_html = ""
    if has_permission(request, "employees_view"):
        search_html = (
            "<form class='dashboard-search-form' action='/employees' method='get' role='search'>"
            "<input class='dashboard-search' type='search' name='q' "
            "placeholder='Search employee by name, ID or phone' aria-label='Search employees'>"
            "<button class='btn secondary' type='submit'>Search</button></form>")

    # --- KPI tiles -------------------------------------------------------
    kpis = [
        _kpi_card("check", "present", "Present today", str(present),
                  f"{present_pct}% of {employees} active staff", present_pct),
        _kpi_card("clock", "late", "Late arrivals", str(late),
                  ("All on time today" if late == 0 else f"{late_pct}% of staff — already counted in Present"),
                  late_pct),
        _kpi_card("x", "absent", "Absent", str(absent),
                  f"{absent_pct}% of {employees} active staff", absent_pct),
        _kpi_card("calendar-minus", "leave", "On approved leave", str(on_leave),
                  f"{leave_pct}% of {employees} active staff", leave_pct),
        _kpi_card("trending-up", "overtime", "Checked out today", str(checked_out),
                  f"{checked_out} of {present} have checked out · Overtime: manual only",
                  _pct(checked_out, present) if present else 0.0),
    ]
    if has_permission(request, "approvals_view"):
        kpis.append(_kpi_card(
            "user", "late", "Selfies awaiting review", str(pending_selfie),
            "Open the review queue" if pending_selfie else "Queue is clear",
            _pct(pending_selfie, max(present, 1)), "/duplicates?review=pending"))
    if has_permission(request, "face_security_view"):
        spoof = int(face_today["spoof"] or 0)
        checks = int(face_today["checks"] or 0)
        kpis.append(_kpi_card(
            "shield", "absent", "Spoof attempts today", str(spoof),
            f"{checks} face checks · {face_today['rejected']} rejected",
            _pct(spoof, max(checks, 1)), "/face-security"))
    kpi_html = "".join(kpis)

    # --- workforce breakdown donut (segments actually add up to 100%) ----
    seg_present = _pct(present, employees)
    seg_leave = _pct(on_leave, employees)
    stop_1 = seg_present
    stop_2 = min(100.0, seg_present + seg_leave)
    legend = [
        ("present", "Present", present, present_pct),
        ("leave", "On leave", on_leave, leave_pct),
        ("absent", "Absent", absent, absent_pct),
    ]
    legend_html = "".join(
        f"<div class='legend-row'><span class='legend-dot legend-{key}'></span>"
        f"<span>{label}</span><b>{value}</b><span class='legend-pct'>{pc}%</span></div>"
        for key, label, value, pc in legend)
    reg_note = (f"{registered} of {employees} face-registered"
                + (f" · {unregistered} still pending enrolment" if unregistered else " · all enrolled"))

    # --- pending work ----------------------------------------------------
    pending_items = []
    if has_permission(request, "approvals_view"):
        pending_items.append(("/pending", "user", "Employee registrations",
                              "New staff waiting for approval", pending_registration))
        pending_items.append(("/duplicates?review=pending", "search", "Selfie review",
                              "Check-in photos flagged for a decision", pending_selfie))
    if has_permission(request, "leave_view"):
        pending_items.append(("/hr-operations", "calendar-minus", "Leave requests",
                              "Submitted and not yet decided", pending_leave))
    if has_permission(request, "attendance_edit"):
        pending_items.append(("/hr-operations", "file-text", "Attendance corrections",
                              "Staff-reported time fixes", pending_correction))
    if has_permission(request, "payroll_view"):
        pending_items.append(("/payroll", "banknote", f"Payroll — {now.strftime('%B %Y')}",
                              "Employees with no payslip prepared yet", payroll_pending))

    pending_total = sum(count for *_rest, count in pending_items)
    pending_html = "".join(
        f"<a class='pending-item{'' if count else ' is-clear'}' href='{url}'>"
        f"<span class='pending-icon'>{ui.icon(ic)}</span>"
        f"<span><b>{escape(title)}</b><div class='sub'>{escape(sub)}</div></span>"
        f"<span class='count-chip{'' if count else ' chip-zero'}'>{count}</span>"
        f"<span class='pending-go' aria-hidden='true'></span></a>"
        for url, ic, title, sub, count in pending_items)
    if not pending_html:
        pending_html = "<div class='empty-cell'>Nothing is waiting on you right now.</div>"

    # --- quick actions ---------------------------------------------------
    quick = []
    if has_permission(request, "employees_add"):
        quick.append(("/employees", "plus", "Add employee"))
    if has_permission(request, "attendance_edit"):
        quick.append(("/attendance", "clock", "Mark attendance"))
    if has_permission(request, "duty_view"):
        quick.append(("/duty", "calendar-check", "Assign duty"))
    if has_permission(request, "leave_view"):
        quick.append(("/hr-operations", "calendar-minus", "Record leave"))
    if has_permission(request, "payroll_view"):
        quick.append(("/payroll", "banknote", "Run payroll"))
    if has_permission(request, "reports_view"):
        quick.append(("/reports", "chart-bar", "View reports"))
    quick_html = "".join(
        f"<a href='{url}'><span class='qicon'>{ui.icon(ic)}</span><span>{escape(label)}</span></a>"
        for url, ic, label in quick)
    quick_block = (f"<div class='section-head'><h3>Quick actions</h3></div>"
                   f"<div class='dashboard-quick'>{quick_html}</div>") if quick_html else ""

    report_btn = ("<a class='btn secondary' href='/reports'>Full report</a>"
                  if has_permission(request, "reports_view") else "")
    all_btn = ("<a class='btn secondary' href='/employees'>All employees</a>"
               if has_permission(request, "employees_view") else "")

    body = f"""
    <div class='dashboard-head'>
      <div class='dashboard-greet'>
        <h1>Good {'morning' if now.hour < 12 else 'afternoon' if now.hour < 17 else 'evening'}, {name}</h1>
        <div class='dashboard-date'>
          <span class='status ok'>{role}</span>
          <span>{now.strftime('%A, %d %B %Y')}</span>
          <span class='live-dot' title='Refreshes automatically'></span>
          <span class='sub' id='dash-updated'>Updated {now.strftime('%I:%M %p')}</span>
        </div>
      </div>
      <div class='dashboard-tools'>{search_html}</div>
    </div>

    <div class='dashboard-kpis'>{kpi_html}</div>

    <div class='section-gap'></div>
    <div class='dashboard-main-grid'>
      <div class='card dashboard-panel'>
        <div class='card-head'>
          <div><h3>Attendance, last 7 days</h3>
          <div class='sub'>{week_total} check-ins · {week_avg} per day on average</div></div>
          {report_btn}
        </div>
        <div class='chart'>{chart}</div>
      </div>
      <div class='card dashboard-panel'>
        <div class='card-head'><div><h3>Where the team is today</h3>
        <div class='sub'>{reg_note}</div></div></div>
        <div class='readiness-wrap'>
          <div class='donut donut-stack' style='--stop1:{stop_1};--stop2:{stop_2}'
               role='img' aria-label='{present} present, {on_leave} on leave, {absent} absent'>
            <div class='donut-value'><b>{round(present_pct)}%</b><span class='sub'>Present</span></div>
          </div>
          <div class='legend-list'>{legend_html}</div>
        </div>
      </div>
    </div>

    <div class='section-gap'></div>
    <div class='dashboard-main-grid'>
      <div class='card'>
        <div class='card-head'><div><h3>Live attendance</h3>
        <div class='sub'>Latest check-ins first</div></div>{all_btn}</div>
        <div class='table-scroll'>
          <table class='dashboard-table'>
            <thead><tr><th>Employee</th><th>Status</th><th class='num'>Check in</th><th class='num'>Check out</th></tr></thead>
            <tbody>{live_table}</tbody>
          </table>
        </div>
      </div>
      <div class='card'>
        <div class='card-head'><div><h3>Needs your attention</h3>
        <div class='sub'>Items waiting for a decision</div></div>
        <span class='pill{'' if pending_total else ' chip-zero'}'>{pending_total}</span></div>
        <div class='pending-list'>{pending_html}</div>
      </div>
    </div>

    <div class='section-gap'></div>
    {quick_block}

    <script>
    (function () {{
      // Refresh every 60s, but only while the tab is visible and the user
      // is not typing in the search box. Prevents the page reloading under
      // someone's hands, which the old dashboard never handled at all.
      var WAIT = 60000, timer = null;
      function armed() {{
        var a = document.activeElement;
        return document.visibilityState === 'visible' &&
               !(a && (a.tagName === 'INPUT' || a.tagName === 'TEXTAREA' || a.tagName === 'SELECT'));
      }}
      function tick() {{ if (armed()) {{ location.reload(); }} else {{ schedule(); }} }}
      function schedule() {{ clearTimeout(timer); timer = setTimeout(tick, WAIT); }}
      document.addEventListener('visibilitychange', schedule);
      schedule();
    }})();
    </script>
    """
    return layout("Dashboard", body, request, "dashboard")


@app.get("/attendance", response_class=HTMLResponse)
def attendance_center(request: Request):
    require_login(request)
    cards=[]
    if has_permission(request,"reports_view"): cards.append(("📊","Attendance Reports","Daily records, late, overtime and employee attendance history.","/reports"))
    if has_permission(request,"duty_view"): cards.append(("🗓","Duty Schedule","Regular, custom, Friday and night duty with reminder status.","/duty-schedules"))
    if has_permission(request,"attendance_edit"): cards.append(("✅","Missing Duty Days","Scheduled duty with no attendance at all. Record a past day directly.","/attendance/missing"))
    if has_permission(request,"leave_view"): cards.append(("🏖","Leave & Corrections","Leave approval, attendance correction, shifts and departments.","/hr-operations"))
    if has_permission(request,"reports_export"): cards.append(("📥","Reports & Export","Download filtered attendance as Excel, PDF or CSV.","/reports"))
    if not cards: raise HTTPException(403,"Permission denied")
    content=''.join(f"<a class='card control-card' href='{url}'><div class='control-icon'>{icon}</div><h3>{title}</h3><div class='sub'>{description}</div></a>" for icon,title,description,url in cards)
    body=f"<div class='hero'><div><div class='eyebrow'>One Simple Workspace</div><h2>Attendance Center</h2><div class='sub'>Attendance, duty, leave, corrections and exports are organized here.</div></div></div><div class='control-grid'>{content}</div>"
    return layout("Attendance",body,request,"attendance")

@app.get("/admin", response_class=HTMLResponse)
def admin_center(request: Request):
    require_login(request)
    cards=[]
    if has_permission(request,"approvals_view"):
        cards.extend([("✅","All Approvals","Registration, leave, correction and duplicate review in one place.","/approvals"),("🔎","Duplicate Review","Open duplicate attendance evidence directly.","/duplicates")])
    if has_permission(request,"face_security_view"): cards.append(("◉","Face Security","Monitor verification decisions, liveness failures and spoof alerts.","/face-security"))
    if has_permission(request,"user_accounts_view"): cards.append(("👤","Users & Permissions","Manage HR accounts, roles and access permissions.","/hr-accounts"))
    if has_permission(request,"audit_view"): cards.append(("🧾","Activity Logs","See who changed attendance, payroll or system data.","/audit-logs"))
    if has_permission(request,"settings_view"): cards.append(("⚙️","Settings & Backup","WhatsApp connection, webhook, password and backups.","/settings"))
    if has_permission(request,"shift_manage") or has_permission(request,"department_manage"): cards.append(("🏢","Office Setup","Manage shifts and departments from HR Operations.","/hr-operations"))
    if not cards: raise HTTPException(403,"Permission denied")
    content=''.join(f"<a class='card control-card' href='{url}'><div class='control-icon'>{icon}</div><h3>{title}</h3><div class='sub'>{description}</div></a>" for icon,title,description,url in cards)
    body=f"<div class='hero'><div><div class='eyebrow'>Restricted Control</div><h2>Admin Center</h2><div class='sub'>Approvals, security, accounts, logs and settings in one place.</div></div></div><div class='control-grid'>{content}</div>"
    return layout("Admin",body,request,"admin")

@app.get("/approvals", response_class=HTMLResponse)
def approvals_center(request: Request):
    require_permission(request,"approvals_view")
    cards=[("👤","Registration","Approve or reject new employee WhatsApp registrations.","/pending"),("🔎","Duplicate Attendance","Review duplicate evidence and Accept/Pending/Reject decisions.","/duplicates")]
    if has_permission(request,"leave_view"): cards.append(("🏖","Leave & Corrections","Review leave and attendance correction requests.","/hr-operations"))
    content=''.join(f"<a class='card control-card' href='{url}'><div class='control-icon'>{icon}</div><h3>{title}</h3><div class='sub'>{description}</div></a>" for icon,title,description,url in cards)
    body=f"<div class='hero'><div><div class='eyebrow'>Review Queue</div><h2>All Approvals</h2><div class='sub'>Choose the approval type instead of searching separate menus.</div></div></div><div class='control-grid'>{content}</div>"
    return layout("Approvals",body,request,"admin")


def face_security_summary(now: datetime | None = None) -> dict[str, int]:
    """Return today's Face AI totals without ever breaking the main dashboard."""
    local_now = now or datetime.now(ZoneInfo(settings.timezone))
    start = local_now.date().isoformat()
    end = (local_now.date() + timedelta(days=1)).isoformat()
    try:
        with get_db() as c:
            row = c.execute(
                """SELECT COUNT(*) checks,
                    SUM(CASE WHEN decision='accepted' THEN 1 ELSE 0 END) accepted,
                    SUM(CASE WHEN decision='rejected' THEN 1 ELSE 0 END) rejected,
                    SUM(CASE WHEN LOWER(COALESCE(reason,'')) LIKE '%liveness%'
                              OR LOWER(COALESCE(liveness_verdict,'')) IN ('spoof','failed','fail')
                             THEN 1 ELSE 0 END) spoof
                   FROM face_events WHERE created_at>=? AND created_at<?""",
                (start, end),
            ).fetchone()
        return {key: int(row[key] or 0) for key in ("checks", "accepted", "rejected", "spoof")}
    except Exception:
        logger.warning("face security summary failed", exc_info=True)
        return {"checks": 0, "accepted": 0, "rejected": 0, "spoof": 0}


@app.get("/face-security", response_class=HTMLResponse)
def face_security_page(request: Request, view: str = "all"):
    require_permission(request, "face_security_view")
    if view not in {"all", "rejected", "spoof"}:
        view = "all"
    now = datetime.now(ZoneInfo(settings.timezone))
    totals = face_security_summary(now)
    since = (now.date() - timedelta(days=6)).isoformat()
    clauses = ["f.created_at>=?"]
    params: list = [since]
    if view == "rejected":
        clauses.append("f.decision='rejected'")
    elif view == "spoof":
        clauses.append("(LOWER(COALESCE(f.reason,'')) LIKE '%liveness%' OR LOWER(COALESCE(f.liveness_verdict,'')) IN ('spoof','failed','fail'))")
    where = " AND ".join(clauses)
    try:
        with get_db() as c:
            events = c.execute(
                f"""SELECT f.*,e.name employee_name,e.staff_id
                      FROM face_events f LEFT JOIN employees e ON e.id=f.employee_id
                     WHERE {where} ORDER BY f.created_at DESC LIMIT 200""",
                params,
            ).fetchall()
    except Exception:
        logger.warning("face security events failed", exc_info=True)
        events = []

    rows = ""
    for event in events:
        decision = str(event["decision"] or "unknown").lower()
        badge = "ok" if decision == "accepted" else ("bad" if decision == "rejected" else "warn")
        employee = escape(str(event["employee_name"] or "Unknown employee"))
        staff_id = escape(str(event["staff_id"] or "—"))
        created = escape(str(event["created_at"] or "—").replace("T", " ")[:19])
        stage = escape(str(event["stage"] or "—").replace("_", " ").title())
        action = escape(str(event["action"] or "—").replace("_", " ").title())
        reason = escape(str(event["reason"] or "—"))
        liveness = escape(str(event["liveness_verdict"] or "—").replace("_", " ").title())
        rows += f"""<tr><td><b>{created}</b></td><td><b>{employee}</b><div class='sub'>{staff_id}</div></td>
        <td>{stage}<div class='sub'>{action}</div></td><td><span class='status {badge}'>{escape(decision.title())}</span></td>
        <td>{float(event['match_score'] or 0):.3f}<div class='sub'>margin {float(event['margin'] or 0):.3f}</div></td>
        <td>{float(event['quality'] or 0):.1f}</td><td>{liveness}<div class='sub'>{float(event['liveness_score'] or 0):.2f}</div></td>
        <td>{reason}</td><td>{float(event['elapsed_ms'] or 0):.0f} ms</td></tr>"""

    def tab(label: str, key: str, count: int | None = None) -> str:
        active = " active" if view == key else ""
        suffix = f" <span class='pill'>{count}</span>" if count is not None else ""
        return f"<a class='btn secondary{active}' href='/face-security?view={key}'>{label}{suffix}</a>"

    tabs = tab("All checks", "all", totals["checks"]) + tab("Rejected", "rejected", totals["rejected"]) + tab("Spoof alerts", "spoof", totals["spoof"])
    review_link = "<a class='btn secondary' href='/duplicates?review=pending'>Open Selfie Review</a>" if has_permission(request, "approvals_view") else ""
    body = f"""
    <div class='hero'><div><div class='eyebrow'>Face AI Monitoring</div><h2>Face Security</h2>
      <div class='sub'>Verification decisions, liveness failures and suspected spoof attempts in one place.</div></div>{review_link}</div>
    <div class='dashboard-kpis face-security-kpis'>
      <div class='card dashboard-kpi'><div class='metric-label'>Face Checks Today</div><div class='metric'>{totals['checks']}</div><div class='metric-foot'>All enrollment and attendance checks</div></div>
      <div class='card dashboard-kpi'><div class='metric-label'>Verified Today</div><div class='metric'>{totals['accepted']}</div><div class='metric-foot'>Face verification accepted</div></div>
      <div class='card dashboard-kpi'><div class='metric-label'>Rejected Today</div><div class='metric'>{totals['rejected']}</div><div class='metric-foot'>Needs attention or retry</div></div>
      <div class='card dashboard-kpi'><div class='metric-label'>Spoof Alerts Today</div><div class='metric'>{totals['spoof']}</div><div class='metric-foot'>Liveness-related failures</div></div>
    </div>
    <div class='card section-gap'><div class='card-head'><div><h3>Security Event Log</h3><div class='sub'>Latest 200 events from the last 7 days</div></div><div class='actions'>{tabs}</div></div>
      <div style='overflow:auto'><table class='dashboard-table'><thead><tr><th>Time</th><th>Employee</th><th>Stage / Action</th><th>Decision</th><th>Match</th><th>Quality</th><th>Liveness</th><th>Reason</th><th>Speed</th></tr></thead>
      <tbody>{rows or '<tr><td colspan="9"><div class="empty"><h3>No face security events</h3><div class="sub">New face checks will appear here automatically.</div></div></td></tr>'}</tbody></table></div></div>
    """
    return layout("Face Security", body, request, "face-security")

@app.post("/test-message")
async def test_message(request: Request, phone: str = Form(...), message: str = Form(...)):
    require_permission(request, "whatsapp_settings")
    result = await send_text(phone.strip(), message.strip())
    return RedirectResponse("/dashboard" if result.get("sent") else "/settings?error=send", 303)

def mask_secret(value: str, visible: int = 4) -> str:
    if not value: return "Not configured"
    if len(value) <= visible * 2: return "•" * len(value)
    return value[:visible] + "•" * 12 + value[-visible:]

@app.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request, saved: str = "", error: str = ""):
    require_permission(request, "settings_view")
    access=get_setting("whatsapp_access_token"); phone=get_setting("whatsapp_phone_number_id"); verify=get_setting("whatsapp_verify_token")
    notice = "<div class='notice'>Settings সফলভাবে save হয়েছে।</div>" if saved else ("<div class='notice' style='background:#fee2e2;color:#991b1b'>Action ব্যর্থ হয়েছে। তথ্য পরীক্ষা করুন।</div>" if error else "")
    webhook=f"{base_url(request)}/webhook/whatsapp"
    if has_permission(request, "whatsapp_settings"):
        body=f"{notice}<div class='two'><div class='card'><h2>WhatsApp Connection</h2><p><span class='status {'ok' if configured() else 'warn'}'>{'Connected' if configured() else 'Setup needed'}</span></p><div class='sub'>Access Token</div><div class='masked'>{escape(mask_secret(access))}</div><br><div class='sub'>Phone Number ID</div><div class='masked'>{escape(mask_secret(phone))}</div><br><div class='sub'>Verify Token</div><div class='masked'>{escape(mask_secret(verify))}</div><br><details><summary class='btn secondary'>Edit credentials</summary><form method='post' style='margin-top:15px'><label>New Access Token</label><input type='password' name='access_token'><label>New Phone Number ID</label><input name='phone_id'><label>New Verify Token</label><input type='password' name='verify_token'><button class='btn'>Save securely</button></form></details></div><div class='card'><h2>Connection & Webhook</h2><div class='sub'>Callback URL</div><div class='code'>{escape(webhook)}</div><p class='sub'>Sensitive access is controlled by Super Admin permission.</p></div></div>"
    else:
        body=f"{notice}<div class='card'><h2>General Settings</h2><div class='notice'>আপনার account-এ WhatsApp Settings permission নেই। Token, Phone Number ID, Verify Token এবং Webhook URL গোপন রাখা হয়েছে।</div></div>"
    if request.session.get("role") == "super_admin":
        recovery=backup_status(); offsite=recovery["offsite_configured"]
        latest=escape(recovery.get("latest_file") or "No backup yet")
        local_success=escape(recovery.get("last_local_success") or "Waiting for first backup")
        remote_success=escape(recovery.get("last_offsite_success") or ("Waiting for first upload" if offsite else "Not configured"))
        recovery_error=escape(recovery.get("last_error") or "None")
        admin_email=escape(get_setting("admin_email","admin@buraq.com"))
        body += f"""<div class='card' style='margin-top:18px'><h2>Admin Login Settings</h2>
        <p class='sub'>Initial Setup আবার করতে হবে না। এখান থেকে email ও password পরিবর্তন করুন।</p>
        <form method='post' action='/settings/password'><label>Current password</label><input type='password' name='current_password' required autocomplete='current-password'>
        <label>Admin email</label><input type='email' name='new_email' value='{admin_email}' required autocomplete='email'>
        <label>New password</label><input type='password' name='new_password' minlength='8' required autocomplete='new-password'>
        <label>Confirm new password</label><input type='password' name='confirm_password' minlength='8' required autocomplete='new-password'>
        <button class='btn'>Update Password</button></form></div>""" + f"""<div class='card' style='margin-top:18px'><h2>Disaster Recovery</h2>
        <p><span class='status {'ok' if recovery.get('verified') else 'warn'}'>{'Latest backup verified' if recovery.get('verified') else 'Verification pending'}</span>
        <span class='status {'ok' if offsite else 'warn'}'>{'Off-site active' if offsite else 'Local only'}</span>
        <span class='status {'ok' if recovery.get('encrypted') else 'bad'}'>{'Encrypted' if recovery.get('encrypted') else 'Encryption missing'}</span></p>
        <div class='two'><div><div class='sub'>Latest local backup</div><b>{latest}</b><p class='sub'>{local_success} · {recovery.get('local_count',0)} retained</p></div>
        <div><div class='sub'>Latest off-site copy</div><b>{remote_success}</b><p class='sub'>Last error: {recovery_error}</p></div></div>
        <p class='sub'>Full backup-এ employee, face embedding, attendance, duty, payroll, approval, user, settings ও audit history থাকে। প্রতিদিন automatic backup হয়।</p>
        <div class='table-actions'><a class='btn' href='/settings/full-backup'>Download Full Backup</a>
        <form method='post' action='/settings/full-backup/offsite'><button class='btn secondary'>Backup Now</button></form></div>
        <hr style='border:0;border-top:1px solid var(--line);margin:20px 0'>
        <details><summary class='btn secondary'>Verify a backup</summary><form method='post' action='/settings/full-backup/inspect' enctype='multipart/form-data' style='margin-top:14px'>
        <input type='file' name='backup_file' accept='.buraq,.gz' required><button class='btn secondary'>Check Without Restoring</button></form></details>
        <details><summary class='btn danger'>Restore on this server</summary>
        <div class='notice' style='background:#fee2e2;color:#991b1b;margin-top:14px'>Restore বর্তমান database replace করবে। Restore-এর আগে automatic safety backup রাখা হবে।</div>
        <form method='post' action='/settings/full-restore' enctype='multipart/form-data'>
        <label>BURAQ encrypted full backup (.buraq)</label><input type='file' name='backup_file' accept='.buraq,.gz' required>
        <label>Confirmation</label><input name='confirmation' placeholder='RESTORE BURAQ' required>
        <button class='btn danger'>Restore Full Database</button></form></details></div>"""
    return layout("Settings", body, request, "settings")

@app.post("/settings")
def save_settings(request: Request, access_token: str = Form(""), phone_id: str = Form(""), verify_token: str = Form("")):
    require_permission(request, "whatsapp_settings")
    if access_token.strip(): set_setting("whatsapp_access_token", access_token.strip())
    if phone_id.strip(): set_setting("whatsapp_phone_number_id", phone_id.strip())
    if verify_token.strip(): set_setting("whatsapp_verify_token", verify_token.strip())
    return RedirectResponse("/settings?saved=1", 303)

@app.post("/settings/password")
def change_password(request: Request, current_password: str = Form(...), new_email: str = Form(...), new_password: str = Form(...), confirm_password: str = Form(...)):
    require_super_admin(request)
    if not verify_password(current_password, admin_setup_hash()) or len(new_password) < 8 or new_password != confirm_password:
        return RedirectResponse("/settings?error=password", 303)
    set_setting("admin_password_hash", hash_password(new_password))
    set_setting("admin_email",new_email.strip().lower())
    audit(request,"login_settings_changed","user_account","super_admin","Admin email/password changed")
    return RedirectResponse("/settings?saved=password", 303)

@app.get("/settings/backup")
def backup_settings(request: Request):
    require_super_admin(request)
    payload={"version":1,"encrypted":True,"created_at":datetime.now(ZoneInfo(settings.timezone)).isoformat(),"settings":{key:get_stored_setting(key) for key in ("whatsapp_access_token","whatsapp_phone_number_id","whatsapp_verify_token")},"plain_settings":{"meta_api_version":get_setting("meta_api_version","v23.0")}}
    data=json.dumps(payload,ensure_ascii=False,indent=2).encode("utf-8")
    return StreamingResponse(io.BytesIO(data),media_type="application/json",headers={"Content-Disposition":"attachment; filename=BURAQ-Config-Backup.json"})

@app.post("/settings/restore")
async def restore_settings(request: Request):
    require_super_admin(request)
    form=await request.form(); upload=form.get("backup_file")
    try:
        data=json.loads((await upload.read()).decode("utf-8"))
        values=data.get("settings",{})
        if data.get("encrypted"):
            for key in ("whatsapp_access_token","whatsapp_phone_number_id","whatsapp_verify_token"):
                if values.get(key): restore_stored_setting(key,str(values[key]))
            plain=data.get("plain_settings",{})
            if plain.get("meta_api_version"): set_setting("meta_api_version",str(plain["meta_api_version"]))
        else:
            for key in ("whatsapp_access_token","whatsapp_phone_number_id","whatsapp_verify_token","meta_api_version"):
                if values.get(key): set_setting(key,str(values[key]))
    except Exception:
        return RedirectResponse("/settings?error=restore",303)
    return RedirectResponse("/settings?saved=restore",303)

@app.get("/settings/payroll-backup")
def payroll_backup(request: Request):
    require_super_admin(request)
    with get_db() as c:
        payload={"version":2,"type":"buraq_payroll_backup","created_at":datetime.now(ZoneInfo(settings.timezone)).isoformat(),"employee_salary_master":[dict(r) for r in c.execute("SELECT id,staff_id,name,fixed_salary,default_overtime_rate FROM employees ORDER BY id").fetchall()],"payroll_records":[dict(r) for r in c.execute("SELECT * FROM payroll_records ORDER BY salary_month,id").fetchall()],"payroll_change_logs":[dict(r) for r in c.execute("SELECT * FROM payroll_change_logs ORDER BY id").fetchall()]}
    data=json.dumps(payload,ensure_ascii=False,indent=2,default=str).encode("utf-8")
    stamp=datetime.now(ZoneInfo(settings.timezone)).strftime("%Y%m%d-%H%M%S")
    return StreamingResponse(io.BytesIO(data),media_type="application/json",headers={"Content-Disposition":f"attachment; filename=BURAQ-Payroll-Backup-{stamp}.json"})

@app.get("/settings/full-backup")
def full_backup_download(request: Request):
    require_super_admin(request)
    path=create_full_backup()
    audit(request,"full_backup_downloaded","system","database",f"file={path.name}")
    data=path.read_bytes()
    return StreamingResponse(io.BytesIO(data),media_type="application/octet-stream",headers={"Content-Disposition":f"attachment; filename={path.name}","Cache-Control":"no-store"})

@app.post("/settings/full-backup/offsite")
def full_backup_offsite(request: Request):
    require_super_admin(request)
    try:
        path=create_full_backup(); uploaded=upload_offsite(path)
        audit(request,"full_backup_created","system","database",f"file={path.name}; offsite={uploaded}")
        return RedirectResponse("/settings?saved=backup" if uploaded else "/settings?saved=backup-local",303)
    except Exception:
        logger.exception("Manual full backup failed")
        return RedirectResponse("/settings?error=backup",303)

@app.post("/settings/full-backup/inspect", response_class=HTMLResponse)
async def full_backup_inspect(request: Request):
    require_super_admin(request)
    form=await request.form(); upload=form.get("backup_file")
    temporary=Path(tempfile.gettempdir())/f"buraq-inspect-{uuid.uuid4().hex}.buraq"
    try:
        content=await upload.read()
        if len(content) > 250 * 1024 * 1024: raise ValueError("Backup is too large")
        temporary.write_bytes(content); info=inspect_backup(temporary)
        body=f"""<div class='card'><h2>Backup Verification Passed</h2><p><span class='status ok'>Valid & readable</span></p>
        <div class='two'><div><div class='sub'>Created</div><b>{escape(str(info['created_at']))}</b><br><div class='sub'>Source</div><b>{escape(str(info['source_database']))}</b></div>
        <div><div class='sub'>App version</div><b>{escape(str(info['app_version']))}</b><br><div class='sub'>Contents</div><b>{info['tables']} tables · {info['rows']} rows</b></div></div>
        <p class='sub'>কোনো data restore বা পরিবর্তন করা হয়নি।</p><a class='btn' href='/settings'>Back to Settings</a></div>"""
        return layout("Backup Verification",body,request,"settings")
    except Exception as exc:
        logger.warning("Backup inspection failed: %s",exc)
        body=f"<div class='card'><h2>Backup Verification Failed</h2><div class='notice' style='background:#fee2e2;color:#991b1b'>{escape(str(exc))}</div><a class='btn' href='/settings'>Back to Settings</a></div>"
        return layout("Backup Verification",body,request,"settings")
    finally:
        temporary.unlink(missing_ok=True)

@app.post("/settings/full-restore")
async def full_backup_restore(request: Request):
    require_super_admin(request)
    form=await request.form(); upload=form.get("backup_file"); confirmation=str(form.get("confirmation", "")).strip()
    if confirmation != "RESTORE BURAQ" or not upload:
        return RedirectResponse("/settings?error=restore-confirmation",303)
    temporary=Path(tempfile.gettempdir())/f"buraq-restore-{uuid.uuid4().hex}.buraq"
    try:
        content=await upload.read()
        if len(content) > 250 * 1024 * 1024: raise ValueError("Backup is too large")
        temporary.write_bytes(content)
        read_backup(temporary)
        result=restore_full_backup(temporary)
        logger.warning("Full database restored created_at=%s safety=%s",result["created_at"],result["safety_backup"])
    except Exception:
        logger.exception("Full restore failed")
        return RedirectResponse("/settings?error=full-restore",303)
    finally:
        temporary.unlink(missing_ok=True)
    return RedirectResponse("/settings?saved=full-restore",303)

@app.get("/employees", response_class=HTMLResponse)
def employees_page(request: Request, q: str = "", department: str = "", shift: str = "", status: str = ""):
    require_permission(request, "employees_view")
    clauses=["1=1"]; params=[]
    if q.strip():
        clauses.append("(LOWER(e.staff_id) LIKE ? OR LOWER(e.name) LIKE ? OR LOWER(COALESCE(e.phone,'')) LIKE ? OR LOWER(COALESCE(e.designation,'')) LIKE ?)")
        term=f"%{q.strip().lower()}%"; params += [term,term,term,term]
    if department: clauses.append("e.department=?"); params.append(department)
    if shift: clauses.append("e.shift=?"); params.append(shift)
    if status in {"active","inactive"}: clauses.append("e.is_active=?"); params.append(1 if status=="active" else 0)
    with get_db() as c:
        rows=c.execute("SELECT e.*,(SELECT COUNT(*) FROM face_samples f WHERE f.employee_id=e.id) face_count,(SELECT MAX(work_date) FROM attendance a WHERE a.employee_id=e.id) last_attendance FROM employees e WHERE "+" AND ".join(clauses)+" ORDER BY e.staff_id", tuple(params)).fetchall()
        deps=c.execute("SELECT DISTINCT department FROM employees WHERE COALESCE(department,'')<>'' ORDER BY department").fetchall()
    can_add=has_permission(request,"employees_add"); can_edit=has_permission(request,"employees_edit"); can_reset=has_permission(request,"face_reset")
    tr=[]
    for r in rows:
        initials=''.join(x[:1] for x in (r['name'] or '?').split()[:2]).upper()
        reg='ok' if r['registration_status']=='approved' else 'warn'; face='ok' if r['face_count']>=3 else 'bad'
        actions=f"<a class='btn secondary' href='/employees/{r['id']}'>Profile</a>"
        if has_permission(request,'duty_view'): actions += f"<a class='btn secondary' href='/employees/{r['id']}/duty'>Duty</a>"
        if can_reset: actions += f"<a class='btn danger' href='/employees/{r['id']}/reset'>Reset</a>"
        tr.append(f"<tr><td><input class='checkbox' type='checkbox' name='employee_ids' value='{r['id']}'></td><td><div style='display:flex;gap:9px;align-items:center'><span class='avatar'>{escape(initials)}</span><span><b>{escape(r['name'])}</b><br><span class='sub'>{escape(r['staff_id'])}</span></span></div></td><td>{escape(r['designation'] or '—')}</td><td>{escape(r['department'] or '—')}</td><td>{escape(r['shift'])}</td><td><span class='status {reg}'>{escape(r['registration_status'])}</span><br><span class='status {face}' style='margin-top:5px'>{r['face_count']}/3 Face</span></td><td>{escape(r['last_attendance'] or 'Never')}</td><td><div class='table-actions'>{actions}</div></td></tr>")
    depopts=''.join(f"<option {'selected' if department==d['department'] else ''}>{escape(d['department'])}</option>" for d in deps)
    add=''
    if can_add:
        add="""<details class='card'><summary style='cursor:pointer;font-weight:850'>＋ Add Employee</summary><form method='post' style='margin-top:16px'><div class='two'><div><label>Staff ID</label><input name='staff_id' required><label>Name</label><input name='name' required><label>Phone</label><input name='phone'></div><div><label>Department</label><input name='department'><label>Designation</label><input name='designation'><label>Shift</label><select name='shift'><option value='morning'>Morning</option><option value='evening'>Evening</option></select></div></div><button class='btn'>Add Employee</button></form></details><div class='section-gap'></div>"""
    bulk=''
    if can_edit:
        bulk="""<div class='card' style='margin-bottom:15px'><b>Bulk Actions</b><div class='actions' style='margin-top:10px'><select name='bulk_action' style='width:auto;margin:0'><option value=''>Choose action</option><option value='activate'>Activate</option><option value='deactivate'>Deactivate</option><option value='shift'>Change shift</option><option value='department'>Change department</option></select><input name='bulk_value' placeholder='Shift/department value' style='width:220px;margin:0'><button class='btn'>Apply to selected</button></div></div>"""
    body=f"""<div class='hero'><div><div class='eyebrow'>Employee Center</div><h2>Workforce Directory</h2><div class='sub'>360° profiles, advanced search and bulk workforce operations.</div></div><span class='pill'>{len(rows)} results</span></div>{add}<div class='card' style='margin-bottom:15px'><form method='get' class='searchbar'><div><label>Global search</label><input name='q' value='{escape(q)}' placeholder='Name, Staff ID, phone or designation'></div><div><label>Department</label><select name='department'><option value=''>All</option>{depopts}</select></div><div><label>Shift</label><select name='shift'><option value=''>All</option><option {'selected' if shift=='morning' else ''} value='morning'>Morning</option><option {'selected' if shift=='evening' else ''} value='evening'>Evening</option></select></div><div><label>Status</label><select name='status'><option value=''>All</option><option {'selected' if status=='active' else ''} value='active'>Active</option><option {'selected' if status=='inactive' else ''} value='inactive'>Inactive</option></select></div><button class='btn'>Search</button></form></div><form method='post' action='/employees/bulk'>{bulk}<div class='card'><div style='overflow:auto'><table><thead><tr><th></th><th>Employee</th><th>Designation</th><th>Department</th><th>Shift</th><th>Readiness</th><th>Last attendance</th><th>Action</th></tr></thead><tbody>{''.join(tr) or '<tr><td colspan=8>No employees found</td></tr>'}</tbody></table></div></div></form>"""
    return layout("Employee Center", body, request, "employees")

@app.post("/employees")
def add_employee(request: Request, staff_id: str = Form(...), name: str = Form(...), phone: str = Form(""), department: str = Form(""), designation: str = Form(""), shift: str = Form("morning")):
    require_permission(request, "employees_add")
    try:
        with get_db() as c:
            c.execute("INSERT INTO employees(staff_id,name,phone,department,designation,shift) VALUES(?,?,?,?,?,?)", (staff_id.strip(),name.strip(),phone.strip() or None,department.strip() or None,designation.strip() or None,shift))
            audit(request,"employee_created","employee",staff_id,db=c)
    except Exception as exc: logger.warning("Employee add failed: %s",exc)
    return RedirectResponse("/employees",303)

@app.post("/employees/bulk")
async def employee_bulk(request: Request):
    require_permission(request,"employees_edit")
    form=await request.form(); ids=[int(x) for x in form.getlist('employee_ids') if str(x).isdigit()]
    action=str(form.get('bulk_action','')); value=str(form.get('bulk_value','')).strip()
    if not ids or action not in {'activate','deactivate','shift','department'}: return RedirectResponse('/employees',303)
    marks=','.join('?' for _ in ids)
    with get_db() as c:
        if action in {'activate','deactivate'}: c.execute(f"UPDATE employees SET is_active=?,updated_at=CURRENT_TIMESTAMP WHERE id IN ({marks})", tuple([1 if action=='activate' else 0]+ids))
        elif action=='shift' and value: c.execute(f"UPDATE employees SET shift=?,updated_at=CURRENT_TIMESTAMP WHERE id IN ({marks})", tuple([value]+ids))
        elif action=='department' and value: c.execute(f"UPDATE employees SET department=?,updated_at=CURRENT_TIMESTAMP WHERE id IN ({marks})", tuple([value]+ids))
        audit(request,"employee_bulk_update","employee",','.join(map(str,ids)),f"{action}:{value}",db=c)
    return RedirectResponse('/employees',303)

@app.get("/employees/{employee_id}", response_class=HTMLResponse)
def employee_profile(request: Request, employee_id: int, month: str = ""):
    require_permission(request,"employees_view")
    can_payroll_view=has_permission(request,"payroll_view")
    can_payroll_manage=has_permission(request,"payroll_manage")
    can_payroll_export=has_permission(request,"payroll_export")
    now=datetime.now(ZoneInfo(settings.timezone)); month=month or now.strftime('%Y-%m')
    try: first=datetime.strptime(month+'-01','%Y-%m-%d')
    except ValueError: first=datetime(now.year,now.month,1); month=first.strftime('%Y-%m')
    next_month=(first.replace(day=28)+timedelta(days=4)).replace(day=1); last=next_month-timedelta(days=1)
    with get_db() as c:
        e=c.execute("SELECT e.*,(SELECT COUNT(*) FROM face_samples WHERE employee_id=e.id) face_count FROM employees e WHERE e.id=?",(employee_id,)).fetchone()
        if not e: raise HTTPException(404,'Employee not found')
        attendance=c.execute("SELECT * FROM attendance WHERE employee_id=? AND work_date>=? AND work_date<=? ORDER BY work_date",(employee_id,first.strftime('%Y-%m-%d'),last.strftime('%Y-%m-%d'))).fetchall()
        leaves=c.execute("SELECT * FROM leave_requests WHERE employee_id=? AND status='approved' AND start_date<=? AND end_date>=?",(employee_id,last.strftime('%Y-%m-%d'),first.strftime('%Y-%m-%d'))).fetchall()
        notes=c.execute("SELECT * FROM employee_notes WHERE employee_id=? ORDER BY id DESC LIMIT 50",(employee_id,)).fetchall()
        recent=c.execute("SELECT * FROM attendance WHERE employee_id=? ORDER BY work_date DESC LIMIT 20",(employee_id,)).fetchall()
        reviews=c.execute("SELECT * FROM performance_reviews WHERE employee_id=? ORDER BY reviewed_at DESC,id DESC LIMIT 20",(employee_id,)).fetchall()
        perf=c.execute("SELECT overall_rating,review_period,reviewed_by,reviewed_at FROM performance_reviews WHERE employee_id=? ORDER BY reviewed_at DESC,id DESC LIMIT 1",(employee_id,)).fetchone()
        month_stats=c.execute("SELECT COUNT(*) total,SUM(CASE WHEN status='present' THEN 1 ELSE 0 END) present,SUM(CASE WHEN late_minutes>0 THEN 1 ELSE 0 END) late,SUM(COALESCE(overtime_minutes,0)) overtime FROM attendance WHERE employee_id=? AND work_date>=? AND work_date<=?",(employee_id,first.strftime('%Y-%m-%d'),last.strftime('%Y-%m-%d'))).fetchone()
        payroll=c.execute("SELECT * FROM payroll_records WHERE employee_id=? ORDER BY salary_month DESC LIMIT 24",(employee_id,)).fetchall() if can_payroll_view else []
    amap={a['work_date']:a for a in attendance}; leave_dates=set()
    for l in leaves:
        d=datetime.strptime(l['start_date'],'%Y-%m-%d'); end=datetime.strptime(l['end_date'],'%Y-%m-%d')
        while d<=end: leave_dates.add(d.strftime('%Y-%m-%d')); d+=timedelta(days=1)
    cells=['<div class="sub"><b>'+x+'</b></div>' for x in ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']]
    cells += ['<div class="cal-day empty"></div>']*first.weekday()
    today=now.strftime('%Y-%m-%d')
    for day in range(1,last.day+1):
        ds=f'{month}-{day:02d}'; a=amap.get(ds); cls=''; detail=''
        if ds in leave_dates: cls='leave'; detail='Leave'
        elif a: cls='late' if (a['late_minutes'] or 0)>0 else 'present'; detail=format_time_12h(a['check_in'])
        elif ds<today and datetime.strptime(ds,'%Y-%m-%d').weekday()<5: cls='absent'; detail='Absent'
        cells.append(f"<div class='cal-day {cls}'><b>{day}</b><div class='sub' style='margin-top:9px'>{escape(detail)}</div></div>")
    timeline=''.join(f"<div class='timeline-item'><span class='avatar'>{escape(str(a['work_date'])[-2:])}</span><div><b>{escape(a['work_date'])}</b><div class='sub'>In {escape(format_time_12h(a['check_in']) or '—')} • Out {escape(format_time_12h(a['check_out']) or '—')}</div></div><span class='pill'>{a['late_minutes'] or 0}m late • {a['overtime_minutes'] or 0}m OT</span></div>" for a in recent) or '<div class="sub">No attendance history</div>'
    notehtml=''.join(f"<div style='padding:12px 0;border-bottom:1px solid var(--line)'><span class='tag'>{escape(n['note_type'])}</span> <b>{escape(n['created_by'] or 'HR')}</b><div style='margin-top:6px'>{escape(n['note'])}</div><div class='sub'>{escape(str(n['created_at']))}</div></div>" for n in notes) or '<div class="sub">No HR notes</div>'
    edit=''
    if has_permission(request,'employees_edit'):
        edit=f"""<div class='card'><h3>Edit profile</h3><form method='post' action='/employees/{employee_id}/profile'><div class='two'><div><label>Name</label><input name='name' value='{escape(e['name'])}'><label>Designation</label><input name='designation' value='{escape(e['designation'] or '')}'><label>Department</label><input name='department' value='{escape(e['department'] or '')}'><label>Shift</label><select name='shift'><option {'selected' if e['shift']=='morning' else ''} value='morning'>Morning</option><option {'selected' if e['shift']=='evening' else ''} value='evening'>Evening</option></select></div><div><label>Reporting manager</label><input name='reporting_manager' value='{escape(e['reporting_manager'] or '')}'><label>Office</label><input name='office_name' value='{escape(e['office_name'] or '')}'><label>Join date</label><input type='date' name='join_date' value='{escape(e['join_date'] or '')}'><label>Phone</label><input name='phone' value='{escape(e['phone'] or '')}'></div></div><h3>Emergency contact</h3><div class='grid'><input name='emergency_name' placeholder='Name' value='{escape(e['emergency_name'] or '')}'><input name='emergency_relation' placeholder='Relation' value='{escape(e['emergency_relation'] or '')}'><input name='emergency_phone' placeholder='Phone' value='{escape(e['emergency_phone'] or '')}'><select name='is_active'><option value='1' {'selected' if e['is_active'] else ''}>Active</option><option value='0' {'selected' if not e['is_active'] else ''}>Inactive</option></select></div><button class='btn'>Save profile</button></form></div>"""
    noteform=f"<form method='post' action='/employees/{employee_id}/notes'><div class='two'><select name='note_type'><option>general</option><option>performance</option><option>warning</option><option>promotion</option></select><input name='note' required placeholder='Write a private HR note'></div><button class='btn'>Add note</button></form>" if has_permission(request,'employees_edit') else ''
    reviewhtml=''.join(f"<div style='padding:14px 0;border-bottom:1px solid var(--line)'><div class='card-head'><b>{escape(r['review_period'])}</b><span class='pill'>{(float(r['overall_rating']) if r['overall_rating'] is not None else 0):.1f}/5</span></div><div class='sub'>Reviewed by {escape(r['reviewed_by'] or 'HR')} • {escape(str(r['reviewed_at']))}</div><div style='margin-top:8px'>{escape(r['comments'] or 'No comments')}</div>{f"<div class='sub' style='margin-top:6px'>Goals: {escape(r['goals'])}</div>" if r['goals'] else ''}</div>" for r in reviews) or '<div class="sub">No performance reviews yet</div>'
    reviewform=''
    if has_permission(request,'performance_manage'):
        reviewform=f"""<form method='post' action='/employees/{employee_id}/performance'><div class='two'><div><label>Review period</label><input type='month' name='review_period' value='{now.strftime('%Y-%m')}' required><label>Attendance</label><select name='attendance_rating'>{''.join(f'<option>{x}</option>' for x in range(1,6))}</select><label>Discipline</label><select name='discipline_rating'>{''.join(f'<option>{x}</option>' for x in range(1,6))}</select><label>Work quality</label><select name='work_quality_rating'>{''.join(f'<option>{x}</option>' for x in range(1,6))}</select></div><div><label>Teamwork</label><select name='teamwork_rating'>{''.join(f'<option>{x}</option>' for x in range(1,6))}</select><label>Communication</label><select name='communication_rating'>{''.join(f'<option>{x}</option>' for x in range(1,6))}</select><label>Responsibility</label><select name='responsibility_rating'>{''.join(f'<option>{x}</option>' for x in range(1,6))}</select><label>Comments</label><textarea name='comments' rows='3'></textarea><label>Goals</label><textarea name='goals' rows='2'></textarea></div></div><button class='btn'>Save Review</button></form>"""
    latest_rating=float(perf['overall_rating']) if perf else 0
    attendance_total=int(month_stats['total'] or 0); attendance_present=int(month_stats['present'] or 0); attendance_rate=round((attendance_present/attendance_total*100),1) if attendance_total else 0
    initials=''.join(x[:1] for x in e['name'].split()[:2]).upper()
    payroll_section=''; payroll_tab=''
    if can_payroll_view:
        payroll_tab="<a class='tab' href='#payroll'>Payroll</a>"
        current_payroll=next((p for p in payroll if p['salary_month']==month),None)
        fixed=float(current_payroll['fixed_salary']) if current_payroll else float(e['fixed_salary'] or 0); hours=float(current_payroll['overtime_hours']) if current_payroll else 0; rate=float(current_payroll['overtime_rate']) if current_payroll else float(e['default_overtime_rate'] or 0); bonus=float(current_payroll['bonus']) if current_payroll else 0; deduction=float(current_payroll['deduction']) if current_payroll else 0; advance=float(current_payroll['advance_amount']) if current_payroll else 0; fine=float(current_payroll['fine_amount']) if current_payroll else 0; adjustment_reason=str(current_payroll['adjustment_reason'] or '') if current_payroll else ''
        if current_payroll and str(current_payroll['overtime_mode'] or 'auto')!='manual': hours=0
        payroll_form=''
        if can_payroll_manage:
            payroll_form=f"""<div class='card'><div class='card-head'><div><h3>{'Update' if current_payroll else 'Create'} Salary</h3><div class='sub'>Basic salary stays active until HR changes it.</div></div><span class='tag'>Private</span></div><form method='post' action='/payroll'><input type='hidden' name='employee_id' value='{employee_id}'><input type='hidden' name='profile_employee_id' value='{employee_id}'><div class='two'><div><label>Salary Month</label><input type='month' name='salary_month' value='{escape(month)}' required></div><div><label>Basic Salary</label><input type='number' min='0' step='0.01' name='fixed_salary' value='{fixed:.2f}' required></div></div><div class='two'><div><label>Overtime Mode</label><select name='overtime_mode'><option value='auto'>Automatic</option><option value='manual'>Manual</option></select><label>Manual OT Hours</label><input type='number' min='0' step='0.01' name='overtime_hours' value='{hours:.2f}'></div><div><label>Default OT Rate</label><input type='number' min='0' step='0.01' name='overtime_rate' value='{rate:.2f}'></div></div><div class='two'><div><label>Bonus</label><input type='number' min='0' step='0.01' name='bonus' value='{bonus:.2f}'><label>Advance</label><input type='number' min='0' step='0.01' name='advance' value='{advance:.2f}'></div><div><label>Fine</label><input type='number' min='0' step='0.01' name='fine' value='{fine:.2f}'><label>Other Deduction</label><input type='number' min='0' step='0.01' name='deduction' value='{deduction:.2f}'></div></div><label>Adjustment Reason</label><input name='adjustment_reason' value='{escape(adjustment_reason)}'><label>Private Note</label><textarea name='note'>{escape(current_payroll['note'] or '') if current_payroll else ''}</textarea><button class='btn'>Calculate & Save Draft</button></form></div>"""
        payroll_form=payroll_form.replace("<label>Overtime Mode</label><select name='overtime_mode'><option value='auto'>Automatic</option><option value='manual'>Manual</option></select>","<input type='hidden' name='overtime_mode' value='manual'><div class='sub'>Overtime is never added automatically.</div>")
        payroll_form=payroll_form.replace("Default OT Rate","Manual OT Rate")
        history=[]
        for p in payroll:
            actions=f"<a class='btn secondary' href='/payroll/{p['id']}/payslip.pdf'>PDF</a>" if can_payroll_export else ''
            total_ded=float(p['total_deduction'] or 0) if 'total_deduction' in p.keys() else float(p['deduction'] or 0)
            history.append(f"<tr><td><b>{escape(p['salary_month'])}</b></td><td>{_money(p['fixed_salary'])}</td><td>{_money(p['overtime_amount'])}</td><td>{_money(p['bonus'])}</td><td>{_money(total_ded)}</td><td><b>{_money(p['net_salary'])}</b></td><td><span class='status {'ok' if p['payment_status']=='paid' else 'warn'}'>{escape(p['payment_status'])}</span></td><td>{actions}</td></tr>")
        if current_payroll:
            net=float(current_payroll['net_salary']); summary=f"""<div class='card payroll-panel'><div class='card-head'><div><span class='confidential'>🔒 HR CONFIDENTIAL</span><h2 style='margin-top:12px'>{escape(month)} Payroll</h2></div><span class='status {'ok' if current_payroll['payment_status']=='paid' else 'warn'}'>{escape(current_payroll['payment_status'])}</span></div><div class='payroll-amount'>৳{_money(net)}</div><div class='sub'>Net salary</div><div class='salary-breakdown' style='margin-top:18px'><div><div class='sub'>Fixed</div><b>৳{_money(current_payroll['fixed_salary'])}</b></div><div><div class='sub'>Overtime</div><b>৳{_money(current_payroll['overtime_amount'])}</b></div><div><div class='sub'>Bonus</div><b>৳{_money(current_payroll['bonus'])}</b></div><div><div class='sub'>Deduction</div><b>৳{_money(current_payroll['deduction'])}</b></div></div></div>"""
        else: summary=f"<div class='card payroll-panel'><span class='confidential'>🔒 HR CONFIDENTIAL</span><h2 style='margin-top:14px'>No salary for {escape(month)}</h2><div class='sub'>Create this employee's monthly salary using the form.</div></div>"
        payroll_section=f"""<div id='payroll' class='section-gap'></div><div class='hero'><div><div class='eyebrow'>Confidential Compensation</div><h2>Employee Payroll</h2><div class='sub'>Only authorized HR/Admin can view or change this information.</div></div><a class='btn secondary' href='/payroll?month={escape(month)}'>Monthly Payroll</a></div><div class='two'>{summary}{payroll_form}</div><div class='section-gap'></div><div class='card' style='overflow:auto'><div class='card-head'><div><h3>Salary History</h3><div class='sub'>Latest 24 months</div></div></div><table><thead><tr><th>Month</th><th>Fixed</th><th>OT</th><th>Bonus</th><th>Deduction</th><th>Net</th><th>Status</th><th>Action</th></tr></thead><tbody>{''.join(history) or '<tr><td colspan=8>No salary history.</td></tr>'}</tbody></table></div>"""
    body=f"""<div class='card profile-hero'><div class='profile-photo'>{escape(initials)}</div><div><div class='eyebrow'>Employee 360°</div><h2>{escape(e['name'])}</h2><div class='sub'>{escape(e['staff_id'])} • {escape(e['designation'] or 'No designation')} • {escape(e['department'] or 'No department')}</div><div class='actions' style='margin-top:10px'><span class='status {'ok' if e['is_active'] else 'bad'}'>{'Active' if e['is_active'] else 'Inactive'}</span><span class='status {'ok' if e['registration_status']=='approved' else 'warn'}'>WhatsApp {escape(e['registration_status'])}</span><span class='status {'ok' if e['face_count']>=3 else 'warn'}'>Face {e['face_count']}/3</span></div></div><a class='btn secondary' href='/employees'>Back</a></div><div class='section-gap'></div><div class='summary-strip'><div class='card'><div class='sub'>Attendance</div><div class='metric'>{attendance_rate}%</div></div><div class='card'><div class='sub'>Present</div><div class='metric'>{attendance_present}</div></div><div class='card'><div class='sub'>Late</div><div class='metric'>{int(month_stats['late'] or 0)}</div></div><div class='card'><div class='sub'>Overtime</div><div class='metric'>{round(int(month_stats['overtime'] or 0)/60,1)}h</div></div><div class='card'><div class='sub'>Performance</div><div class='rating'>{latest_rating:.1f}/5</div><div class='stars'>{'★'*round(latest_rating)}{'☆'*(5-round(latest_rating))}</div></div></div><div class='tabs'><a class='tab' href='#profile'>Profile</a><a class='tab' href='#attendance'>Attendance</a><a class='tab' href='#leave'>Leave</a><a class='tab' href='#performance'>Performance</a>{payroll_tab}<a class='tab' href='#activity'>Activity</a></div><div id='profile' class='facts'><div class='fact'><span class='sub'>Shift</span><b>{escape(e['shift'])}</b></div><div class='fact'><span class='sub'>Manager</span><b>{escape(e['reporting_manager'] or '—')}</b></div><div class='fact'><span class='sub'>Office</span><b>{escape(e['office_name'] or 'BURAQ Office')}</b></div><div class='fact'><span class='sub'>Join date</span><b>{escape(e['join_date'] or '—')}</b></div><div class='fact'><span class='sub'>WhatsApp</span><b>{escape(e['whatsapp_phone'] or 'Not registered')}</b></div><div class='fact'><span class='sub'>Emergency</span><b>{escape(e['emergency_name'] or '—')} {escape(e['emergency_phone'] or '')}</b></div></div><div class='section-gap'></div><div class='two'><div class='card'><div class='card-head'><div><h3>Attendance Calendar</h3><div class='sub'>{escape(month)}</div></div><form method='get'><input type='month' name='month' value='{escape(month)}' style='margin:0' onchange='this.form.submit()'></form></div><div class='calendar'>{''.join(cells)}</div></div><div class='card'><h3>Recent timeline</h3><div class='timeline'>{timeline}</div></div></div><div class='section-gap'></div><div id='performance' class='card'><div class='card-head'><div><h3>Performance Review</h3><div class='sub'>Simple 1–5 ratings with comments and goals</div></div><span class='pill'>{latest_rating:.1f}/5 latest</span></div>{reviewform}<div class='section-gap'></div>{reviewhtml}</div>{payroll_section}<div class='section-gap'></div><div class='two'>{edit}<div id='activity' class='card'><h3>HR Notes</h3>{noteform}<div class='section-gap'></div>{notehtml}</div></div>"""
    return layout(e['name'],body,request,'employees')

@app.get("/employees/{employee_id}/duty", response_class=HTMLResponse)
def employee_duty_page(request: Request, employee_id: int, saved: str=""):
    require_permission(request,'duty_view'); can_manage=has_permission(request,'duty_manage')
    today=datetime.now(ZoneInfo(settings.timezone)).date().isoformat(); days=['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
    with get_db() as c:
        e=c.execute("SELECT * FROM employees WHERE id=?",(employee_id,)).fetchone()
        weekly=c.execute("SELECT * FROM duty_schedules WHERE employee_id=? ORDER BY weekday",(employee_id,)).fetchall()
        custom=c.execute("SELECT * FROM custom_duties WHERE employee_id=? AND duty_date>=? ORDER BY duty_date",(employee_id,today)).fetchall()
    if not e: raise HTTPException(404,'Employee not found')
    first_start,first_end=shift_window('first'); second_start,second_end=shift_window('second')
    first_preset=f"{first_start.strftime('%H:%M')}-{first_end.strftime('%H:%M')}"
    second_preset=f"{second_start.strftime('%H:%M')}-{second_end.strftime('%H:%M')}"
    forms=''
    if can_manage:
        day_options=''.join(f"<option value='{i}'>{d}</option>" for i,d in enumerate(days))
        forms=f"""<div class='two'><div class='card'><div class='eyebrow'>Repeating Schedule</div><h2>Regular Duty by Shift</h2><form method='post' action='/employees/{employee_id}/duty/regular'><label>Weekday</label><select name='weekday'>{day_options}</select><label>Shift preset</label><select name='preset'><option value='morning'>First Shift ({first_preset})</option><option value='evening'>Second Shift ({second_preset})</option><option value='night'>Night (22:00-06:00)</option><option value='custom'>Custom selectable time</option></select><div class='two'><div><label>Custom start (optional)</label><input type='time' name='start_time'></div><div><label>Custom end (optional)</label><input type='time' name='end_time'></div></div><label>Break (minutes)</label><input type='number' name='break_minutes' min='0' step='5' value='60'><label>Office</label><input name='office_name' value='{escape(e['office_name'] or 'BURAQ Office')}'><button class='btn'>Assign Regular Duty</button></form></div><div class='card money-card'><div class='eyebrow'>One Specific Date</div><h2>Custom Duty</h2><form method='post' action='/employees/{employee_id}/duty/custom'><label>Date</label><input type='date' name='duty_date' required><div class='two'><div><label>Start</label><input type='time' name='start_time' required></div><div><label>End</label><input type='time' name='end_time' required></div></div><label>Break (minutes)</label><input type='number' name='break_minutes' min='0' step='5' value='60'><label>Office</label><input name='office_name' value='{escape(e['office_name'] or 'BURAQ Office')}'><label>Note</label><input name='note' placeholder='Special duty reason'><button class='btn'>Assign Custom Duty</button></form></div></div><div class='section-gap'></div><div class='two'><div class='card'><div class='eyebrow'>Quick Assignment</div><h2>Assign Friday Duty</h2><form method='post' action='/employees/{employee_id}/duty/friday'><div class='two'><div><label>Start</label><input type='time' name='start_time' required></div><div><label>End</label><input type='time' name='end_time' required></div></div><label>Break (minutes)</label><input type='number' name='break_minutes' min='0' step='5' value='60'><label>Office</label><input name='office_name' value='{escape(e['office_name'] or 'BURAQ Office')}'><button class='btn'>Assign Every Friday</button></form></div><div class='card payroll-panel'><div class='eyebrow' style='color:#8ff0cb'>Overnight Assignment</div><h2>Assign Night Duty</h2><form method='post' action='/employees/{employee_id}/duty/night'><label>Starting date</label><input type='date' name='duty_date' required><div class='two'><div><label>Night start</label><input type='time' name='start_time' value='22:00' required></div><div><label>Next-day end</label><input type='time' name='end_time' value='06:00' required></div></div><label>Break (minutes)</label><input type='number' name='break_minutes' min='0' step='5' value='60'><label>Repeat</label><select name='repeat'><option value='once'>One-time night duty</option><option value='weekly'>Repeat every week on this weekday</option></select><label>Office</label><input name='office_name' value='{escape(e['office_name'] or 'BURAQ Office')}'><button class='btn'>Assign Night Duty</button></form></div></div>"""
    weekly_rows=''.join(f"<tr><td>{days[int(r['weekday'])]}</td><td>{escape(format_time_12h(r['start_time']))} - {escape(format_time_12h(r['end_time']))}{' (+1 day)' if r['end_time']<=r['start_time'] else ''}<div class='sub'>Break: {int(r['break_minutes'] or 0)} min</div></td><td>{escape(r['office_name'] or 'BURAQ Office')}</td><td>{f'''<form method='post' action='/employees/{employee_id}/duty/weekly/{r['id']}/delete'><button class='btn danger'>Delete</button></form>''' if can_manage else ''}</td></tr>" for r in weekly) or '<tr><td colspan=4>No regular duty.</td></tr>'
    custom_rows=''.join(f"<tr><td>{escape(r['duty_date'])}</td><td>{escape(format_time_12h(r['start_time']))} - {escape(format_time_12h(r['end_time']))}{' (+1 day)' if r['end_time']<=r['start_time'] else ''}<div class='sub'>Break: {int(r['break_minutes'] or 0)} min</div></td><td>{escape(r['office_name'] or 'BURAQ Office')}<div class='sub'>{escape(r['note'] or '')}</div></td><td>{f'''<form method='post' action='/employees/{employee_id}/duty/custom/{r['id']}/delete'><button class='btn danger'>Delete</button></form>''' if can_manage else ''}</td></tr>" for r in custom) or '<tr><td colspan=4>No upcoming custom duty.</td></tr>'
    notice="<div class='notice'>Duty assignment saved.</div>" if saved else ''
    body=f"""{notice}<div class='card profile-hero'><div class='profile-photo'>{escape(''.join(x[:1] for x in e['name'].split()[:2]).upper())}</div><div><div class='eyebrow'>Employee Duty Control</div><h2>{escape(e['name'])}</h2><div class='sub'>{escape(e['staff_id'])} • Current shift: {escape(e['shift'])}</div></div><div class='actions'><a class='btn secondary' href='/employees/{employee_id}'>Profile</a><a class='btn secondary' href='/employees'>Employees</a></div></div><div class='section-gap'></div>{forms}<div class='section-gap'></div><div class='two'><div class='card' style='overflow:auto'><h2>Regular Weekly Duty</h2><table><thead><tr><th>Day</th><th>Time</th><th>Office</th><th></th></tr></thead><tbody>{weekly_rows}</tbody></table></div><div class='card' style='overflow:auto'><h2>Upcoming Custom Duty</h2><table><thead><tr><th>Date</th><th>Time</th><th>Office</th><th></th></tr></thead><tbody>{custom_rows}</tbody></table></div></div>"""
    return layout(f"{e['name']} Duty",body,request,'employees')

def _duty_times(preset: str, start_time: str, end_time: str):
    """Presets follow the Admin-configured global shift rules."""
    if start_time and end_time: return start_time,end_time
    first_start,first_end=shift_window('first'); second_start,second_end=shift_window('second')
    presets={'morning':(first_start.strftime('%H:%M'),first_end.strftime('%H:%M')),
             'evening':(second_start.strftime('%H:%M'),second_end.strftime('%H:%M')),
             'night':('22:00','06:00')}
    if preset in presets: return presets[preset]
    raise HTTPException(400,'Select start and end time')

@app.post("/employees/{employee_id}/duty/regular")
def assign_regular_duty(request: Request, employee_id: int, weekday: int=Form(...), preset: str=Form('morning'), start_time: str=Form(''), end_time: str=Form(''), office_name: str=Form('BURAQ Office'), break_minutes: int=Form(0)):
    require_permission(request,'duty_manage'); start_time,end_time=_duty_times(preset,start_time,end_time)
    if weekday not in range(7): raise HTTPException(400,'Invalid weekday')
    break_minutes=_validated_break_minutes(start_time,end_time,break_minutes)
    with get_db() as c: c.execute("INSERT INTO duty_schedules(employee_id,weekday,start_time,end_time,break_minutes,office_name,created_by) VALUES(?,?,?,?,?,?,?) ON CONFLICT(employee_id,weekday) DO UPDATE SET start_time=excluded.start_time,end_time=excluded.end_time,break_minutes=excluded.break_minutes,office_name=excluded.office_name,is_active=excluded.is_active,updated_at=CURRENT_TIMESTAMP",(employee_id,weekday,start_time,end_time,break_minutes,office_name.strip() or 'BURAQ Office',str(request.session.get('hr_id') or 'super_admin')))
    return RedirectResponse(f'/employees/{employee_id}/duty?saved=1',303)

@app.post("/employees/{employee_id}/duty/custom")
def assign_employee_custom_duty(request: Request, employee_id: int, duty_date: str=Form(...), start_time: str=Form(...), end_time: str=Form(...), office_name: str=Form('BURAQ Office'), note: str=Form(''), break_minutes: int=Form(0)):
    require_permission(request,'duty_manage')
    try: datetime.strptime(duty_date,'%Y-%m-%d')
    except ValueError: raise HTTPException(400,'Invalid date')
    break_minutes=_validated_break_minutes(start_time,end_time,break_minutes)
    with get_db() as c: c.execute("INSERT INTO custom_duties(employee_id,duty_date,start_time,end_time,break_minutes,office_name,note,created_by) VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(employee_id,duty_date) DO UPDATE SET start_time=excluded.start_time,end_time=excluded.end_time,break_minutes=excluded.break_minutes,office_name=excluded.office_name,note=excluded.note,is_active=excluded.is_active,updated_at=CURRENT_TIMESTAMP",(employee_id,duty_date,start_time,end_time,break_minutes,office_name.strip() or 'BURAQ Office',note.strip() or None,str(request.session.get('hr_id') or 'super_admin')))
    return RedirectResponse(f'/employees/{employee_id}/duty?saved=1',303)

@app.post("/employees/{employee_id}/duty/friday")
def assign_friday_duty(request: Request, employee_id: int, start_time: str=Form(...), end_time: str=Form(...), office_name: str=Form('BURAQ Office'), break_minutes: int=Form(0)):
    return assign_regular_duty(request,employee_id,4,'custom',start_time,end_time,office_name,break_minutes)

@app.post("/employees/{employee_id}/duty/night")
def assign_night_duty(request: Request, employee_id: int, duty_date: str=Form(...), start_time: str=Form(...), end_time: str=Form(...), repeat: str=Form('once'), office_name: str=Form('BURAQ Office'), break_minutes: int=Form(0)):
    require_permission(request,'duty_manage')
    try: day=datetime.strptime(duty_date,'%Y-%m-%d')
    except ValueError: raise HTTPException(400,'Invalid date')
    if repeat=='weekly': return assign_regular_duty(request,employee_id,day.weekday(),'custom',start_time,end_time,office_name,break_minutes)
    return assign_employee_custom_duty(request,employee_id,duty_date,start_time,end_time,office_name,'Night duty',break_minutes)

@app.post("/employees/{employee_id}/duty/weekly/{schedule_id}/delete")
def delete_employee_weekly_duty(request: Request, employee_id: int, schedule_id: int):
    require_permission(request,'duty_manage')
    with get_db() as c: c.execute("DELETE FROM duty_schedules WHERE id=? AND employee_id=?",(schedule_id,employee_id))
    return RedirectResponse(f'/employees/{employee_id}/duty',303)

@app.post("/employees/{employee_id}/duty/custom/{duty_id}/delete")
def delete_employee_custom_duty(request: Request, employee_id: int, duty_id: int):
    require_permission(request,'duty_manage')
    with get_db() as c: c.execute("DELETE FROM custom_duties WHERE id=? AND employee_id=?",(duty_id,employee_id))
    return RedirectResponse(f'/employees/{employee_id}/duty',303)


@app.post("/employees/{employee_id}/performance")
def add_performance_review(request: Request, employee_id: int, review_period: str=Form(...), attendance_rating: int=Form(...), discipline_rating: int=Form(...), work_quality_rating: int=Form(...), teamwork_rating: int=Form(...), communication_rating: int=Form(...), responsibility_rating: int=Form(...), comments: str=Form(""), goals: str=Form("")):
    require_permission(request,"performance_manage")
    ratings=[attendance_rating,discipline_rating,work_quality_rating,teamwork_rating,communication_rating,responsibility_rating]
    if any(x<1 or x>5 for x in ratings): raise HTTPException(400,"Ratings must be 1 to 5")
    overall=round(sum(ratings)/len(ratings),2)
    with get_db() as c:
        c.execute("INSERT INTO performance_reviews(employee_id,review_period,attendance_rating,discipline_rating,work_quality_rating,teamwork_rating,communication_rating,responsibility_rating,overall_rating,comments,goals,reviewed_by) VALUES(?,?,?,?,?,?,?,?,?,?,?,?)",(employee_id,review_period,*ratings,overall,comments.strip() or None,goals.strip() or None,request.session.get('user_name','Admin')))
        audit(request,'performance_review_created','employee',str(employee_id),f'{review_period}: {overall}/5',db=c)
    return RedirectResponse(f'/employees/{employee_id}#performance',303)

@app.get("/performance", response_class=HTMLResponse)
def performance_page(request: Request):
    require_permission(request,"performance_view")
    with get_db() as c:
        rows=c.execute("SELECT e.id,e.staff_id,e.name,e.department,e.designation,p.overall_rating,p.review_period,p.reviewed_at FROM employees e LEFT JOIN performance_reviews p ON p.id=(SELECT p2.id FROM performance_reviews p2 WHERE p2.employee_id=e.id ORDER BY p2.reviewed_at DESC,p2.id DESC LIMIT 1) WHERE e.is_active ORDER BY e.name").fetchall()
    trs=''.join(f"<tr><td><b>{escape(r['name'])}</b><div class='sub'>{escape(r['staff_id'])}</div></td><td>{escape(r['department'] or '—')}</td><td>{escape(r['designation'] or '—')}</td><td>{(float(r['overall_rating']) if r['overall_rating'] is not None else 0):.1f}/5"+f"<div class='sub'>{escape(r['review_period'] or '')}</div>"+f"</td><td><a class='btn secondary' href='/employees/{r['id']}#performance'>{'Review' if r['overall_rating'] is None else 'Open'}</a></td></tr>" for r in rows) or "<tr><td colspan='5'>No employees</td></tr>"
    body=f"""<div class='hero'><div><div class='eyebrow'>People Development</div><h2>Performance Reviews</h2><div class='sub'>Simple, consistent employee reviews without unnecessary complexity.</div></div><span class='pill'>{len(rows)} employees</span></div><div class='card'><table><tr><th>Employee</th><th>Department</th><th>Designation</th><th>Latest rating</th><th></th></tr>{trs}</table></div>"""
    return layout('Performance',body,request,'performance')

@app.post("/employees/{employee_id}/profile")
def update_employee_profile(request: Request, employee_id: int, name: str=Form(...), phone: str=Form(""), designation: str=Form(""), department: str=Form(""), shift: str=Form("morning"), reporting_manager: str=Form(""), office_name: str=Form(""), join_date: str=Form(""), emergency_name: str=Form(""), emergency_relation: str=Form(""), emergency_phone: str=Form(""), is_active: int=Form(1)):
    require_permission(request,'employees_edit')
    with get_db() as c:
        c.execute("UPDATE employees SET name=?,phone=?,designation=?,department=?,shift=?,reporting_manager=?,office_name=?,join_date=?,emergency_name=?,emergency_relation=?,emergency_phone=?,is_active=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",(name.strip(),phone.strip() or None,designation.strip() or None,department.strip() or None,shift,reporting_manager.strip() or None,office_name.strip() or None,join_date or None,emergency_name.strip() or None,emergency_relation.strip() or None,emergency_phone.strip() or None,1 if is_active else 0,employee_id))
        audit(request,'employee_profile_updated','employee',str(employee_id),db=c)
    return RedirectResponse(f'/employees/{employee_id}',303)

@app.post("/employees/{employee_id}/notes")
def add_employee_note(request: Request, employee_id: int, note_type: str=Form('general'), note: str=Form(...)):
    require_permission(request,'employees_edit')
    with get_db() as c:
        c.execute("INSERT INTO employee_notes(employee_id,note_type,note,created_by) VALUES(?,?,?,?)",(employee_id,note_type[:30],note.strip(),request.session.get('user_name','Admin')))
        audit(request,'employee_note_added','employee',str(employee_id),note_type,db=c)
    return RedirectResponse(f'/employees/{employee_id}',303)

@app.get("/employees/{employee_id}/reset", response_class=HTMLResponse)
def employee_reset_page(request: Request, employee_id: int):
    require_permission(request, "face_reset")
    with get_db() as c:
        employee=c.execute("SELECT id,staff_id,name,registration_status,whatsapp_phone,phone FROM employees WHERE id=?",(employee_id,)).fetchone()
    if not employee:
        raise HTTPException(404,"Employee not found")
    can_reset_all = request.session.get("role") == "super_admin" and bool(request.session.get("admin"))
    reset_all_card = ""
    if can_reset_all:
        reset_all_card=f"""<div class='card' style='border-color:#fecaca'>
        <div class='eyebrow' style='color:#b91c1c'>Danger zone</div><h3>Reset All</h3>
        <p class='sub'>Attendance, leave, payroll, performance, duty, face data and onboarding history for this employee will be permanently removed. The employee master record and Basic Salary field remain available for setup again.</p>
        <form method='post' action='/employees/{employee_id}/reset-all'>
          <label>Type RESET ALL to confirm</label><input name='confirmation' autocomplete='off' placeholder='RESET ALL' required>
          <button class='btn danger'>Reset All</button>
        </form></div>"""
    body=f"""<div class='hero'><div><div class='eyebrow'>Employee reset</div><h2>{escape(employee['name'])}</h2><div class='sub'>{escape(employee['staff_id'])} · {escape(employee['registration_status'])}</div></div><a class='btn secondary' href='/employees'>Back</a></div>
    <div class='two'><div class='card'><div class='eyebrow'>Recommended</div><h3>Reset</h3>
    <p class='sub'>Restarts WhatsApp registration and Face AI setup only. Employee profile, attendance, leave, payroll, performance and duty history stay unchanged.</p>
    <form method='post' action='/employees/{employee_id}/reset-registration'><button class='btn'>Reset</button></form></div>{reset_all_card}</div>"""
    return layout("Reset Employee",body,request,"employees")

@app.post("/employees/{employee_id}/reset-registration")
def reset_employee_registration(request: Request, employee_id: int, background_tasks: BackgroundTasks):
    require_permission(request, "face_reset")
    notify_phone=None
    with get_db() as c:
        employee=c.execute("SELECT staff_id,name,whatsapp_phone,phone FROM employees WHERE id=?",(employee_id,)).fetchone()
        if not employee: raise HTTPException(404,"Employee not found")
        notify_phone=employee["whatsapp_phone"] or employee["phone"]
        for phone in {employee["whatsapp_phone"], employee["phone"]}:
            if phone: c.execute("DELETE FROM conversation_states WHERE phone=?",(re.sub(r"\D","",phone),))
        c.execute("DELETE FROM face_samples WHERE employee_id=?",(employee_id,))
        c.execute("DELETE FROM face_profiles WHERE employee_id=?",(employee_id,))
        c.execute("DELETE FROM pending_registrations WHERE employee_id=?",(employee_id,))
        c.execute("UPDATE employees SET registration_status='unregistered',whatsapp_phone=NULL,updated_at=CURRENT_TIMESTAMP WHERE id=?",(employee_id,))
        if notify_phone:
            normalized=re.sub(r"\D","",notify_phone)
            c.execute("INSERT INTO conversation_states(phone,state) VALUES(?,?) ON CONFLICT(phone) DO UPDATE SET state=excluded.state,updated_at=CURRENT_TIMESTAMP",(normalized,"awaiting_staff_id"))
        audit(request,'employee_registration_reset','employee',str(employee_id),'face and WhatsApp onboarding reset',db=c)
    if notify_phone:
        background_tasks.add_task(send_text,notify_phone,"🔄 আপনার BURAQ Attendance registration reset করা হয়েছে।\n\nআবার শুরু করতে আপনার Staff ID পাঠান।")
    return RedirectResponse(f"/employees/{employee_id}?reset=registration",303)

@app.post("/employees/{employee_id}/reset-all")
def reset_employee_all(request: Request, employee_id: int, confirmation: str=Form(...)):
    require_super_admin(request)
    if confirmation.strip().upper() != "RESET ALL":
        return RedirectResponse(f"/employees/{employee_id}/reset?error=confirmation",303)
    with get_db() as c:
        employee=c.execute("SELECT whatsapp_phone,phone FROM employees WHERE id=?",(employee_id,)).fetchone()
        if not employee: raise HTTPException(404,"Employee not found")
        payroll_rows=c.execute("SELECT id FROM payroll_records WHERE employee_id=?",(employee_id,)).fetchall()
        for row in payroll_rows:
            c.execute("DELETE FROM payroll_change_logs WHERE payroll_id=?",(row['id'],))
        # Delete dependent records in a safe order. The employee master row is preserved.
        for table in ("attendance_fingerprints","attendance_evidence","attendance_corrections","leave_requests","performance_reviews","employee_notes","duty_reminder_logs","custom_duties","duty_schedules","payroll_records","attendance","pending_registrations","face_samples","face_profiles"):
            c.execute(f"DELETE FROM {table} WHERE employee_id=?",(employee_id,))
        for phone in {employee["whatsapp_phone"], employee["phone"]}:
            if phone: c.execute("DELETE FROM conversation_states WHERE phone=?",(re.sub(r"\D","",phone),))
        c.execute("UPDATE employees SET registration_status='unregistered',whatsapp_phone=NULL,updated_at=CURRENT_TIMESTAMP WHERE id=?",(employee_id,))
        audit(request,'employee_all_reset','employee',str(employee_id),'all employee operational history reset; master record preserved',db=c)
    return RedirectResponse(f"/employees/{employee_id}?reset=all",303)

# Backward-compatible endpoint for older bookmarks/forms.
@app.post("/employees/{employee_id}/reset-face")
def reset_employee_face_legacy(request: Request, employee_id: int, background_tasks: BackgroundTasks):
    return reset_employee_registration(request,employee_id,background_tasks)

@app.get("/pending", response_class=HTMLResponse)
def pending_page(request: Request):
    require_permission(request, "approvals_view")
    with get_db() as c:
        rows=c.execute("SELECT p.id,p.whatsapp_phone,p.created_at,e.staff_id,e.name FROM pending_registrations p JOIN employees e ON e.id=p.employee_id WHERE p.status='pending' ORDER BY p.id DESC").fetchall()
    can_manage=has_permission(request,"approvals_manage")
    parts=[]
    for r in rows:
        approve = f"<form method='post' action='/pending/{r['id']}/approve'><button class='btn'>Approve</button></form>" if can_manage else "<span class='sub'>View only</span>"
        reject = f"<form method='post' action='/pending/{r['id']}/reject'><button class='btn danger'>Reject</button></form>" if can_manage else ""
        parts.append(f"<tr><td>{escape(r['staff_id'])}</td><td>{escape(r['name'])}</td><td>{escape(r['whatsapp_phone'])}</td><td>{approve}</td><td>{reject}</td></tr>")
    trs=''.join(parts) or "<tr><td colspan='5'>কোনো pending registration নেই</td></tr>"
    return layout("Approvals", f"<div class='card'><table><thead><tr><th>Staff ID</th><th>Name</th><th>WhatsApp</th><th></th><th></th></tr></thead><tbody>{trs}</tbody></table></div>", request, "pending")

@app.post("/pending/{pending_id}/approve")
def approve_pending(request: Request, pending_id: int, background_tasks: BackgroundTasks):
    require_permission(request,"approvals_manage")
    notify = None
    with get_db() as c:
        row=c.execute("SELECT p.*,e.name,e.staff_id FROM pending_registrations p JOIN employees e ON e.id=p.employee_id WHERE p.id=? AND p.status='pending'", (pending_id,)).fetchone()
        if row:
            c.execute("UPDATE employees SET whatsapp_phone=?,registration_status='approved',updated_at=CURRENT_TIMESTAMP WHERE id=?", (row['whatsapp_phone'],row['employee_id']))
            c.execute("UPDATE pending_registrations SET status='approved',updated_at=CURRENT_TIMESTAMP WHERE id=?", (pending_id,))
            c.execute("INSERT INTO conversation_states(phone,state) VALUES(?,?) ON CONFLICT(phone) DO UPDATE SET state=excluded.state,updated_at=CURRENT_TIMESTAMP", (row['whatsapp_phone'],'awaiting_face_registration'))
            notify = (row['whatsapp_phone'], row['name'], row['staff_id'])
    if notify:
        background_tasks.add_task(send_approval_flow, *notify)
    return RedirectResponse("/pending",303)

@app.post("/pending/{pending_id}/reject")
def reject_pending(request: Request, pending_id: int):
    require_permission(request, "approvals_manage")
    with get_db() as c: c.execute("UPDATE pending_registrations SET status='rejected',updated_at=CURRENT_TIMESTAMP WHERE id=?", (pending_id,))
    return RedirectResponse("/pending",303)

@app.get("/export/attendance.csv")
def export_attendance(request: Request):
    require_permission(request, "reports_export")
    with get_db() as c: rows=c.execute("SELECT a.work_date,e.staff_id,e.name,e.department,CASE WHEN a.attendance_shift='second' THEN 'Second Shift' ELSE 'First Shift' END AS shift,a.check_in,a.check_out,a.late_minutes,a.early_leave_minutes,a.overtime_minutes,a.status FROM attendance a JOIN employees e ON e.id=a.employee_id ORDER BY a.work_date DESC,e.staff_id").fetchall()
    output=io.StringIO(); writer=csv.writer(output); writer.writerow(["Date","Staff ID","Name","Department","Shift","Check In","Check Out","Late Minutes","Early Leave Minutes","Overtime Minutes","Status"])
    for r in rows: writer.writerow(list(r.values()))
    data=output.getvalue().encode("utf-8-sig")
    return StreamingResponse(io.BytesIO(data),media_type="text/csv",headers={"Content-Disposition":"attachment; filename=BURAQ-Attendance.csv"})


@app.get("/hr-accounts", response_class=HTMLResponse)
def hr_accounts_page(request: Request, saved: str = "", error: str = ""):
    require_permission(request, "user_accounts_view")
    can_manage = has_permission(request, "user_accounts_manage")
    is_super = request.session.get("role") == "super_admin" and request.session.get("admin")
    with get_db() as c:
        rows = c.execute("SELECT * FROM hr_accounts ORDER BY is_active DESC, name").fetchall()
        permission_rows = c.execute("SELECT account_id,permission FROM account_permissions").fetchall()
    by_account = {}
    for pr in permission_rows:
        by_account.setdefault(pr["account_id"], set()).add(pr["permission"])
    notice = "<div class='notice'>Account or permissions saved successfully.</div>" if saved else ("<div class='notice' style='background:#fee2e2;color:#991b1b'>Action failed. Check account data and permissions.</div>" if error else "")
    trs=[]
    for r in rows:
        perms=by_account.get(r['id']) or DEFAULT_ROLE_PERMISSIONS.get(r['role'], set())
        chips=' '.join(f"<span class='status ok' style='margin:2px'>{escape(PERMISSION_CATALOG[p][0])}</span>" for p in sorted(perms) if p in PERMISSION_CATALOG) or "<span class='sub'>No permissions</span>"
        actions=""
        if can_manage:
            actions=f"<div class='actions'><form method='post' action='/hr-accounts/{r['id']}/toggle'><button class='btn secondary'>{'Disable' if r['is_active'] else 'Enable'}</button></form><form method='post' action='/hr-accounts/{r['id']}/delete' onsubmit=\"return confirm('Delete this account?')\"><button class='btn danger'>Delete</button></form></div>"
        if is_super:
            actions += f"<a class='btn secondary' href='/hr-accounts/{r['id']}/permissions'>Permissions</a>"
        trs.append(f"<tr><td><b>{escape(r['name'])}</b><div class='sub'>{escape(r['email'])}</div></td><td>{escape(r['role'].replace('_',' ').title())}</td><td><span class='status {'ok' if r['is_active'] else 'bad'}'>{'Active' if r['is_active'] else 'Disabled'}</span></td><td><div style='max-width:520px'>{chips}</div></td><td>{actions or '<span class=\'sub\'>View only</span>'}</td></tr>")
    create_card=""
    if can_manage:
        create_card="""<div class='card'><h2>Add Admin / HR Account</h2><form method='post'><label>Full Name</label><input name='name' required><label>Email</label><input type='email' name='email' required><label>Role</label><select name='role'><option value='admin'>Admin</option><option value='hr_manager'>HR Manager</option><option value='hr_executive'>HR Executive</option><option value='hr_officer'>HR Officer</option><option value='viewer'>Viewer</option></select><label>Temporary Password</label><input type='password' name='password' minlength='8' required><button class='btn'>Create Account</button></form></div>"""
    body=f"{notice}<div class='two'>{create_card}<div class='card'><h2>Dynamic Permission Rules</h2><p>Super Admin প্রতিটি account-এর জন্য আলাদা permission নির্বাচন করবে। Permission না থাকলে menu লুকানো থাকবে এবং direct URL-এ 403 হবে।</p><p class='sub'>Permission changes audit log-এ সংরক্ষিত হয় এবং পরের request থেকেই কার্যকর হয়।</p></div></div><div class='section-gap'></div><div class='card'><h2>Admin & HR Accounts</h2><div style='overflow:auto'><table><thead><tr><th>User</th><th>Role</th><th>Status</th><th>Current Permissions</th><th>Action</th></tr></thead><tbody>{''.join(trs) or '<tr><td colspan=\'5\'>No accounts yet.</td></tr>'}</tbody></table></div></div>"
    return layout("User Accounts", body, request, "hr")

@app.post("/hr-accounts")
def add_hr_account(request: Request, name: str = Form(...), email: str = Form(...), role: str = Form(...), password: str = Form(...)):
    require_permission(request, "user_accounts_manage")
    if role not in DEFAULT_ROLE_PERMISSIONS or len(password) < 8:
        return RedirectResponse("/hr-accounts?error=1",303)
    try:
        with get_db() as c:
            cur=c.execute("INSERT INTO hr_accounts(name,email,password_hash,role) VALUES(?,?,?,?)", (name.strip(), email.strip().lower(), hash_password(password), role))
            account_id=getattr(cur, 'lastrowid', None)
            if not account_id:
                row=c.execute("SELECT id FROM hr_accounts WHERE LOWER(email)=LOWER(?)",(email.strip(),)).fetchone(); account_id=row['id']
            c.execute("INSERT INTO account_permissions(account_id,permission) VALUES(?,?) ON CONFLICT(account_id,permission) DO NOTHING",(account_id,"__configured__"))
            for permission in DEFAULT_ROLE_PERMISSIONS.get(role,set()):
                c.execute("INSERT INTO account_permissions(account_id,permission) VALUES(?,?) ON CONFLICT(account_id,permission) DO NOTHING",(account_id,permission))
        audit(request,"create","hr_account",email.strip().lower(),f"Role: {role}")
    except Exception as exc:
        logger.warning("Account create failed: %s", exc)
        return RedirectResponse("/hr-accounts?error=1",303)
    return RedirectResponse("/hr-accounts?saved=1",303)

@app.post("/hr-accounts/{account_id}/toggle")
def toggle_hr_account(request: Request, account_id: int):
    require_permission(request, "user_accounts_manage")
    with get_db() as c:
        row=c.execute("SELECT is_active,email FROM hr_accounts WHERE id=?",(account_id,)).fetchone()
        if row:
            c.execute("UPDATE hr_accounts SET is_active=?,updated_at=CURRENT_TIMESTAMP WHERE id=?", (not bool(row['is_active']), account_id))
    audit(request,"toggle","hr_account",str(account_id),row['email'] if row else '')
    return RedirectResponse("/hr-accounts",303)

@app.post("/hr-accounts/{account_id}/delete")
def delete_hr_account(request: Request, account_id: int):
    require_permission(request, "user_accounts_manage")
    with get_db() as c:
        row=c.execute("SELECT email FROM hr_accounts WHERE id=?",(account_id,)).fetchone()
        c.execute("DELETE FROM hr_accounts WHERE id=?",(account_id,))
    audit(request,"delete","hr_account",str(account_id),row['email'] if row else '')
    return RedirectResponse("/hr-accounts",303)

@app.get("/hr-accounts/{account_id}/permissions", response_class=HTMLResponse)
def account_permissions_page(request: Request, account_id: int, saved: str = ""):
    require_super_admin(request)
    with get_db() as c:
        account=c.execute("SELECT * FROM hr_accounts WHERE id=?",(account_id,)).fetchone()
        rows=c.execute("SELECT permission FROM account_permissions WHERE account_id=?",(account_id,)).fetchall()
    if not account: raise HTTPException(404,"Account not found")
    selected={r['permission'] for r in rows} or DEFAULT_ROLE_PERMISSIONS.get(account['role'],set())
    checks=''.join(f"<label style='display:flex;gap:10px;align-items:flex-start;padding:12px;border:1px solid var(--line);border-radius:12px;margin-bottom:9px'><input style='width:auto;margin:3px 0' type='checkbox' name='permissions' value='{escape(key)}' {'checked' if key in selected else ''}><span><b>{escape(label)}</b><div class='sub'>{escape(desc)}</div></span></label>" for key,(label,desc) in PERMISSION_CATALOG.items())
    notice="<div class='notice'>Permissions updated.</div>" if saved else ""
    body=f"{notice}<div class='card' style='max-width:900px'><h2>{escape(account['name'])}</h2><p class='sub'>{escape(account['email'])} • {escape(account['role'].replace('_',' ').title())}</p><form method='post'>{checks}<div class='actions'><button class='btn'>Save Permissions</button><a class='btn secondary' href='/hr-accounts'>Back</a></div></form></div>"
    return layout("Account Permissions",body,request,"hr")

@app.post("/hr-accounts/{account_id}/permissions")
async def save_account_permissions(request: Request, account_id: int):
    require_super_admin(request)
    form=await request.form()
    selected={p for p in form.getlist('permissions') if p in PERMISSION_CATALOG}
    with get_db() as c:
        account=c.execute("SELECT email FROM hr_accounts WHERE id=?",(account_id,)).fetchone()
        if not account: raise HTTPException(404,"Account not found")
        c.execute("DELETE FROM account_permissions WHERE account_id=?",(account_id,))
        c.execute("INSERT INTO account_permissions(account_id,permission) VALUES(?,?)",(account_id,"__configured__"))
        for permission in sorted(selected):
            c.execute("INSERT INTO account_permissions(account_id,permission) VALUES(?,?)",(account_id,permission))
    audit(request,"permissions_update","hr_account",str(account_id),", ".join(sorted(selected)))
    return RedirectResponse(f"/hr-accounts/{account_id}/permissions?saved=1",303)

@app.get("/audit-logs", response_class=HTMLResponse)
def audit_logs_page(request: Request):
    require_permission(request, "audit_view")
    with get_db() as c:
        rows=c.execute("SELECT * FROM audit_logs ORDER BY id DESC LIMIT 300").fetchall()
    trs=''.join(f"<tr><td>{escape(str(r['created_at']))}</td><td><b>{escape(r['actor_name'] or r['actor_type'])}</b><div class='sub'>{escape(r['actor_type'])}</div></td><td>{escape(r['action'])}</td><td>{escape((r['target_type'] or '')+' '+(r['target_id'] or ''))}</td><td>{escape(r['details'] or '')}</td><td>{escape(r['ip_address'] or '')}</td></tr>" for r in rows) or "<tr><td colspan='6'>No activity yet.</td></tr>"
    return layout("Activity Logs",f"<div class='card'><h2>Security & Activity Audit</h2><div style='overflow:auto'><table><thead><tr><th>Time</th><th>Actor</th><th>Action</th><th>Target</th><th>Details</th><th>IP</th></tr></thead><tbody>{trs}</tbody></table></div></div>",request,"audit")


def _attendance_report_rows(start_date: str, end_date: str, status: str = "", department: str = ""):
    clauses = ["a.work_date>=?", "a.work_date<=?"]
    params = [start_date, end_date]
    if status:
        clauses.append("a.status=?"); params.append(status)
    if department:
        clauses.append("e.department=?"); params.append(department)
    sql = "SELECT a.*,e.staff_id,e.name,e.department,CASE WHEN a.attendance_shift='second' THEN 'Second Shift' ELSE 'First Shift' END AS shift FROM attendance a JOIN employees e ON e.id=a.employee_id WHERE " + " AND ".join(clauses) + " ORDER BY a.work_date DESC,e.staff_id"
    with get_db() as c:
        return c.execute(sql, params).fetchall()

def _payroll_rows(month: str):
    with get_db() as c:
        return c.execute("""SELECT p.*,e.staff_id,e.name,e.department,e.designation
            FROM payroll_records p JOIN employees e ON e.id=p.employee_id
            WHERE p.salary_month=? ORDER BY e.staff_id""",(month,)).fetchall()

def _duty_duration_minutes(start_time: str, end_time: str) -> int:
    start=datetime.strptime(start_time,'%H:%M')
    end=datetime.strptime(end_time,'%H:%M')
    if end<=start: end+=timedelta(days=1)
    return int((end-start).total_seconds()//60)

def _validated_break_minutes(start_time: str, end_time: str, break_minutes: int) -> int:
    if not re.fullmatch(r'\d{2}:\d{2}',start_time or '') or not re.fullmatch(r'\d{2}:\d{2}',end_time or ''):
        raise HTTPException(400,'Invalid duty time')
    minutes=int(break_minutes or 0)
    if minutes<0 or minutes>=_duty_duration_minutes(start_time,end_time):
        raise HTTPException(400,'Break must be shorter than the duty time')
    return minutes

def _payroll_duty_metrics(employee_id: int, month: str):
    first=datetime.strptime(month+'-01','%Y-%m-%d').date(); next_month=(first.replace(day=28)+timedelta(days=4)).replace(day=1); last=next_month-timedelta(days=1)
    today=datetime.now(ZoneInfo(settings.timezone)).date()
    effective_last=min(last,today) if first<=today<=last else last
    if today<first: effective_last=first-timedelta(days=1)
    with get_db() as c:
        weekly=c.execute("SELECT * FROM duty_schedules WHERE employee_id=? AND is_active",(employee_id,)).fetchall()
        custom=c.execute("SELECT * FROM custom_duties WHERE employee_id=? AND duty_date>=? AND duty_date<=? AND is_active",(employee_id,first.isoformat(),effective_last.isoformat())).fetchall() if effective_last>=first else []
        attendance=c.execute("SELECT work_date,check_in,check_out,status,late_minutes FROM attendance WHERE employee_id=? AND work_date>=? AND work_date<=?",(employee_id,first.isoformat(),effective_last.isoformat())).fetchall() if effective_last>=first else []
        leaves=c.execute("SELECT leave_type,start_date,end_date FROM leave_requests WHERE employee_id=? AND status='approved' AND start_date<=? AND end_date>=?",(employee_id,effective_last.isoformat(),first.isoformat())).fetchall() if effective_last>=first else []
    weekly_by_day={int(r['weekday']):r for r in weekly}
    custom_by_date={r['duty_date']:r for r in custom}
    scheduled={}; day=first
    while day<=effective_last:
        duty=custom_by_date.get(day.isoformat()) or weekly_by_day.get(day.weekday())
        if duty:
            duration=_duty_duration_minutes(duty['start_time'],duty['end_time'])
            scheduled[day.isoformat()]=max(duration-int(duty['break_minutes'] or 0),1)
        day+=timedelta(days=1)
    attendance_by_date={r['work_date']:r for r in attendance if r['work_date'] in scheduled}; worked_units=0.0; incomplete=[]; late_minutes=0; late_fraction_units=0.0
    for work_date,row in attendance_by_date.items():
        if not row['check_in'] or not row['check_out']:
            incomplete.append(work_date)
            continue
        status=str(row['status'] or '').lower(); unit=0.5 if status in {'half_day','half-day','half day'} else 1.0
        worked_units += unit
        daily_late=max(int(row['late_minutes'] or 0),0)
        daily_late=min(daily_late,scheduled[work_date])
        late_minutes+=daily_late
        late_fraction_units+=min(daily_late/scheduled[work_date],unit)
    paid_leave_dates=set(); unpaid_leave_dates=set()
    for leave in leaves:
        day=max(datetime.fromisoformat(leave['start_date']).date(),first); end=min(datetime.fromisoformat(leave['end_date']).date(),effective_last)
        while day<=end:
            if day.isoformat() in scheduled and (day.isoformat() not in attendance_by_date or day.isoformat() in incomplete):
                leave_name=str(leave['leave_type'] or '').strip().lower()
                target=unpaid_leave_dates if leave_name in {'unpaid','unpaid leave','lwp','leave without pay','without pay'} else paid_leave_dates
                target.add(day.isoformat())
            day+=timedelta(days=1)
    scheduled_units=float(len(scheduled)); paid_units=float(len(paid_leave_dates)); unpaid_units=float(len(unpaid_leave_dates)); absent_units=max(scheduled_units-worked_units-paid_units-unpaid_units,0)
    return {"scheduled":scheduled_units,"worked":worked_units,"paid_leave":paid_units,"unpaid_leave":unpaid_units,"absent":absent_units,"late_minutes":late_minutes,"late_fraction_units":late_fraction_units,"payable_duty_minutes":sum(scheduled.values()),"incomplete_dates":incomplete}

def _calculate_employee_payroll(employee_id: int, month: str, fixed_salary: float, overtime_rate: float, overtime_mode: str="manual", manual_overtime_hours: float=0, bonus: float=0, advance: float=0, fine: float=0, deduction: float=0):
    duty=_payroll_duty_metrics(employee_id,month)
    per_day=(float(fixed_salary or 0)/duty['scheduled']) if duty['scheduled'] else 0
    late_deduction=per_day*duty['late_fraction_units']
    result=calculate_payroll(PayrollInput(fixed_salary=fixed_salary,scheduled_units=duty['scheduled'],worked_units=duty['worked'],paid_leave_units=duty['paid_leave'],unpaid_leave_units=duty['unpaid_leave'],late_minutes=duty['late_minutes'],late_deduction=late_deduction,payable_duty_minutes=duty['payable_duty_minutes'],overtime_hours=manual_overtime_hours,overtime_rate=overtime_rate,bonus=bonus,advance=advance,fine=fine,other_deduction=deduction))
    result['incomplete_dates']=duty['incomplete_dates']; result['overtime_mode']='manual'
    return result

def _salary_sheet_rows(month: str):
    with get_db() as c:
        rows=c.execute("""SELECT e.id employee_id,e.staff_id,e.name,e.department,e.designation,
            p.id payroll_id,COALESCE(p.fixed_salary,e.fixed_salary,0) fixed_salary,p.overtime_hours,
            COALESCE(p.overtime_rate,e.default_overtime_rate,0) overtime_rate,p.overtime_amount,p.bonus,p.deduction,
            p.advance_amount,p.fine_amount,p.overtime_mode,p.adjustment_reason,p.payment_method,p.payment_reference,p.payment_status,p.calculation_snapshot,p.note
            FROM employees e LEFT JOIN payroll_records p ON p.employee_id=e.id AND p.salary_month=?
            WHERE e.is_active ORDER BY e.staff_id""",(month,)).fetchall()
    output=[]
    for row in rows:
        item=dict(row); item['id']=item['payroll_id']; fixed=float(row['fixed_salary'] or 0); rate=float(row['overtime_rate'] or 0); mode=str(row['overtime_mode'] or 'auto'); manual_hours=float(row['overtime_hours'] or 0) if mode=='manual' else 0
        if row['payroll_id'] and row['payment_status'] in {'finalized','paid'} and row['calculation_snapshot']:
            try: calculated=json.loads(row['calculation_snapshot'])
            except Exception: calculated=_calculate_employee_payroll(row['employee_id'],month,fixed,rate,mode,manual_hours,float(row['bonus'] or 0),float(row['advance_amount'] or 0),float(row['fine_amount'] or 0),float(row['deduction'] or 0))
        else: calculated=_calculate_employee_payroll(row['employee_id'],month,fixed,rate,mode,manual_hours,float(row['bonus'] or 0),float(row['advance_amount'] or 0),float(row['fine_amount'] or 0),float(row['deduction'] or 0))
        calculated.setdefault('earned_basic_salary',max(fixed-float(calculated.get('absent_deduction') or 0)-float(calculated.get('unpaid_leave_deduction') or 0),0))
        calculated.setdefault('late_minutes',0); calculated.setdefault('late_deduction',0); calculated.setdefault('payable_duty_minutes',0)
        item.update(calculated)
        output.append(item)
    return output

def _money(value):
    return f"{float(value or 0):,.2f}"

def _payroll_actor(request: Request) -> str:
    return str(request.session.get('user_name') or request.session.get('hr_id') or 'Super Admin')

def _log_payroll_change(db, payroll_id: int, action: str, actor: str, reason: str=""):
    row=db.execute("SELECT * FROM payroll_records WHERE id=?",(payroll_id,)).fetchone()
    db.execute("INSERT INTO payroll_change_logs(payroll_id,action,actor,reason,snapshot) VALUES(?,?,?,?,?)",(payroll_id,action,actor,reason,json.dumps(dict(row),default=str) if row else '{}'))

PAYROLL_STAGES = {
    "not_prepared": ("Not prepared", "stage-none",
                     "No payslip exists yet for this month."),
    "draft":        ("Draft", "stage-draft",
                     "Saved, still editable. Nothing is committed."),
    "finalized":    ("Finalized", "stage-final",
                     "Locked. Only a Super Admin can reopen it."),
    "paid":         ("Paid", "stage-paid",
                     "Payment recorded with a method and reference."),
}


@app.get("/payroll/preview")
def payroll_preview(request: Request, employee_id: int, month: str,
                    fixed_salary: float = -1, overtime_rate: float = -1,
                    overtime_hours: float = 0, bonus: float = 0,
                    advance: float = 0, fine: float = 0, deduction: float = 0):
    """Live calculation for the payroll form.

    Previously this endpoint existed but nothing ever called it, and it
    accepted no adjustment values — so the form's 'Preview' label was a
    promise it could not keep. The extra arguments are all optional and
    default to the employee's saved figures, so old callers behave
    exactly as before.
    """
    require_permission(request, "payroll_view")
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise HTTPException(400, "Invalid salary month")
    with get_db() as c:
        employee = c.execute(
            "SELECT staff_id,name,fixed_salary,default_overtime_rate "
            "FROM employees WHERE id=? AND is_active", (employee_id,)).fetchone()
    if not employee:
        raise HTTPException(404, "Employee not found")
    if fixed_salary < 0:
        fixed_salary = float(employee["fixed_salary"] or 0)
    if overtime_rate < 0:
        overtime_rate = float(employee["default_overtime_rate"] or 0)
    for value in (fixed_salary, overtime_rate, overtime_hours, bonus, advance, fine, deduction):
        if value < 0:
            raise HTTPException(400, "Payroll values cannot be negative")
    result = _calculate_employee_payroll(
        employee_id, month, fixed_salary, overtime_rate, "manual",
        overtime_hours, bonus, advance, fine, deduction)
    result["staff_id"] = employee["staff_id"]
    result["employee_name"] = employee["name"]
    return result


@app.get("/payroll", response_class=HTMLResponse)
def payroll_page(request: Request, month: str = "", saved: str = "", error: str = "",
                 made: int = 0, skipped: int = 0):
    require_permission(request, "payroll_view")
    current = datetime.now(ZoneInfo(settings.timezone)).strftime("%Y-%m")
    month = month or current
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise HTTPException(400, "Invalid salary month")

    month_label = datetime.strptime(month, "%Y-%m").strftime("%B %Y")
    can_manage = has_permission(request, "payroll_manage")
    can_export = has_permission(request, "payroll_export")
    is_super = request.session.get("role") == "super_admin"

    # Every active employee, including those with no payslip yet. The old
    # page dropped them with `if row['payroll_id']`, so the one question a
    # payroll officer actually has — "who is still left?" — was unanswerable.
    all_rows = _salary_sheet_rows(month)
    prepared = [r for r in all_rows if r["payroll_id"]]
    missing = [r for r in all_rows if not r["payroll_id"]]

    def _stage(row):
        if not row["payroll_id"]:
            return "not_prepared"
        return str(row["payment_status"] or "draft")

    counts = {k: 0 for k in PAYROLL_STAGES}
    for r in all_rows:
        counts[_stage(r)] = counts.get(_stage(r), 0) + 1

    total_staff = len(all_rows)
    net_total = sum(float(r["net_salary"] or 0) for r in prepared)
    paid_total = sum(float(r["net_salary"] or 0) for r in prepared
                     if r["payment_status"] == "paid")
    outstanding = net_total - paid_total

    # Employees who have no basic salary on record. Preparing a payslip for
    # them can only ever produce 0.00, which is what makes a bulk prepare
    # look broken.
    no_salary = [r for r in all_rows if float(r["fixed_salary"] or 0) <= 0]

    # ---------------- notice ----------------
    if saved == "discard":
        notices = (f"<div class='notice notice-ok'>Discarded {made} draft payslip"
                   f"{'' if made == 1 else 's'}. Finalized and paid payslips were "
                   f"left untouched.</div>" if made else
                   "<div class='notice notice-ok'>Draft payslip discarded.</div>")
    elif saved == "bulk":
        notices = f"<div class='notice notice-ok'>Prepared {made} payslip{'' if made == 1 else 's'}."
        if skipped:
            notices += (f" Skipped {skipped} employee{'' if skipped == 1 else 's'} with no basic "
                        f"salary on record — set their salary first, then prepare again.")
        notices += "</div>"
    elif saved:
        notices = ("<div class='notice notice-ok'>Payslip saved as a draft. "
                   "Check the figures below, then finalize it to lock.</div>")
    elif error == "confirm":
        notices = ("<div class='notice notice-bad'>Nothing was discarded. You must type "
                   "DISCARD to confirm.</div>")
    elif error == "reason":
        notices = ("<div class='notice notice-bad'>A reason is required whenever you "
                   "add a bonus, advance, fine or other deduction.</div>")
    elif error:
        notices = ("<div class='notice notice-bad'>The payslip could not be saved. "
                   "Check that no amount is negative and try again.</div>")
    else:
        notices = ""

    # ---------------- progress strip ----------------
    def _seg(key):
        return _pct(counts.get(key, 0), total_staff)

    progress = (
        f"<div class='card payroll-progress'>"
        f"<div class='card-head'><div>"
        f"<h3>{len(prepared)} of {total_staff} payslips prepared</h3>"
        f"<div class='sub'>{month_label} · {counts['paid']} paid, "
        f"{counts['finalized']} locked, {counts['draft']} still editable</div>"
        f"</div>"
        + (f"<form method='post' action='/payroll/bulk-prepare'>"
           f"<input type='hidden' name='month' value='{month}'>"
           f"<button class='btn'>Prepare the remaining {len(missing)}</button></form>"
           if can_manage and missing else "")
        + (f"<details class='pay-details'><summary class='btn secondary'>Discard {counts['draft']} draft"
           f"{'' if counts['draft'] == 1 else 's'}</summary>"
           f"<form method='post' action='/payroll/bulk-discard' class='pay-form'>"
           f"<input type='hidden' name='month' value='{month}'>"
           f"<div class='hint'>This deletes every draft payslip for {month_label}. "
           f"Finalized and paid payslips are not touched. It cannot be undone.</div>"
           f"<label>Type DISCARD to confirm</label>"
           f"<input name='confirm' placeholder='DISCARD' required>"
           f"<button class='btn danger'>Discard drafts</button></form></details>"
           if can_manage and counts["draft"] else "")
        + "</div>"
        f"<div class='stage-bar' role='img' aria-label='"
        f"{counts['paid']} paid, {counts['finalized']} finalized, "
        f"{counts['draft']} draft, {counts['not_prepared']} not prepared'>"
        f"<span class='seg-paid'  style='width:{_seg('paid')}%'></span>"
        f"<span class='seg-final' style='width:{_seg('finalized')}%'></span>"
        f"<span class='seg-draft' style='width:{_seg('draft')}%'></span>"
        f"<span class='seg-none'  style='width:{_seg('not_prepared')}%'></span>"
        f"</div>"
        f"<div class='stage-key'>"
        + "".join(
            f"<span><i class='seg-dot {cls}'></i>{label} <b>{counts.get(key, 0)}</b></span>"
            for key, (label, cls, _tip) in PAYROLL_STAGES.items())
        + "</div></div>"
    )

    # ---------------- money summary ----------------
    summary = (
        "<div class='payroll-summary'>"
        f"<div class='card'><div class='metric-label'>Net payroll</div>"
        f"<div class='metric'>৳{_money(net_total)}</div>"
        f"<div class='metric-foot'>Total payable for {month_label}</div></div>"
        f"<div class='card'><div class='metric-label'>Already paid</div>"
        f"<div class='metric'>৳{_money(paid_total)}</div>"
        f"<div class='metric-foot'>{counts['paid']} of {len(prepared)} payslips</div></div>"
        f"<div class='card'><div class='metric-label'>Still outstanding</div>"
        f"<div class='metric'>৳{_money(outstanding)}</div>"
        f"<div class='metric-foot'>"
        + ("Everything is paid" if outstanding <= 0 else "Waiting to be disbursed")
        + "</div></div></div>"
    )

    # ---------------- entry form with live breakdown ----------------
    form = ""
    if can_manage:
        with get_db() as c:
            employees = c.execute(
                "SELECT id,staff_id,name,COALESCE(fixed_salary,0) fixed_salary,"
                "COALESCE(default_overtime_rate,0) ot_rate "
                "FROM employees WHERE is_active ORDER BY staff_id").fetchall()
        options = "".join(
            f"<option value='{e['id']}' data-salary='{float(e['fixed_salary'] or 0):.2f}' "
            f"data-ot='{float(e['ot_rate'] or 0):.2f}'>"
            f"{escape(e['staff_id'])} — {escape(e['name'])}</option>"
            for e in employees)
        form = f"""
        <div class='card'>
          <div class='card-head'><div><h3>Prepare a payslip</h3>
          <div class='sub'>Pick a person, adjust the amounts, and watch the
          breakdown on the right update as you type.</div></div></div>
          <form method='post' action='/payroll' id='payroll-form'>
            <input type='hidden' name='return_month' value='{month}'>
            <input type='hidden' name='overtime_mode' value='manual'>

            <label for='pf-emp'>Employee</label>
            <select id='pf-emp' name='employee_id' required>{options}</select>

            <label for='pf-month'>Salary month</label>
            <input id='pf-month' type='month' name='salary_month' value='{month}' required>

            <div class='field-group'>
              <div class='field-group-title'>Salary</div>
              <div class='two'>
                <div><label for='pf-basic'>Monthly basic salary</label>
                <input id='pf-basic' type='number' min='0' step='0.01' name='fixed_salary' required>
                <div class='hint'>Filled from the employee record. Editing it updates their record too.</div></div>
                <div><label for='pf-otrate'>Overtime rate per hour</label>
                <input id='pf-otrate' type='number' min='0' step='0.01' name='overtime_rate' value='0'></div>
              </div>
              <label for='pf-othours'>Overtime hours</label>
              <input id='pf-othours' type='number' min='0' step='0.01' name='overtime_hours' value='0'>
              <div class='hint'>Entered by hand. Overtime is never added automatically.</div>
            </div>

            <div class='field-group'>
              <div class='field-group-title'>Adds to the salary</div>
              <label for='pf-bonus'>Bonus</label>
              <input id='pf-bonus' type='number' min='0' step='0.01' name='bonus' value='0'>
            </div>

            <div class='field-group'>
              <div class='field-group-title'>Cut from the salary</div>
              <div class='two'>
                <div><label for='pf-advance'>Salary advance already taken</label>
                <input id='pf-advance' type='number' min='0' step='0.01' name='advance' value='0'></div>
                <div><label for='pf-fine'>Fine</label>
                <input id='pf-fine' type='number' min='0' step='0.01' name='fine' value='0'></div>
              </div>
              <label for='pf-other'>Other deduction</label>
              <input id='pf-other' type='number' min='0' step='0.01' name='deduction' value='0'>
              <div class='hint'>Late-minute cuts are calculated from attendance — do not enter them here.</div>
            </div>

            <label for='pf-reason'>Reason for the adjustment</label>
            <input id='pf-reason' name='adjustment_reason'
                   placeholder='Required if you entered a bonus, advance, fine or other deduction'>

            <label for='pf-note'>Private HR note</label>
            <textarea id='pf-note' name='note'></textarea>

            <button class='btn' type='submit'>Save as draft</button>
            <div class='hint'>Saving does not pay anyone. You finalize and mark
            as paid from the salary sheet below.</div>
          </form>
        </div>"""

    # ---------------- live breakdown panel ----------------
    breakdown = f"""
    <div class='card' id='payroll-breakdown'>
      <div class='card-head'><div><h3>How this salary is worked out</h3>
      <div class='sub' id='pb-who'>Select an employee to see their figures.</div></div></div>
      <div class='calc-sheet' id='pb-body' aria-live='polite'>
        <div class='empty-cell'>Waiting for an employee…</div>
      </div>
    </div>"""

    # ---------------- salary sheet ----------------
    table = []
    for r in all_rows:
        stage = _stage(r)
        label, cls, _tip = PAYROLL_STAGES[stage]
        scheduled = float(r["scheduled"] or 0)
        worked = float(r["worked"] or 0)
        duty_pct = _pct(worked, scheduled)

        if stage == "not_prepared":
            actions = (f"<span class='sub'>Use the form to prepare</span>")
            net_cell = "<span class='sub'>—</span>"
        else:
            net_cell = f"<b>৳{_money(r['net_salary'])}</b>"
            bits = []
            if can_manage and stage == "draft":
                bits.append(
                    f"<form method='post' action='/payroll/{r['id']}/status'>"
                    f"<input type='hidden' name='month' value='{month}'>"
                    f"<input type='hidden' name='status' value='finalized'>"
                    f"<button class='btn small'>Finalize</button></form>")
                bits.append(
                    f"<form method='post' action='/payroll/{r['id']}/discard'>"
                    f"<input type='hidden' name='month' value='{month}'>"
                    f"<button class='btn small secondary' title='Delete this draft payslip'>Discard</button></form>")
            if can_manage and stage == "finalized":
                bits.append(
                    f"<details class='pay-details'><summary class='btn small'>Mark as paid</summary>"
                    f"<form method='post' action='/payroll/{r['id']}/status' class='pay-form'>"
                    f"<input type='hidden' name='month' value='{month}'>"
                    f"<input type='hidden' name='status' value='paid'>"
                    f"<label>How it was paid</label>"
                    f"<input name='payment_method' placeholder='Cash, bKash, bank transfer…' required>"
                    f"<label>Reference number</label>"
                    f"<input name='payment_reference' placeholder='Transaction or voucher no.' required>"
                    f"<button class='btn small'>Confirm payment</button></form></details>")
            if is_super and stage == "finalized":
                bits.append(
                    f"<details class='pay-details'><summary class='btn small secondary'>Reopen</summary>"
                    f"<form method='post' action='/payroll/{r['id']}/reopen' class='pay-form'>"
                    f"<input type='hidden' name='month' value='{month}'>"
                    f"<label>Why is this being reopened?</label>"
                    f"<input name='reason' required>"
                    f"<button class='btn small secondary'>Unlock for editing</button></form></details>")
            if can_export:
                bits.append(f"<a class='btn small secondary' "
                            f"href='/payroll/{r['id']}/payslip.pdf'>Payslip</a>")
            actions = "".join(bits) or "<span class='sub'>—</span>"

        table.append(
            f"<tr class='row-{stage}'>"
            f"<td><b>{escape(str(r['name']))}</b>"
            f"<div class='sub'>{escape(str(r['staff_id']))}</div></td>"
            f"<td><span class='status-badge {cls}' title='{escape(_tip)}'>{label}</span></td>"
            f"<td class='num'>{worked:g} / {scheduled:g}"
            f"<div class='mini-line'><span class='mini-present' style='width:{duty_pct}%'></span></div></td>"
            f"<td class='num'>৳{_money(r['earned_basic_salary'])}"
            f"<div class='sub'>of ৳{_money(r['fixed_salary'])}</div></td>"
            f"<td class='num'>"
            + (f"+৳{_money(r['overtime_amount'])}" if float(r['overtime_amount'] or 0) else "—")
            + f"</td>"
            f"<td class='num'>"
            + (f"−৳{_money(r['total_deduction'])}" if float(r['total_deduction'] or 0) else "—")
            + f"<div class='sub'>"
            + (f"{int(r['late_minutes'] or 0)} min late" if int(r['late_minutes'] or 0) else "&nbsp;")
            + f"</div></td>"
            f"<td class='num'>{net_cell}</td>"
            f"<td class='cell-actions'>{actions}</td></tr>")

    export_buttons = ""
    if can_export:
        export_buttons += (f"<a class='btn secondary' href='/payroll/export.xlsx?month={month}'>Excel</a>"
                           f"<a class='btn secondary' href='/payroll/export.pdf?month={month}'>PDF</a>")
    if is_super:
        export_buttons += "<a class='btn secondary' href='/settings/payroll-backup'>Backup</a>"

    salary_warning = ""
    if no_salary:
        names = ", ".join(escape(str(r["name"])) for r in no_salary[:6])
        more = f" and {len(no_salary) - 6} more" if len(no_salary) > 6 else ""
        salary_warning = (
            f"<div class='card warn-card'><b>{len(no_salary)} employee"
            f"{'' if len(no_salary) == 1 else 's'} have no basic salary on record.</b>"
            f"<div class='sub'>{names}{more}. Their payslips can only come out as ৳0.00. "
            f"Set a monthly basic salary on each employee first — either in the form below "
            f"or from their profile page.</div></div><div class='section-gap'></div>")

    body = f"""
    {notices}
    <div class='dashboard-head'>
      <div class='dashboard-greet'>
        <h1>Salary &amp; payroll</h1>
        <div class='dashboard-date'><span>{month_label}</span>
        <span class='sub'>{total_staff} active employees</span></div>
      </div>
      <div class='dashboard-tools'>
        <form method='get' class='month-picker'>
          <input type='month' name='month' value='{month}' aria-label='Salary month'>
          <button class='btn secondary'>Open month</button>
        </form>
        {export_buttons}
      </div>
    </div>

    {salary_warning}
    {progress}
    <div class='section-gap'></div>
    {summary}
    <div class='section-gap'></div>

    <div class='payroll-work'>{form}{breakdown}</div>

    <div class='section-gap'></div>
    <div class='card'>
      <div class='card-head'><div><h3>Salary sheet — {month_label}</h3>
      <div class='sub'>Net = earned basic + overtime + bonus − (late + advance + fine + other)</div></div></div>
      <div class='table-scroll'>
        <table class='dashboard-table payroll-table'>
          <thead><tr>
            <th>Employee</th><th>Stage</th><th class='num'>Duty done</th>
            <th class='num'>Earned basic</th><th class='num'>Overtime</th>
            <th class='num'>Deductions</th><th class='num'>Net payable</th><th>Action</th>
          </tr></thead>
          <tbody>{''.join(table) or "<tr><td colspan='8'><div class='empty-cell'>No active employees.</div></td></tr>"}</tbody>
        </table>
      </div>
    </div>

    <script>
    (function () {{
      var form = document.getElementById('payroll-form');
      if (!form) return;
      var emp = document.getElementById('pf-emp');
      var basic = document.getElementById('pf-basic');
      var otRate = document.getElementById('pf-otrate');
      var body = document.getElementById('pb-body');
      var who = document.getElementById('pb-who');
      var timer = null, seq = 0;

      function taka(n) {{
        return '৳' + Number(n || 0).toLocaleString('en-US',
          {{ minimumFractionDigits: 2, maximumFractionDigits: 2 }});
      }}
      function val(id) {{
        var el = document.getElementById(id);
        var n = parseFloat(el && el.value);
        return isFinite(n) && n >= 0 ? n : 0;
      }}
      function line(label, note, amount, kind) {{
        return '<div class="calc-row ' + (kind || '') + '"><span>' + label +
               (note ? '<i>' + note + '</i>' : '') + '</span><b>' + amount + '</b></div>';
      }}

      // Fill basic salary and OT rate from the chosen employee, so the
      // officer is not retyping figures the system already knows.
      function fillFromEmployee() {{
        var opt = emp.options[emp.selectedIndex];
        if (!opt) return;
        basic.value = opt.dataset.salary || '0';
        otRate.value = opt.dataset.ot || '0';
      }}

      function render(d) {{
        who.textContent = d.employee_name + ' · ' + d.staff_id;
        var rows = '';
        rows += line('Earned basic',
                     d.worked + ' of ' + d.scheduled + ' duty units at ' +
                     taka(d.per_day_salary) + ' each',
                     taka(d.earned_basic_salary));
        if (d.paid_leave > 0) rows += line('Includes paid leave', d.paid_leave + ' units', '', 'muted');
        if (d.absent > 0) rows += line('Absent', d.absent + ' units not earned',
                                       '−' + taka(d.absent_deduction), 'muted');
        if (d.overtime_amount > 0)
          rows += line('Overtime', d.overtime_hours + 'h at ' + taka(d.overtime_rate) + '/h',
                       '+' + taka(d.overtime_amount), 'plus');
        if (d.bonus > 0) rows += line('Bonus', '', '+' + taka(d.bonus), 'plus');
        rows += line('Gross', '', taka(d.gross_salary), 'subtotal');
        if (d.late_deduction > 0)
          rows += line('Late arrival', Math.round(d.late_minutes) + ' minutes this month',
                       '−' + taka(d.late_deduction), 'minus');
        if (d.advance > 0) rows += line('Advance already taken', '', '−' + taka(d.advance), 'minus');
        if (d.fine > 0) rows += line('Fine', '', '−' + taka(d.fine), 'minus');
        if (d.deduction > 0) rows += line('Other deduction', '', '−' + taka(d.deduction), 'minus');
        if (d.total_deduction > 0) rows += line('Total deductions', '', '−' + taka(d.total_deduction), 'subtotal');
        rows += line('Net payable', '', taka(d.net_salary), 'total');
        if (d.incomplete_dates && d.incomplete_dates.length)
          rows += '<div class="calc-warn">' + d.incomplete_dates.length +
                  ' day(s) have a check-in but no check-out. Fix those before finalizing.</div>';
        body.innerHTML = rows;
      }}

      function refresh() {{
        if (!emp.value) return;
        var mine = ++seq;
        var q = new URLSearchParams({{
          employee_id: emp.value,
          month: (document.getElementById('pf-month').value || '{month}'),
          fixed_salary: val('pf-basic'),
          overtime_rate: val('pf-otrate'),
          overtime_hours: val('pf-othours'),
          bonus: val('pf-bonus'),
          advance: val('pf-advance'),
          fine: val('pf-fine'),
          deduction: val('pf-other')
        }});
        fetch('/payroll/preview?' + q.toString(), {{ credentials: 'same-origin' }})
          .then(function (r) {{ if (!r.ok) throw new Error(r.status); return r.json(); }})
          .then(function (d) {{ if (mine === seq) render(d); }})
          .catch(function () {{
            if (mine === seq)
              body.innerHTML = '<div class="calc-warn">Could not calculate right now. ' +
                               'The figures will still be correct when you save.</div>';
          }});
      }}

      function debounced() {{ clearTimeout(timer); timer = setTimeout(refresh, 280); }}

      emp.addEventListener('change', function () {{ fillFromEmployee(); refresh(); }});
      document.getElementById('pf-month').addEventListener('change', refresh);
      ['pf-basic','pf-otrate','pf-othours','pf-bonus','pf-advance','pf-fine','pf-other']
        .forEach(function (id) {{
          var el = document.getElementById(id);
          if (el) el.addEventListener('input', debounced);
        }});

      // Block submit when an adjustment has no reason, instead of letting the
      // server bounce the officer back and lose everything they typed.
      form.addEventListener('submit', function (e) {{
        var needs = val('pf-bonus') + val('pf-advance') + val('pf-fine') + val('pf-other');
        var reason = document.getElementById('pf-reason');
        if (needs > 0 && !reason.value.trim()) {{
          e.preventDefault();
          reason.focus();
          reason.setAttribute('aria-invalid', 'true');
        }}
      }});

      fillFromEmployee();
      refresh();
    }})();
    </script>
    """
    return layout("Payroll", body, request, "payroll")


@app.post("/payroll")
def save_payroll(request: Request, employee_id: int=Form(...), salary_month: str=Form(...), fixed_salary: float=Form(...), overtime_hours: float=Form(0), overtime_rate: float=Form(0), overtime_mode: str=Form("manual"), bonus: float=Form(0), advance: float=Form(0), fine: float=Form(0), deduction: float=Form(0), adjustment_reason: str=Form(""), note: str=Form(""), return_month: str=Form(""), profile_employee_id: int=Form(0)):
    require_permission(request,"payroll_manage")
    values=(fixed_salary,overtime_hours,overtime_rate,bonus,advance,fine,deduction); overtime_mode='manual'
    if not re.fullmatch(r"\d{4}-\d{2}",salary_month) or any(v<0 for v in values): return RedirectResponse(f"/payroll?month={return_month or salary_month}&error=1",303)
    if adjustment_reason_required(bonus,advance,fine,deduction) and not adjustment_reason.strip(): return RedirectResponse(f"/payroll?month={salary_month}&error=reason",303)
    actor=_payroll_actor(request); calc=_calculate_employee_payroll(employee_id,salary_month,fixed_salary,overtime_rate,overtime_mode,overtime_hours,bonus,advance,fine,deduction)
    with get_db() as c:
        existing=c.execute("SELECT id,payment_status FROM payroll_records WHERE employee_id=? AND salary_month=?",(employee_id,salary_month)).fetchone()
        if existing and existing['payment_status'] in {'finalized','paid'}: raise HTTPException(409,"Finalized payroll is locked. Super Admin must reopen it first.")
        c.execute("UPDATE employees SET fixed_salary=?,default_overtime_rate=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",(fixed_salary,overtime_rate,employee_id))
        payload=(fixed_salary,calc['overtime_hours'],overtime_rate,calc['overtime_amount'],bonus,deduction,calc['net_salary'],note.strip(),actor,int(calc['scheduled']),int(calc['worked']),int(calc['paid_leave']),int(calc['absent']),calc['absent_deduction'],calc['worked'],calc['paid_leave'],calc['unpaid_leave'],calc['absent'],calc['unpaid_leave_deduction'],calc['earned_basic_salary'],int(calc['late_minutes']),calc['late_deduction'],int(calc['payable_duty_minutes']),advance,fine,calc['gross_salary'],calc['total_deduction'],overtime_mode,adjustment_reason.strip(),json.dumps(calc,default=str))
        if existing:
            c.execute("""UPDATE payroll_records SET fixed_salary=?,overtime_hours=?,overtime_rate=?,overtime_amount=?,bonus=?,deduction=?,net_salary=?,note=?,updated_by=?,scheduled_duty_days=?,worked_duty_days=?,paid_leave_days=?,absent_days=?,absent_deduction=?,worked_duty_units=?,paid_leave_units=?,unpaid_leave_units=?,absent_duty_units=?,unpaid_leave_deduction=?,earned_basic_salary=?,late_minutes=?,late_deduction=?,payable_duty_minutes=?,advance_amount=?,fine_amount=?,gross_salary=?,total_deduction=?,overtime_mode=?,adjustment_reason=?,calculation_snapshot=?,payment_status='draft',updated_at=CURRENT_TIMESTAMP WHERE id=?""",payload+(existing['id'],)); payroll_id=existing['id']
        else:
            insert_values=(employee_id,salary_month)+payload[:9]+(actor,)+payload[9:]
            c.execute("""INSERT INTO payroll_records(employee_id,salary_month,fixed_salary,overtime_hours,overtime_rate,overtime_amount,bonus,deduction,net_salary,note,created_by,updated_by,scheduled_duty_days,worked_duty_days,paid_leave_days,absent_days,absent_deduction,worked_duty_units,paid_leave_units,unpaid_leave_units,absent_duty_units,unpaid_leave_deduction,earned_basic_salary,late_minutes,late_deduction,payable_duty_minutes,advance_amount,fine_amount,gross_salary,total_deduction,overtime_mode,adjustment_reason,calculation_snapshot,payment_status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'draft')""",insert_values); payroll_id=c.execute("SELECT id FROM payroll_records WHERE employee_id=? AND salary_month=?",(employee_id,salary_month)).fetchone()['id']
        _log_payroll_change(c,payroll_id,"saved",actor,adjustment_reason.strip())
    audit(request,"save","payroll",f"{employee_id}:{salary_month}",f"Net salary: {calc['net_salary']:.2f}")
    if profile_employee_id==employee_id: return RedirectResponse(f"/employees/{employee_id}?month={salary_month}#payroll",303)
    return RedirectResponse(f"/payroll?month={salary_month}&saved=1",303)

@app.post("/payroll/{payroll_id}/status")
def payroll_status(request: Request, background_tasks: BackgroundTasks, payroll_id: int, status: str=Form(...), month: str=Form(...), payment_method: str=Form(""), payment_reference: str=Form(""), return_employee_id: int=Form(0)):
    require_permission(request,"payroll_manage")
    if status not in {"finalized","paid"}: raise HTTPException(400,"Invalid payroll status")
    actor=_payroll_actor(request)
    with get_db() as c:
        row=c.execute("SELECT * FROM payroll_records WHERE id=?",(payroll_id,)).fetchone()
        if not row: raise HTTPException(404,"Payroll not found")
        if status=='finalized':
            if row['payment_status']!='draft': raise HTTPException(409,"Only draft payroll can be finalized")
            snapshot=json.loads(row['calculation_snapshot'] or '{}')
            if float(row['fixed_salary'] or 0)<=0: raise HTTPException(409,"Basic Salary is missing")
            if float(snapshot.get('scheduled') or 0)<=0: raise HTTPException(409,"No scheduled duty found for this month")
            if float(snapshot.get('net_salary') or 0)<0: raise HTTPException(409,"Net salary cannot be negative")
            if snapshot.get('incomplete_dates'): raise HTTPException(409,"Incomplete checkout must be reviewed before finalizing")
            c.execute("UPDATE payroll_records SET payment_status='finalized',finalized_at=CURRENT_TIMESTAMP,locked_at=CURRENT_TIMESTAMP,locked_by=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",(actor,payroll_id))
        else:
            if row['payment_status']!='finalized': raise HTTPException(409,"Finalize payroll before payment")
            if not payment_method.strip() or not payment_reference.strip(): raise HTTPException(400,"Payment method and reference are required")
            c.execute("UPDATE payroll_records SET payment_status='paid',payment_method=?,payment_reference=?,paid_at=CURRENT_TIMESTAMP,updated_at=CURRENT_TIMESTAMP WHERE id=?",(payment_method.strip(),payment_reference.strip(),payroll_id))
        _log_payroll_change(c,payroll_id,status,actor,payment_reference.strip())
        delivery=c.execute("SELECT p.*,e.staff_id,e.name,e.department,e.designation,e.whatsapp_phone,e.phone FROM payroll_records p JOIN employees e ON e.id=p.employee_id WHERE p.id=?",(payroll_id,)).fetchone() if status=='paid' else None
    audit(request,"payment_status","payroll",str(payroll_id),status)
    if delivery and (delivery['whatsapp_phone'] or delivery['phone']):
        pdf_bytes=_build_payslip_pdf(delivery); filename=f"BURAQ-Payslip-{delivery['staff_id']}-{delivery['salary_month']}.pdf"
        background_tasks.add_task(send_document_bytes,delivery['whatsapp_phone'] or delivery['phone'],pdf_bytes,filename,f"Salary payslip - {delivery['salary_month']}")
    if return_employee_id: return RedirectResponse(f"/employees/{return_employee_id}?month={month}#payroll",303)
    return RedirectResponse(f"/payroll?month={month}",303)

@app.post("/payroll/{payroll_id}/reopen")
def payroll_reopen(request: Request, payroll_id: int, month: str=Form(...), reason: str=Form(...)):
    require_super_admin(request)
    if len(reason.strip())<5: raise HTTPException(400,"Reopen reason is required")
    actor=_payroll_actor(request)
    with get_db() as c:
        row=c.execute("SELECT payment_status FROM payroll_records WHERE id=?",(payroll_id,)).fetchone()
        if not row: raise HTTPException(404,"Payroll not found")
        if row['payment_status']=='paid': raise HTTPException(409,"Paid payroll cannot be reopened")
        c.execute("UPDATE payroll_records SET payment_status='draft',reopened_at=CURRENT_TIMESTAMP,reopen_reason=?,locked_at=NULL,locked_by=NULL,updated_at=CURRENT_TIMESTAMP WHERE id=?",(reason.strip(),payroll_id)); _log_payroll_change(c,payroll_id,"reopened",actor,reason.strip())
    audit(request,"reopen","payroll",str(payroll_id),reason.strip())
    return RedirectResponse(f"/payroll?month={month}",303)

@app.post("/payroll/bulk-prepare")
def payroll_bulk_prepare(request: Request, month: str=Form(...)):
    require_permission(request,"payroll_manage")
    if not re.fullmatch(r"\d{4}-\d{2}",month): raise HTTPException(400,"Invalid salary month")
    actor=_payroll_actor(request); prepared=0; skipped=0
    with get_db() as c:
        employees=c.execute("SELECT id,fixed_salary,default_overtime_rate FROM employees WHERE is_active ORDER BY id").fetchall()
        for employee in employees:
            exists=c.execute("SELECT id FROM payroll_records WHERE employee_id=? AND salary_month=?",(employee['id'],month)).fetchone()
            if exists: continue
            # A payslip for someone with no basic salary is always 0.00 and only
            # creates clutter that then has to be cleaned up. Skip and report.
            if float(employee['fixed_salary'] or 0) <= 0:
                skipped += 1
                continue
            calc=_calculate_employee_payroll(employee['id'],month,float(employee['fixed_salary'] or 0),float(employee['default_overtime_rate'] or 0))
            c.execute("""INSERT INTO payroll_records(employee_id,salary_month,fixed_salary,overtime_hours,overtime_rate,overtime_amount,bonus,deduction,net_salary,created_by,updated_by,scheduled_duty_days,worked_duty_days,paid_leave_days,absent_days,absent_deduction,worked_duty_units,paid_leave_units,unpaid_leave_units,absent_duty_units,unpaid_leave_deduction,earned_basic_salary,late_minutes,late_deduction,payable_duty_minutes,advance_amount,fine_amount,gross_salary,total_deduction,overtime_mode,calculation_snapshot,payment_status) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,'draft')""",(employee['id'],month,calc['fixed_salary'],calc['overtime_hours'],calc['overtime_rate'],calc['overtime_amount'],0,0,calc['net_salary'],actor,actor,int(calc['scheduled']),int(calc['worked']),int(calc['paid_leave']),int(calc['absent']),calc['absent_deduction'],calc['worked'],calc['paid_leave'],calc['unpaid_leave'],calc['absent'],calc['unpaid_leave_deduction'],calc['earned_basic_salary'],int(calc['late_minutes']),calc['late_deduction'],int(calc['payable_duty_minutes']),0,0,calc['gross_salary'],calc['total_deduction'],'manual',json.dumps(calc,default=str))); prepared+=1
    audit(request,"bulk_prepare","payroll",month,f"Prepared {prepared} employee payrolls, skipped {skipped} with no basic salary")
    return RedirectResponse(f"/payroll?month={month}&saved=bulk&made={prepared}&skipped={skipped}",303)

@app.post("/payroll/{payroll_id}/discard")
def payroll_discard(request: Request, payroll_id: int, month: str=Form(...)):
    """Delete a single DRAFT payslip.

    Until now nothing in the application could remove a payroll_records row,
    so a mistaken 'Prepare' was permanent. Only drafts can be discarded:
    finalized rows must be reopened by a Super Admin first, and paid rows can
    never be removed, which keeps the payment trail intact.
    """
    require_permission(request,"payroll_manage")
    if not re.fullmatch(r"\d{4}-\d{2}",month): raise HTTPException(400,"Invalid salary month")
    actor=_payroll_actor(request)
    with get_db() as c:
        row=c.execute("SELECT id,employee_id,salary_month,payment_status,net_salary FROM payroll_records WHERE id=?",(payroll_id,)).fetchone()
        if not row: raise HTTPException(404,"Payroll not found")
        if row['payment_status']!='draft':
            raise HTTPException(409,"Only a draft payslip can be discarded. Reopen it first.")
        c.execute("DELETE FROM payroll_change_logs WHERE payroll_id=?",(payroll_id,))
        c.execute("DELETE FROM payroll_records WHERE id=?",(payroll_id,))
    audit(request,"discard","payroll",str(payroll_id),f"Discarded draft payslip for {row['salary_month']} (net {float(row['net_salary'] or 0):.2f}) by {actor}")
    return RedirectResponse(f"/payroll?month={month}&saved=discard",303)

@app.post("/payroll/bulk-discard")
def payroll_bulk_discard(request: Request, month: str=Form(...), confirm: str=Form("")):
    """Undo a bulk prepare: delete every DRAFT payslip for the month.

    Finalized and paid payslips are left untouched, so this can never wipe
    out work that has already been committed or disbursed.
    """
    require_permission(request,"payroll_manage")
    if not re.fullmatch(r"\d{4}-\d{2}",month): raise HTTPException(400,"Invalid salary month")
    if confirm.strip().upper()!="DISCARD":
        return RedirectResponse(f"/payroll?month={month}&error=confirm",303)
    actor=_payroll_actor(request)
    with get_db() as c:
        rows=c.execute("SELECT id FROM payroll_records WHERE salary_month=? AND payment_status='draft'",(month,)).fetchall()
        for row in rows:
            c.execute("DELETE FROM payroll_change_logs WHERE payroll_id=?",(row['id'],))
            c.execute("DELETE FROM payroll_records WHERE id=?",(row['id'],))
    audit(request,"bulk_discard","payroll",month,f"Discarded {len(rows)} draft payslips by {actor}")
    return RedirectResponse(f"/payroll?month={month}&saved=discard&made={len(rows)}",303)

@app.post("/employees/{employee_id}/salary-master")
def salary_master(request: Request, employee_id: int, fixed_salary: float=Form(...), overtime_rate: float=Form(0), return_month: str=Form("")):
    require_permission(request,"payroll_manage")
    if fixed_salary<0 or overtime_rate<0: raise HTTPException(400,"Salary values cannot be negative")
    with get_db() as c: c.execute("UPDATE employees SET fixed_salary=?,default_overtime_rate=?,updated_at=CURRENT_TIMESTAMP WHERE id=?",(fixed_salary,overtime_rate,employee_id))
    audit(request,"salary_master","employee",str(employee_id),f"Fixed salary and OT rate updated")
    return RedirectResponse(f"/employees/{employee_id}?month={return_month}#payroll",303)

@app.get("/payroll/export.xlsx")
def payroll_xlsx(request: Request, month: str):
    require_permission(request,"payroll_export")
    from openpyxl import Workbook
    from openpyxl.worksheet.page import PageMargins
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    if not re.fullmatch(r"\d{4}-\d{2}",month): raise HTTPException(400,"Invalid salary month")
    rows=_salary_sheet_rows(month); wb=Workbook(); summary=wb.active; summary.title="Summary"; ws=wb.create_sheet("Salary Sheet")
    dark="0D3B2E"; green="087F5B"; mint="EAF7F2"; pale="F4F7F6"; amber="FFF3CD"; red="FDE2E2"; white="FFFFFF"; grey="64748B"
    thin=Side(style="thin",color="D9E4E0"); border=Border(bottom=thin)
    headers=["SL","Staff ID","Employee Name","Department","Designation","Scheduled Duty","Worked Duty","Paid Leave","Unpaid Leave","Absent","Basic Salary","Per Duty Salary","Duty Earned Salary","Late Minutes","Late Deduction","Manual OT Hours","Manual OT Amount","Bonus","Gross Salary","Advance","Fine","Other Deduction","Total Deduction","Net Salary","Status","HR Note"]
    ws.merge_cells("A1:Z1"); ws["A1"]="BURAQ MONTHLY SALARY SHEET"; ws["A1"].font=Font(bold=True,size=20,color=white); ws["A1"].fill=PatternFill("solid",fgColor=dark); ws["A1"].alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[1].height=34
    ws.merge_cells("A2:Z2"); ws["A2"]=f"Salary Month: {month}  |  Generated: {datetime.now(ZoneInfo(settings.timezone)).strftime('%d %b %Y, %I:%M %p')}  |  HR/Admin Confidential"; ws["A2"].font=Font(italic=True,color=grey); ws["A2"].alignment=Alignment(horizontal="center")
    for col,title in enumerate(headers,1):
        cell=ws.cell(4,col,title); cell.font=Font(bold=True,color=white); cell.fill=PatternFill("solid",fgColor=green); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    ws.row_dimensions[4].height=42
    for index,r in enumerate(rows,1):
        row=4+index
        note_text=" | ".join(x for x in [f"Adjustment: {r.get('adjustment_reason')}" if r.get('adjustment_reason') else "",r.get('note') or ""] if x)
        values=[index,r['staff_id'],r['name'],r['department'] or "",r['designation'] or "",r['scheduled'],r['worked'],r['paid_leave'],r['unpaid_leave'],r['absent'],float(r['fixed_salary'] or 0),float(r['per_day_salary'] or 0),float(r['earned_basic_salary'] or 0),int(r['late_minutes'] or 0),float(r['late_deduction'] or 0),float(r['overtime_hours'] or 0),float(r['overtime_amount'] or 0),float(r['bonus'] or 0),float(r['gross_salary'] or 0),float(r.get('advance_amount') or 0),float(r.get('fine_amount') or 0),float(r['deduction'] or 0),float(r['total_deduction'] or 0),float(r['net_salary'] or 0),(r['payment_status'] or "not prepared").title() if r['payroll_id'] else "Not Prepared",note_text]
        for col,value in enumerate(values,1): ws.cell(row,col,value)
        fill=PatternFill("solid",fgColor=white if index%2 else pale)
        for cell in ws[row]: cell.fill=fill; cell.border=border; cell.alignment=Alignment(vertical="center",wrap_text=cell.column in {3,21})
        status=ws.cell(row,25); status.alignment=Alignment(horizontal="center"); status.fill=PatternFill("solid",fgColor=(mint if status.value=="Paid" else amber if status.value in {"Draft","Finalized"} else red))
    first_data=5; last_data=max(first_data,4+len(rows)); total_row=last_data+1
    ws.cell(total_row,1,"TOTAL"); ws.merge_cells(start_row=total_row,start_column=1,end_row=total_row,end_column=5)
    for col in [6,7,8,9,10,11,13,14,15,16,17,18,19,20,21,22,23,24]: ws.cell(total_row,col,f"=SUM({get_column_letter(col)}{first_data}:{get_column_letter(col)}{last_data})" if rows else 0)
    for cell in ws[total_row]: cell.font=Font(bold=True,color=white); cell.fill=PatternFill("solid",fgColor=dark); cell.border=border
    ws.cell(total_row,1).alignment=Alignment(horizontal="right")
    money_fmt='#,##0.00;[Red](#,##0.00);-'
    for row in ws.iter_rows(min_row=5,max_row=total_row):
        for col in [11,12,13,15,17,18,19,20,21,22,23,24]: row[col-1].number_format=money_fmt
    ws.freeze_panes="F5"; ws.auto_filter.ref=f"A4:Z{last_data}"; ws.sheet_view.showGridLines=False
    widths=[6,13,23,15,15,11,11,11,11,10,14,14,16,12,15,12,15,12,14,12,11,15,15,14,13,24]
    for col,width in enumerate(widths,1): ws.column_dimensions[get_column_letter(col)].width=width
    ws.page_setup.orientation="landscape"; ws.page_setup.paperSize=ws.PAPERSIZE_A4; ws.page_setup.fitToWidth=1; ws.page_setup.fitToHeight=1; ws.print_title_rows="1:4"; ws.print_area=f"A1:Z{total_row}"; ws.sheet_properties.pageSetUpPr.fitToPage=True; ws.sheet_properties.pageSetUpPr.autoPageBreaks=False; ws.print_options.horizontalCentered=True; ws.print_options.verticalCentered=True; ws.page_margins=PageMargins(left=0.15,right=0.15,top=0.25,bottom=0.25,header=0.1,footer=0.1)

    summary.merge_cells("A1:H1"); summary["A1"]="BURAQ PAYROLL SUMMARY"; summary["A1"].font=Font(bold=True,size=20,color=white); summary["A1"].fill=PatternFill("solid",fgColor=dark); summary["A1"].alignment=Alignment(horizontal="center"); summary.row_dimensions[1].height=34
    summary.merge_cells("A2:H2"); summary["A2"]=f"Salary Month: {month}  |  All active employees included"; summary["A2"].font=Font(italic=True,color=grey); summary["A2"].alignment=Alignment(horizontal="center")
    metrics=[("Active Employees",len(rows)),("Payroll Prepared",sum(1 for r in rows if r['payroll_id'])),("Scheduled Duties",sum(r['scheduled'] for r in rows)),("Worked Duties",sum(r['worked'] for r in rows)),("Paid Leave Days",sum(r['paid_leave'] for r in rows)),("Late Minutes",sum(r['late_minutes'] for r in rows)),("Gross Salary",f"='Salary Sheet'!S{total_row}"),("Total Deductions",f"='Salary Sheet'!W{total_row}"),("Net Payroll",f"='Salary Sheet'!X{total_row}")]
    for i,(label,value) in enumerate(metrics):
        row=4+(i//3)*3; col=1+(i%3)*3; summary.merge_cells(start_row=row,start_column=col,end_row=row,end_column=col+1); summary.merge_cells(start_row=row+1,start_column=col,end_row=row+1,end_column=col+1)
        summary.cell(row,col,label).font=Font(bold=True,color=grey); summary.cell(row,col).alignment=Alignment(horizontal="center"); summary.cell(row+1,col,value).font=Font(bold=True,size=18,color=dark); summary.cell(row+1,col).alignment=Alignment(horizontal="center"); summary.cell(row,col).fill=summary.cell(row+1,col).fill=PatternFill("solid",fgColor=mint)
        if i>=6: summary.cell(row+1,col).number_format=money_fmt
    summary.merge_cells("A14:H14"); summary["A14"]="Formula: Basic Salary ÷ Scheduled Duty × Completed/Paid Duty, minus late minutes based on payable duty time after break. Overtime is manual only."; summary["A14"].alignment=Alignment(horizontal="center",wrap_text=True); summary["A14"].font=Font(italic=True,color=grey)
    summary.sheet_view.showGridLines=False
    for col in range(1,9): summary.column_dimensions[get_column_letter(col)].width=17
    summary.page_setup.orientation="landscape"; summary.page_setup.paperSize=summary.PAPERSIZE_A4; summary.page_setup.fitToWidth=1; summary.page_setup.fitToHeight=1; summary.print_area="A1:H14"; summary.sheet_properties.pageSetUpPr.fitToPage=True; summary.sheet_properties.pageSetUpPr.autoPageBreaks=False; summary.print_options.horizontalCentered=True; summary.print_options.verticalCentered=True; summary.page_margins=PageMargins(left=0.35,right=0.35,top=0.5,bottom=0.5,header=0.1,footer=0.1)
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f"attachment; filename=BURAQ-Payroll-{month}.xlsx"})

def _pdf_font():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    path="/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
    if os.path.exists(path):
        try: pdfmetrics.registerFont(TTFont("BuraqUnicode",path))
        except Exception: pass
        return "BuraqUnicode"
    return "Helvetica"

@app.get("/payroll/export.pdf")
def payroll_pdf(request: Request, month: str):
    require_permission(request,"payroll_export")
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    if not re.fullmatch(r"\d{4}-\d{2}",month): raise HTTPException(400,"Invalid salary month")
    month_label=datetime.strptime(month,"%Y-%m").strftime("%B %Y")
    rows=_salary_sheet_rows(month); out=io.BytesIO(); font=_pdf_font(); styles=getSampleStyleSheet(); styles['Title'].fontName=font; styles['Normal'].fontName=font; styles['Normal'].alignment=1; styles['Normal'].textColor=colors.HexColor("#64748B"); styles['Heading1'].fontName=font; styles['Heading1'].fontSize=22; styles['Heading1'].leading=26; styles['Heading1'].alignment=1; styles['Heading1'].textColor=colors.HexColor("#087F5B")
    data=[["Staff ID","Employee","Duty","Earned","Late","Total Ded.","Net","Status"]]+[[str(r['staff_id']),str(r['name']),f"{r['worked']}/{r['scheduled']}",_money(r['earned_basic_salary']),f"{int(r['late_minutes'])}m / {_money(r['late_deduction'])}",_money(r['total_deduction']),_money(r['net_salary']),str(r['payment_status'] or 'not prepared').title()] for r in rows]
    data.append(["","TOTAL","","","","",_money(sum(float(r['net_salary']) for r in rows)),""])
    doc=SimpleDocTemplate(out,pagesize=landscape(A4),leftMargin=24,rightMargin=24,topMargin=24,bottomMargin=24); table=Table(data,repeatRows=1,colWidths=[65,155,75,70,70,75,80,60])
    table.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),font),("BACKGROUND",(0,0),(-1,0),colors.HexColor("#087F5B")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),font),("FONTNAME",(0,-1),(-1,-1),font),("FONTNAME",(0,-1),(-1,-1),font),("GRID",(0,0),(-1,-1),.35,colors.HexColor("#B7C8C2")),("ROWBACKGROUNDS",(0,1),(-1,-2),[colors.white,colors.HexColor("#F4F7F6")]),("ALIGN",(2,1),(-2,-1),"RIGHT"),("TOPPADDING",(0,0),(-1,-1),7),("BOTTOMPADDING",(0,0),(-1,-1),7)]))
    doc.build([Paragraph("BURAQ Payment Sheet",styles['Title']),Spacer(1,4),Paragraph(month_label,styles['Heading1']),Paragraph("HR/Admin confidential",styles['Normal']),Spacer(1,14),table]); out.seek(0)
    return StreamingResponse(out,media_type="application/pdf",headers={"Content-Disposition":f"attachment; filename=BURAQ-Payment-Sheet-{month}.pdf"})

def _build_payslip_pdf(r) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

    def deduction(value):
        amount=float(value or 0)
        return "0.00" if abs(amount)<0.005 else f"- {_money(amount)}"

    out=io.BytesIO(); font=_pdf_font(); styles=getSampleStyleSheet()
    styles['Title'].fontName=font; styles['Title'].fontSize=24; styles['Title'].leading=28; styles['Title'].textColor=colors.HexColor("#0D3B2E")
    styles['Normal'].fontName=font
    month_label=datetime.strptime(str(r['salary_month']),"%Y-%m").strftime("%B %Y").upper()
    month_style=ParagraphStyle("Month",parent=styles['Heading1'],fontName=font,fontSize=20,leading=24,alignment=1,textColor=colors.HexColor("#087F5B"),spaceAfter=12)
    muted=ParagraphStyle("Muted",parent=styles['Normal'],fontName=font,fontSize=9,textColor=colors.HexColor("#64748B"),alignment=1)

    employee_data=[
        ["Employee",str(r['name']),"Staff ID",str(r['staff_id'])],
        ["Department",str(r['department'] or '-'),"Designation",str(r['designation'] or '-')],
        ["Payment status",str(r['payment_status']).title(),"Currency","BDT"],
    ]
    employee_table=Table(employee_data,colWidths=[90,155,95,150])
    employee_table.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),font),("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#F4F7F6")),("TEXTCOLOR",(0,0),(0,-1),colors.HexColor("#64748B")),("TEXTCOLOR",(2,0),(2,-1),colors.HexColor("#64748B")),("FONTNAME",(1,0),(1,-1),font),("FONTNAME",(3,0),(3,-1),font),("GRID",(0,0),(-1,-1),.35,colors.HexColor("#D5E2DD")),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))

    duty_data=[
        ["Scheduled Duty","Worked Duty","Paid Leave","Absent"],
        [f"{float(r['scheduled_duty_days'] or 0):g} days",f"{float(r['worked_duty_days'] or 0):g} days",f"{float(r['paid_leave_days'] or 0):g} days",f"{float(r['absent_days'] or 0):g} days"],
    ]
    duty_table=Table(duty_data,colWidths=[122.5]*4)
    duty_table.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),font),("BACKGROUND",(0,0),(-1,0),colors.HexColor("#E7F5EF")),("TEXTCOLOR",(0,0),(-1,0),colors.HexColor("#087F5B")),("ALIGN",(0,0),(-1,-1),"CENTER"),("GRID",(0,0),(-1,-1),.35,colors.HexColor("#B7C8C2")),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))

    data=[["Salary Item","Amount (BDT)"],["Monthly Basic Salary",_money(r['fixed_salary'])],[f"Earned Basic ({float(r['worked_duty_units'] or 0):g}/{float(r['scheduled_duty_days'] or 0):g} duties)",_money(r['earned_basic_salary'])],[f"Manual Overtime ({r['overtime_hours']:.2f} hours × {_money(r['overtime_rate'])})",_money(r['overtime_amount'])],["Bonus",_money(r['bonus'])],["GROSS SALARY",_money(r['gross_salary'])],[f"Late deduction ({int(r['late_minutes'] or 0)} minutes)",deduction(r['late_deduction'])],["Salary advance",deduction(r['advance_amount'])],["Fine",deduction(r['fine_amount'])],["Other deduction",deduction(r['deduction'])],["TOTAL DEDUCTION",deduction(r['total_deduction'])],["NET SALARY",_money(r['net_salary'])]]
    table=Table(data,colWidths=[330,160]); table.setStyle(TableStyle([("FONTNAME",(0,0),(-1,-1),font),("BACKGROUND",(0,0),(-1,0),colors.HexColor("#087F5B")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.5,colors.HexColor("#B7C8C2")),("ALIGN",(1,1),(1,-1),"RIGHT"),("BACKGROUND",(0,5),(-1,5),colors.HexColor("#F1F5F4")),("BACKGROUND",(0,-1),(-1,-1),colors.HexColor("#D1FAE5")),("TEXTCOLOR",(0,-1),(-1,-1),colors.HexColor("#065F46")),("FONTNAME",(0,5),(-1,5),font),("FONTNAME",(0,-2),(-1,-1),font),("TOPPADDING",(0,0),(-1,-1),8),("BOTTOMPADDING",(0,0),(-1,-1),8)]))
    doc=SimpleDocTemplate(out,pagesize=A4,leftMargin=50,rightMargin=50,topMargin=36,bottomMargin=36,title=f"BURAQ Payment Sheet - {month_label}")
    doc.build([Paragraph("BURAQ PAYMENT SHEET",styles['Title']),Paragraph(month_label,month_style),employee_table,Spacer(1,14),duty_table,Spacer(1,14),table,Spacer(1,14),Paragraph("Confidential • Generated for HR/Admin use only",muted)]); return out.getvalue()

@app.get("/payroll/{payroll_id}/payslip.pdf")
def payroll_payslip(request: Request, payroll_id: int):
    require_permission(request,"payroll_export")
    with get_db() as c: r=c.execute("SELECT p.*,e.staff_id,e.name,e.department,e.designation FROM payroll_records p JOIN employees e ON e.id=p.employee_id WHERE p.id=?",(payroll_id,)).fetchone()
    if not r: raise HTTPException(404,"Payroll not found")
    payslip=dict(r)
    if r['payment_status']=='draft':
        mode=str(r['overtime_mode'] or 'auto'); manual_hours=float(r['overtime_hours'] or 0) if mode=='manual' else 0
        calc=_calculate_employee_payroll(r['employee_id'],r['salary_month'],float(r['fixed_salary'] or 0),float(r['overtime_rate'] or 0),mode,manual_hours,float(r['bonus'] or 0),float(r['advance_amount'] or 0),float(r['fine_amount'] or 0),float(r['deduction'] or 0))
        payslip.update(calc)
        payslip.update({'scheduled_duty_days':calc['scheduled'],'worked_duty_days':calc['worked'],'paid_leave_days':calc['paid_leave'],'absent_days':calc['absent'],'absent_duty_units':calc['absent'],'unpaid_leave_units':calc['unpaid_leave']})
    elif r['calculation_snapshot']:
        try: payslip.update(json.loads(r['calculation_snapshot']))
        except Exception: pass
    payslip.setdefault('earned_basic_salary',max(float(payslip.get('fixed_salary') or 0)-float(payslip.get('absent_deduction') or 0)-float(payslip.get('unpaid_leave_deduction') or 0),0))
    payslip.setdefault('late_minutes',0); payslip.setdefault('late_deduction',0)
    out=io.BytesIO(_build_payslip_pdf(payslip))
    return StreamingResponse(out,media_type="application/pdf",headers={"Content-Disposition":f"attachment; filename=BURAQ-Payment-Sheet-{r['staff_id']}-{r['salary_month']}.pdf"})

@app.get("/reports", response_class=HTMLResponse)
def reports_page(request: Request, start_date: str = "", end_date: str = "", status: str = "", department: str = ""):
    require_permission(request, "reports_view")
    today = datetime.now(ZoneInfo(settings.timezone)).date()
    start_date = start_date or today.replace(day=1).isoformat(); end_date = end_date or today.isoformat()
    rows = _attendance_report_rows(start_date, end_date, status, department)
    with get_db() as c:
        deps = c.execute("SELECT DISTINCT department FROM employees WHERE department IS NOT NULL AND department<>'' ORDER BY department").fetchall()
    dep_options = ''.join(f"<option value='{escape(d['department'])}' {'selected' if department==d['department'] else ''}>{escape(d['department'])}</option>" for d in deps)
    table_rows = ''.join(f"<tr><td>{escape(r['work_date'])}</td><td><b>{escape(r['staff_id'])}</b><div class='sub'>{escape(r['name'])}</div></td><td>{escape(r['department'] or '-')}</td><td>{escape(r['shift'])}</td><td>{escape(format_time_12h(r['check_in']) or '-')}</td><td>{escape(format_time_12h(r['check_out']) or '-')}</td><td>{r['late_minutes']}m</td><td>{r['overtime_minutes']}m</td><td>{escape(r['status'])}</td></tr>" for r in rows) or "<tr><td colspan='9'>No records found.</td></tr>"
    q=f"start_date={start_date}&end_date={end_date}&status={status}&department={department}"
    exports=(f"<a class='btn secondary' href='/reports/export.csv?{q}'>CSV</a><a class='btn secondary' href='/reports/export.xlsx?{q}'>Excel</a><a class='btn secondary' href='/reports/export.pdf?{q}'>PDF</a>" if has_permission(request,'reports_export') else '')
    body=f"""<div class='card'><form method='get'><div class='grid'><div><label>From</label><input type='date' name='start_date' value='{start_date}'></div><div><label>To</label><input type='date' name='end_date' value='{end_date}'></div><div><label>Status</label><select name='status'><option value=''>All</option><option value='present' {'selected' if status=='present' else ''}>Present</option><option value='leave' {'selected' if status=='leave' else ''}>Leave</option><option value='absent' {'selected' if status=='absent' else ''}>Absent</option></select></div><div><label>Department</label><select name='department'><option value=''>All</option>{dep_options}</select></div></div><div class='actions'><button class='btn'>Apply</button>{exports}</div></form></div><div class='section-gap'></div><div class='grid'><div class='card'><div class='sub'>Records</div><div class='metric'>{len(rows)}</div></div><div class='card'><div class='sub'>Late Records</div><div class='metric'>{sum(1 for r in rows if r['late_minutes']>0)}</div></div><div class='card'><div class='sub'>Overtime Minutes</div><div class='metric'>{sum(r['overtime_minutes'] for r in rows)}</div></div><div class='card'><div class='sub'>Leave Records</div><div class='metric'>{sum(1 for r in rows if r['status']=='leave')}</div></div></div><div class='section-gap'></div><div class='card'><h2>Attendance Report</h2><div style='overflow:auto'><table><thead><tr><th>Date</th><th>Employee</th><th>Department</th><th>Shift</th><th>In</th><th>Out</th><th>Late</th><th>OT</th><th>Status</th></tr></thead><tbody>{table_rows}</tbody></table></div></div>"""
    return layout("Attendance Reports", body, request, "reports")

@app.get("/reports/export.csv")
def report_csv(request: Request, start_date: str, end_date: str, status: str = "", department: str = ""):
    require_permission(request,"reports_export"); rows=_attendance_report_rows(start_date,end_date,status,department)
    out=io.StringIO(); w=csv.writer(out); w.writerow(["Date","Staff ID","Name","Department","Shift","Check In","Check Out","Late","Early Leave","Overtime","Status"])
    for r in rows: w.writerow([r['work_date'],r['staff_id'],r['name'],r['department'],r['shift'],format_time_12h(r['check_in']),format_time_12h(r['check_out']),r['late_minutes'],r['early_leave_minutes'],r['overtime_minutes'],r['status']])
    return StreamingResponse(io.BytesIO(out.getvalue().encode("utf-8-sig")),media_type="text/csv",headers={"Content-Disposition":f"attachment; filename=BURAQ-{start_date}-to-{end_date}.csv"})

@app.get("/reports/export.xlsx")
def report_xlsx(request: Request, start_date: str, end_date: str, status: str = "", department: str = ""):
    require_permission(request,"reports_export")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    rows=_attendance_report_rows(start_date,end_date,status,department); wb=Workbook(); ws=wb.active; ws.title="Attendance"
    headers=["Date","Staff ID","Name","Department","Shift","Check In","Check Out","Late","Early Leave","Overtime","Status"]; ws.append(headers)
    for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="087F5B")
    for r in rows: ws.append([r['work_date'],r['staff_id'],r['name'],r['department'],r['shift'],format_time_12h(r['check_in']),format_time_12h(r['check_out']),r['late_minutes'],r['early_leave_minutes'],r['overtime_minutes'],r['status']])
    for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=min(max(len(str(x.value or "")) for x in col)+2,30)
    out=io.BytesIO(); wb.save(out); out.seek(0)
    return StreamingResponse(out,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f"attachment; filename=BURAQ-{start_date}-to-{end_date}.xlsx"})

@app.get("/reports/export.pdf")
def report_pdf(request: Request, start_date: str, end_date: str, status: str = "", department: str = ""):
    require_permission(request,"reports_export")
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    rows=_attendance_report_rows(start_date,end_date,status,department); out=io.BytesIO()
    data=[["Date","Staff ID","Name","Department","Shift","In","Out","Late","OT","Status"]]+[[str(r['work_date'] or ""),str(r['staff_id'] or ""),str(r['name'] or ""),str(r['department'] or ""),str(r['shift'] or ""),format_time_12h(r['check_in']),format_time_12h(r['check_out']),str(r['late_minutes'] or ""),str(r['overtime_minutes'] or ""),str(r['status'] or "")] for r in rows]
    doc=SimpleDocTemplate(out,pagesize=landscape(A4),leftMargin=18,rightMargin=18,topMargin=18,bottomMargin=18); styles=getSampleStyleSheet(); table=Table(data,repeatRows=1)
    table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#087F5B")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTSIZE",(0,0),(-1,-1),7),("GRID",(0,0),(-1,-1),.3,colors.grey)]))
    doc.build([Paragraph("BURAQ Attendance Report",styles["Title"]),Paragraph(f"{start_date} to {end_date}",styles["Normal"]),Spacer(1,8),table]); out.seek(0)
    return StreamingResponse(out,media_type="application/pdf",headers={"Content-Disposition":f"attachment; filename=BURAQ-{start_date}-to-{end_date}.pdf"})

@app.get("/hr-operations", response_class=HTMLResponse)
def operations_page(request: Request, saved: str = ""):
    require_permission(request,"leave_view")
    with get_db() as c:
        employees=c.execute("SELECT id,staff_id,name FROM employees ORDER BY name").fetchall()
        leaves=c.execute("SELECT l.*,e.staff_id,e.name FROM leave_requests l JOIN employees e ON e.id=l.employee_id ORDER BY l.id DESC LIMIT 200").fetchall()
        corrections=c.execute("SELECT x.*,e.staff_id,e.name FROM attendance_corrections x JOIN employees e ON e.id=x.employee_id ORDER BY x.id DESC LIMIT 200").fetchall()
        shifts=c.execute("SELECT * FROM shifts ORDER BY name").fetchall(); deps=c.execute("SELECT * FROM departments ORDER BY name").fetchall()
    opts=''.join(f"<option value='{e['id']}'>{escape(e['staff_id'])} — {escape(e['name'])}</option>" for e in employees)
    can_decide=has_permission(request,"leave_manage")
    leave_rows=[]
    for r in leaves:
        actions="-"
        if can_decide and r['status']=='pending': actions=f"<div class='actions'><form method='post' action='/leave/{r['id']}/approve'><button class='btn'>Approve</button></form><form method='post' action='/leave/{r['id']}/reject'><button class='btn danger'>Reject</button></form></div>"
        leave_rows.append(f"<tr><td>{escape(r['staff_id'])}<div class='sub'>{escape(r['name'])}</div></td><td>{escape(r['leave_type'])}</td><td>{escape(r['start_date'])} → {escape(r['end_date'])}</td><td>{escape(r['reason'] or '')}</td><td>{escape(r['status'])}</td><td>{actions}</td></tr>")
    correction_rows=[]
    for r in corrections:
        actions="-"
        if can_decide and r['status']=='pending': actions=f"<div class='actions'><form method='post' action='/correction/{r['id']}/approve'><button class='btn'>Apply</button></form><form method='post' action='/correction/{r['id']}/reject'><button class='btn danger'>Reject</button></form></div>"
        correction_rows.append(f"<tr><td>{escape(r['staff_id'])}<div class='sub'>{escape(r['name'])}</div></td><td>{escape(r['work_date'])}</td><td>{escape(r['requested_check_in'] or '-')}</td><td>{escape(r['requested_check_out'] or '-')}</td><td>{escape(r['reason'])}</td><td>{escape(r['status'])}</td><td>{actions}</td></tr>")
    correction_form=""
    if has_permission(request,"attendance_edit"):
        correction_form=f"<div class='card'><h2>Attendance Correction</h2><form method='post' action='/correction'><label>Employee</label><select name='employee_id'>{opts}</select><label>Work Date</label><input type='date' name='work_date' required><label>Check In (HH:MM)</label><input name='check_in'><label>Check Out (HH:MM)</label><input name='check_out'><label>Reason</label><textarea name='reason' required></textarea><button class='btn'>Submit</button></form></div>"
    management=""
    if has_permission(request,"shift_manage") or has_permission(request,"department_manage"):
        shift_form=("<div class='card'><h2>Shifts</h2><form method='post' action='/shifts'><label>Name</label><input name='name' required><label>Start</label><input type='time' name='start_time' required><label>End</label><input type='time' name='end_time' required><button class='btn'>Save Shift</button></form><p class='sub'>"+", ".join(escape(x['name']) for x in shifts)+"</p></div>") if has_permission(request,"shift_manage") else ""
        dep_form=("<div class='card'><h2>Departments</h2><form method='post' action='/departments'><label>Name</label><input name='name' required><button class='btn'>Save Department</button></form><p class='sub'>"+", ".join(escape(x['name']) for x in deps)+"</p></div>") if has_permission(request,"department_manage") else ""
        management=f"<div class='section-gap'></div><div class='two'>{shift_form}{dep_form}</div>"
    body=("<div class='notice'>Saved successfully.</div>" if saved else "")+f"<div class='two'><div class='card'><h2>Leave Request</h2><form method='post' action='/leave'><label>Employee</label><select name='employee_id'>{opts}</select><label>Type</label><select name='leave_type'><option>Casual</option><option>Sick</option><option>Annual</option><option>Unpaid</option></select><label>Start</label><input type='date' name='start_date' required><label>End</label><input type='date' name='end_date' required><label>Reason</label><textarea name='reason'></textarea><button class='btn'>Submit</button></form></div>{correction_form}</div>{management}<div class='section-gap'></div><div class='card'><h2>Leave Requests</h2><div style='overflow:auto'><table><thead><tr><th>Employee</th><th>Type</th><th>Dates</th><th>Reason</th><th>Status</th><th>Action</th></tr></thead><tbody>{''.join(leave_rows) or '<tr><td colspan=6>No requests.</td></tr>'}</tbody></table></div></div><div class='section-gap'></div><div class='card'><h2>Attendance Corrections</h2><div style='overflow:auto'><table><thead><tr><th>Employee</th><th>Date</th><th>In</th><th>Out</th><th>Reason</th><th>Status</th><th>Action</th></tr></thead><tbody>{''.join(correction_rows) or '<tr><td colspan=7>No requests.</td></tr>'}</tbody></table></div></div>"
    return layout("HR Operations",body,request,"operations")

@app.post("/leave")
def create_leave(request: Request, employee_id: int=Form(...), leave_type: str=Form(...), start_date: str=Form(...), end_date: str=Form(...), reason: str=Form("")):
    require_permission(request,"leave_view")
    with get_db() as c: c.execute("INSERT INTO leave_requests(employee_id,leave_type,start_date,end_date,reason,requested_by) VALUES(?,?,?,?,?,?)",(employee_id,leave_type,start_date,end_date,reason,str(request.session.get('hr_id') or 'super_admin')))
    audit(request,"create","leave_request",str(employee_id),f"{start_date} to {end_date}"); return RedirectResponse("/hr-operations?saved=1",303)

@app.post("/leave/{request_id}/{action}")
def decide_leave(request: Request, request_id: int, action: str):
    require_permission(request,"leave_manage"); status="approved" if action=="approve" else "rejected" if action=="reject" else None
    if not status: raise HTTPException(400)
    with get_db() as c:
        row=c.execute("SELECT * FROM leave_requests WHERE id=?",(request_id,)).fetchone(); c.execute("UPDATE leave_requests SET status=?,decided_by=?,decided_at=CURRENT_TIMESTAMP WHERE id=?",(status,str(request.session.get('hr_id') or 'super_admin'),request_id))
        if row and status=="approved":
            d=datetime.fromisoformat(row['start_date']).date(); end=datetime.fromisoformat(row['end_date']).date()
            while d<=end:
                c.execute("INSERT INTO attendance(employee_id,work_date,status,source) VALUES(?,?,?,?) ON CONFLICT(employee_id,work_date) DO UPDATE SET status=excluded.status,source=excluded.source,updated_at=CURRENT_TIMESTAMP",(row['employee_id'],d.isoformat(),'leave','hr')); d+=timedelta(days=1)
    audit(request,status,"leave_request",str(request_id),""); return RedirectResponse("/hr-operations?saved=1",303)

def _missing_duty_days(employee_id: int, month: str) -> list[dict]:
    """Scheduled duty days in `month` that have no attendance row at all.

    Deliberately excludes days that DO have a row but are incomplete (a
    check-in with no check-out) — those already surface as incomplete_dates
    in payroll and are a different fix.
    """
    first = datetime.strptime(month + "-01", "%Y-%m-%d").date()
    next_month = (first.replace(day=28) + timedelta(days=4)).replace(day=1)
    last = next_month - timedelta(days=1)
    today = datetime.now(ZoneInfo(settings.timezone)).date()
    # Never offer to backfill a day that has not happened yet.
    effective_last = min(last, today - timedelta(days=1))
    if effective_last < first:
        return []
    with get_db() as c:
        weekly = c.execute(
            "SELECT * FROM duty_schedules WHERE employee_id=? AND is_active",
            (employee_id,)).fetchall()
        custom = c.execute(
            "SELECT * FROM custom_duties WHERE employee_id=? AND duty_date>=? "
            "AND duty_date<=? AND is_active",
            (employee_id, first.isoformat(), effective_last.isoformat())).fetchall()
        have = {str(r["work_date"]) for r in c.execute(
            "SELECT work_date FROM attendance WHERE employee_id=? AND work_date>=? AND work_date<=?",
            (employee_id, first.isoformat(), effective_last.isoformat())).fetchall()}
        leaves = c.execute(
            "SELECT start_date,end_date FROM leave_requests WHERE employee_id=? "
            "AND status='approved' AND start_date<=? AND end_date>=?",
            (employee_id, effective_last.isoformat(), first.isoformat())).fetchall()

    on_leave = set()
    for lv in leaves:
        day = max(datetime.fromisoformat(str(lv["start_date"])).date(), first)
        end = min(datetime.fromisoformat(str(lv["end_date"])).date(), effective_last)
        while day <= end:
            on_leave.add(day.isoformat())
            day += timedelta(days=1)

    weekly_by_day = {int(r["weekday"]): r for r in weekly}
    custom_by_date = {str(r["duty_date"]): r for r in custom}

    out = []
    day = first
    while day <= effective_last:
        key = day.isoformat()
        duty = custom_by_date.get(key) or weekly_by_day.get(day.weekday())
        # Approved leave is not a missing day — it is accounted for already.
        if duty and key not in have and key not in on_leave:
            out.append({
                "date": key,
                "label": day.strftime("%a %d %b"),
                "start": str(duty["start_time"]),
                "end": str(duty["end_time"]),
                "break_minutes": int(duty["break_minutes"] or 0),
            })
        day += timedelta(days=1)
    return out


def _month_payroll_locked(employee_id: int, month: str) -> bool:
    with get_db() as c:
        row = c.execute(
            "SELECT payment_status FROM payroll_records WHERE employee_id=? AND salary_month=?",
            (employee_id, month)).fetchone()
    return bool(row and str(row["payment_status"]) in {"finalized", "paid"})


@app.get("/attendance/missing", response_class=HTMLResponse)
def missing_duty_page(request: Request, employee_id: int = 0, month: str = "",
                      saved: str = "", error: str = "", made: int = 0):
    require_permission(request, "attendance_edit")
    current = datetime.now(ZoneInfo(settings.timezone)).strftime("%Y-%m")
    month = month or current
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise HTTPException(400, "Invalid month")
    month_label = datetime.strptime(month, "%Y-%m").strftime("%B %Y")

    with get_db() as c:
        employees = c.execute(
            "SELECT id,staff_id,name FROM employees WHERE is_active ORDER BY staff_id").fetchall()
    if not employees:
        return layout("Missing duty", "<div class='card'><div class='empty-cell'>"
                      "No active employees.</div></div>", request, "attendance")
    if not employee_id:
        employee_id = int(employees[0]["id"])
    chosen = next((e for e in employees if int(e["id"]) == employee_id), None)
    if not chosen:
        raise HTTPException(404, "Employee not found")

    days = _missing_duty_days(employee_id, month)
    locked = _month_payroll_locked(employee_id, month)

    if saved == "filled":
        notice = (f"<div class='notice notice-ok'>Recorded {made} duty day"
                  f"{'' if made == 1 else 's'}. Payroll for {month_label} will "
                  f"pick this up straight away.</div>")
    elif error == "locked":
        notice = ("<div class='notice notice-bad'>This month's payslip is finalized or "
                  "paid, so attendance cannot be changed. A Super Admin must reopen the "
                  "payslip first.</div>")
    elif error == "exists":
        notice = ("<div class='notice notice-bad'>That day already has an attendance "
                  "record. Use Attendance Correction to change it instead.</div>")
    elif error == "time":
        notice = ("<div class='notice notice-bad'>Check-out must be later than check-in, "
                  "and both must be in HH:MM form.</div>")
    elif error:
        notice = "<div class='notice notice-bad'>That day could not be recorded.</div>"
    else:
        notice = ""

    options = "".join(
        f"<option value='{e['id']}'{' selected' if int(e['id']) == employee_id else ''}>"
        f"{escape(str(e['staff_id']))} — {escape(str(e['name']))}</option>"
        for e in employees)

    if locked:
        rows_html = ("<div class='empty-cell'>The payslip for this month is locked, "
                     "so these days cannot be edited.</div>")
    elif not days:
        rows_html = (f"<div class='empty-cell'>No missing duty days in {month_label}. "
                     f"Every scheduled duty either has attendance or approved leave.</div>")
    else:
        rows = []
        for d in days:
            rows.append(
                f"<tr><td><b>{escape(d['label'])}</b>"
                f"<div class='sub'>{escape(d['date'])}</div></td>"
                f"<td class='sub'>Scheduled {escape(d['start'])} – {escape(d['end'])}"
                + (f", {d['break_minutes']}m break" if d["break_minutes"] else "")
                + "</td>"
                f"<td><form method='post' action='/attendance/backfill' class='fill-form'>"
                f"<input type='hidden' name='employee_id' value='{employee_id}'>"
                f"<input type='hidden' name='month' value='{month}'>"
                f"<input type='hidden' name='work_date' value='{d['date']}'>"
                f"<input type='time' name='check_in' value='{escape(d['start'])}' "
                f"aria-label='Check in for {escape(d['label'])}' required>"
                f"<input type='time' name='check_out' value='{escape(d['end'])}' "
                f"aria-label='Check out for {escape(d['label'])}' required>"
                f"<select name='status' aria-label='Day type'>"
                f"<option value='present'>Full day</option>"
                f"<option value='half_day'>Half day</option></select>"
                f"<button class='btn small'>Record</button></form></td></tr>")
        rows_html = (
            "<div class='table-scroll'><table class='dashboard-table'>"
            "<thead><tr><th>Day</th><th>Duty</th><th>Record it</th></tr></thead>"
            f"<tbody>{''.join(rows)}</tbody></table></div>"
            f"<form method='post' action='/attendance/backfill-all' class='fill-all'>"
            f"<input type='hidden' name='employee_id' value='{employee_id}'>"
            f"<input type='hidden' name='month' value='{month}'>"
            f"<button class='btn secondary'>Record all {len(days)} at their scheduled times</button>"
            f"<div class='hint'>Fills every day above using the duty start and end time. "
            f"Late minutes come out as zero, so only use this when the staff actually "
            f"worked their normal hours.</div></form>")

    body = f"""
    {notice}
    <div class='dashboard-head'>
      <div class='dashboard-greet'>
        <h1>Missing duty days</h1>
        <div class='dashboard-date'><span>Days with a scheduled duty but no attendance at all</span></div>
      </div>
    </div>
    <div class='card' style='margin-bottom:14px'>
      <form method='get' class='missing-picker'>
        <div><label for='md-emp'>Employee</label>
        <select id='md-emp' name='employee_id'>{options}</select></div>
        <div><label for='md-month'>Month</label>
        <input id='md-month' type='month' name='month' value='{month}'></div>
        <button class='btn'>Show</button>
      </form>
    </div>
    <div class='card'>
      <div class='card-head'><div>
        <h3>{escape(str(chosen['name']))} — {month_label}</h3>
        <div class='sub'>{len(days)} day{'' if len(days) == 1 else 's'} missing. Recording a
        day here applies immediately; it does not need a second approval.</div>
      </div></div>
      {rows_html}
    </div>
    """
    return layout("Missing duty", body, request, "attendance")


def _write_backfill(c, employee_id: int, work_date: str, check_in: str,
                    check_out: str, status: str, actor: str) -> None:
    """Insert one attendance row for a past duty day, with late minutes
    measured against that day's scheduled start."""
    start_dt = datetime.fromisoformat(f"{work_date}T{check_in}:00")
    end_dt = datetime.fromisoformat(f"{work_date}T{check_out}:00")
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)
    duty = c.execute(
        "SELECT start_time FROM custom_duties WHERE employee_id=? AND duty_date=? AND is_active",
        (employee_id, work_date)).fetchone()
    if not duty:
        weekday = datetime.fromisoformat(work_date).date().weekday()
        duty = c.execute(
            "SELECT start_time FROM duty_schedules WHERE employee_id=? AND weekday=? AND is_active",
            (employee_id, weekday)).fetchone()
    late = 0
    if duty:
        scheduled_start = datetime.fromisoformat(f"{work_date}T{str(duty['start_time'])}:00")
        late = max(0, int((start_dt - scheduled_start).total_seconds() // 60))
    c.execute(
        "INSERT INTO attendance(employee_id,work_date,check_in,check_out,late_minutes,status,source) "
        "VALUES(?,?,?,?,?,?,'hr_backfill')",
        (employee_id, work_date, start_dt.isoformat(timespec="seconds"),
         end_dt.isoformat(timespec="seconds"), late, status))


@app.post("/attendance/backfill")
def attendance_backfill(request: Request, employee_id: int = Form(...),
                        work_date: str = Form(...), month: str = Form(...),
                        check_in: str = Form(...), check_out: str = Form(...),
                        status: str = Form("present")):
    require_permission(request, "attendance_edit")
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", work_date) or not re.fullmatch(r"\d{4}-\d{2}", month):
        raise HTTPException(400, "Invalid date")
    if not re.fullmatch(r"\d{2}:\d{2}", check_in) or not re.fullmatch(r"\d{2}:\d{2}", check_out):
        return RedirectResponse(f"/attendance/missing?employee_id={employee_id}&month={month}&error=time", 303)
    if status not in {"present", "half_day"}:
        raise HTTPException(400, "Invalid status")
    today = datetime.now(ZoneInfo(settings.timezone)).date()
    if datetime.fromisoformat(work_date).date() >= today:
        raise HTTPException(400, "Only past days can be recorded here")
    if _month_payroll_locked(employee_id, month):
        return RedirectResponse(f"/attendance/missing?employee_id={employee_id}&month={month}&error=locked", 303)
    actor = str(request.session.get("user_name") or request.session.get("hr_id") or "Super Admin")
    with get_db() as c:
        exists = c.execute("SELECT id FROM attendance WHERE employee_id=? AND work_date=?",
                           (employee_id, work_date)).fetchone()
        if exists:
            return RedirectResponse(f"/attendance/missing?employee_id={employee_id}&month={month}&error=exists", 303)
        _write_backfill(c, employee_id, work_date, check_in, check_out, status, actor)
    audit(request, "backfill", "attendance", f"{employee_id}:{work_date}",
          f"Recorded past duty {check_in}-{check_out} ({status}) by {actor}")
    return RedirectResponse(f"/attendance/missing?employee_id={employee_id}&month={month}&saved=filled&made=1", 303)


@app.post("/attendance/backfill-all")
def attendance_backfill_all(request: Request, employee_id: int = Form(...), month: str = Form(...)):
    require_permission(request, "attendance_edit")
    if not re.fullmatch(r"\d{4}-\d{2}", month):
        raise HTTPException(400, "Invalid month")
    if _month_payroll_locked(employee_id, month):
        return RedirectResponse(f"/attendance/missing?employee_id={employee_id}&month={month}&error=locked", 303)
    days = _missing_duty_days(employee_id, month)
    actor = str(request.session.get("user_name") or request.session.get("hr_id") or "Super Admin")
    made = 0
    with get_db() as c:
        for d in days:
            exists = c.execute("SELECT id FROM attendance WHERE employee_id=? AND work_date=?",
                               (employee_id, d["date"])).fetchone()
            if exists:
                continue
            _write_backfill(c, employee_id, d["date"], d["start"], d["end"], "present", actor)
            made += 1
    audit(request, "backfill_all", "attendance", f"{employee_id}:{month}",
          f"Recorded {made} past duty days at scheduled times by {actor}")
    return RedirectResponse(f"/attendance/missing?employee_id={employee_id}&month={month}&saved=filled&made={made}", 303)


@app.post("/correction")
def create_correction(request: Request, employee_id: int=Form(...), work_date: str=Form(...), check_in: str=Form(""), check_out: str=Form(""), reason: str=Form(...)):
    require_permission(request,"attendance_edit")
    with get_db() as c: c.execute("INSERT INTO attendance_corrections(employee_id,work_date,requested_check_in,requested_check_out,reason,requested_by) VALUES(?,?,?,?,?,?)",(employee_id,work_date,check_in or None,check_out or None,reason,str(request.session.get('hr_id') or 'super_admin')))
    audit(request,"create","attendance_correction",str(employee_id),work_date); return RedirectResponse("/hr-operations?saved=1",303)

@app.post("/correction/{request_id}/{action}")
def decide_correction(request: Request, request_id: int, action: str):
    require_permission(request,"leave_manage"); status="approved" if action=="approve" else "rejected" if action=="reject" else None
    if not status: raise HTTPException(400)
    with get_db() as c:
        row=c.execute("SELECT * FROM attendance_corrections WHERE id=?",(request_id,)).fetchone()
        if row and status=="approved":
            ci=row['requested_check_in']; co=row['requested_check_out']
            if ci and len(ci)<=5: ci=f"{row['work_date']}T{ci}:00"
            if co and len(co)<=5: co=f"{row['work_date']}T{co}:00"
            c.execute("INSERT INTO attendance(employee_id,work_date,check_in,check_out,status,source) VALUES(?,?,?,?,?,?) ON CONFLICT(employee_id,work_date) DO UPDATE SET check_in=COALESCE(excluded.check_in,attendance.check_in),check_out=COALESCE(excluded.check_out,attendance.check_out),source='hr_correction',updated_at=CURRENT_TIMESTAMP",(row['employee_id'],row['work_date'],ci,co,'present','hr_correction'))
        c.execute("UPDATE attendance_corrections SET status=?,decided_by=?,decided_at=CURRENT_TIMESTAMP WHERE id=?",(status,str(request.session.get('hr_id') or 'super_admin'),request_id))
    audit(request,status,"attendance_correction",str(request_id),""); return RedirectResponse("/hr-operations?saved=1",303)

@app.post("/shifts")
def save_shift(request: Request, name: str=Form(...), start_time: str=Form(...), end_time: str=Form(...)):
    require_permission(request,"shift_manage")
    with get_db() as c: c.execute("INSERT INTO shifts(name,start_time,end_time) VALUES(?,?,?) ON CONFLICT(name) DO UPDATE SET start_time=excluded.start_time,end_time=excluded.end_time",(name.strip(),start_time,end_time))
    audit(request,"save","shift",name,f"{start_time}-{end_time}"); return RedirectResponse("/hr-operations?saved=1",303)

@app.post("/departments")
def save_department(request: Request, name: str=Form(...)):
    require_permission(request,"department_manage")
    with get_db() as c: c.execute("INSERT INTO departments(name) VALUES(?) ON CONFLICT(name) DO NOTHING",(name.strip(),))
    audit(request,"save","department",name,""); return RedirectResponse("/hr-operations?saved=1",303)


@app.get("/duty", response_class=HTMLResponse)
def duty_management_page(request: Request, saved: str="", error: str=""):
    require_permission(request,"duty_view"); can_manage=has_permission(request,"duty_manage")
    today=datetime.now(ZoneInfo(settings.timezone)).date()
    month_end=(today.replace(day=28)+timedelta(days=4)).replace(day=1)-timedelta(days=1)
    with get_db() as c:
        employees=c.execute("SELECT id,staff_id,name,department,shift FROM employees WHERE is_active ORDER BY staff_id").fetchall()
        assigned=c.execute("SELECT COUNT(DISTINCT employee_id) AS n FROM custom_duties WHERE duty_date>=? AND is_active",(today.isoformat(),)).fetchone()
        upcoming=c.execute("SELECT d.*,e.staff_id,e.name FROM custom_duties d JOIN employees e ON e.id=d.employee_id WHERE d.duty_date>=? ORDER BY d.duty_date,e.staff_id LIMIT 120",(today.isoformat(),)).fetchall()
    total=len(employees); assigned_n=int((assigned or {}).get('n',0)); unassigned=max(0,total-assigned_n)
    rules=get_shift_rules()
    rule_fields=[("first_start","First Shift start",rules[FIRST_START_KEY]),("first_end","First Shift end",rules[FIRST_END_KEY]),
                 ("second_start","Second Shift start",rules[SECOND_START_KEY]),("second_end","Second Shift end",rules[SECOND_END_KEY]),
                 ("second_cutoff","Second Shift detection cutoff",rules[CUTOFF_KEY])]
    rule_inputs=''.join(f"<div><label>{escape(label)}</label><input type='time' name='{name}' value='{escape(value)}' {'required' if can_manage else 'disabled'}><div class='sub'>{escape(format_time_12h(value))}</div></div>" for name,label,value in rule_fields)
    grace_input=f"<div><label>Late-grace minutes</label><input type='number' min='0' max='240' step='1' name='late_grace_minutes' value='{int(rules[GRACE_KEY])}' {'required' if can_manage else 'disabled'}><div class='sub'>Late minutes count only after this grace period.</div></div>"
    save_button="<div class='actions'><button class='btn' type='submit'>Save Shift Rules</button></div>" if can_manage else "<div class='sub'>You do not have permission to change shift rules.</div>"
    shift_card=f"""<section class='card duty-section' id='shiftRules'>
      <div class='card-head'><div><div class='eyebrow'>Global Defaults</div><h2>Shift Rules</h2><p class='sub'>Employee-এর নিজস্ব duty না থাকলে এই সময়গুলোই প্রযোজ্য হবে। একবার save করলে পরের মাসগুলোতেও এই নিয়ম চালু থাকবে।</p></div><span class='pill'>Overtime: Manual only</span></div>
      <form method='post' action='/duty/shift-rules'>
        <div class='shift-rule-grid'>{rule_inputs}{grace_input}</div>
        <div class='notice' style='margin-top:14px'><b>Priority:</b> Employee custom duty (একটি তারিখ) → Employee weekly duty (weekday) → এই global Shift Rules. অর্থাৎ কোনো employee-এর নিজস্ব duty থাকলে সেটিই আগে মানা হবে, global rule তখন প্রযোজ্য নয়।</div>
        <div class='sub' style='margin-top:8px'>Overtime কখনো নিজে থেকে যোগ হয় না — HR/Admin Payroll page-এ manually overtime hours ও rate দেবেন।</div>
        {save_button}
      </form>
    </section>"""
    employee_cards=''.join(f"<label class='employee-pick' data-search='{escape((e['staff_id']+' '+e['name']+' '+(e['department'] or '')).lower())}'><input type='checkbox' name='employee_ids' value='{e['id']}'><span><b>{escape(e['name'])}</b><small>{escape(e['staff_id'])} · {escape(e['department'] or 'No department')} · {escape(e['shift'] or 'morning')}</small></span></label>" for e in employees)
    manage=''
    if can_manage:
        manage=f"""
        <form method='post' action='/duty/bulk' id='bulkDutyForm'>
          <section class='card duty-section'>
            <div class='card-head'><div><div class='eyebrow'>Step 1</div><h2>Select Employees</h2><p class='sub'>সব employee একসাথে অথবা প্রয়োজনীয় employee বেছে নিন।</p></div><div class='actions'><button type='button' class='btn secondary' id='selectAllBtn'>Select All</button><button type='button' class='btn secondary' id='clearAllBtn'>Clear</button></div></div>
            <div class='duty-search'><input id='employeeDutySearch' placeholder='Search by name, Staff ID or department'><span id='selectedDutyCount' class='pill'>0 selected</span></div>
            <div class='employee-pick-grid'>{employee_cards}</div>
          </section>
          <section class='card duty-section'>
            <div class='card-head'><div><div class='eyebrow'>Step 2</div><h2>Duty Date Range</h2><p class='sub'>একদিন, এক সপ্তাহ অথবা পুরো মাসের duty একবারে assign করুন।</p></div></div>
            <div class='two'><div><label>Start Date</label><input id='dutyStartDate' type='date' name='start_date' value='{today.isoformat()}' required></div><div><label>End Date</label><input id='dutyEndDate' type='date' name='end_date' value='{month_end.isoformat()}' required></div></div>
            <div class='actions duty-shortcuts'><button type='button' class='btn secondary' data-range='today'>Today</button><button type='button' class='btn secondary' data-range='week'>This Week</button><button type='button' class='btn secondary' data-range='month'>This Month</button></div>
          </section>
          <section class='card duty-section'>
            <div class='card-head'><div><div class='eyebrow'>Step 3</div><h2>Weekly Duty Schedule</h2><p class='sub'>Friday আলাদা special duty; Sunday–Thursday এবং Saturday regular duty।</p></div></div>
            <div class='schedule-grid'>
              <div class='schedule-card regular'><div class='schedule-days'>Sunday · Monday · Tuesday · Wednesday · Thursday · Saturday</div><h3>Regular Duty</h3><div class='two'><div><label>Start Time</label><input type='time' name='regular_start' value='{rules[FIRST_START_KEY]}' required></div><div><label>End Time</label><input type='time' name='regular_end' value='{rules[FIRST_END_KEY]}' required></div></div><label>Break (minutes)</label><input type='number' name='regular_break_minutes' min='0' step='5' value='60'><label>Office</label><input name='office_name' value='BURAQ Office' required><label>Note (optional)</label><input name='regular_note' placeholder='Regular duty'></div>
              <div class='schedule-card friday'><div class='schedule-days'>Friday · Special Day</div><h3>Friday Duty</h3><div class='two'><div><label>Start Time</label><input type='time' name='friday_start' value='{rules[SECOND_START_KEY]}' required></div><div><label>End Time</label><input type='time' name='friday_end' value='{rules[SECOND_END_KEY]}' required></div></div><label>Break (minutes)</label><input type='number' name='friday_break_minutes' min='0' step='5' value='60'><label>Friday Note (optional)</label><input name='friday_note' value='Friday duty'></div>
            </div>
          </section>
          <section class='duty-actionbar'><div><b id='dutySummary'>Select employees and date range</b><div class='sub'>Existing duty থাকলে নতুন সময় দিয়ে update হবে।</div></div><div class='actions'><button type='button' class='btn secondary' id='previewDutyBtn'>Preview</button><button class='btn' type='submit'>Save Duty</button></div></section>
        </form>"""
    rows=''.join(f"<tr><td><b>{escape(r['duty_date'])}</b></td><td>{escape(r['staff_id'])} - {escape(r['name'])}</td><td>{escape(format_time_12h(r['start_time']))} - {escape(format_time_12h(r['end_time']))}{' (+1 day)' if r['end_time']<=r['start_time'] else ''}</td><td>{escape(r['office_name'] or 'BURAQ Office')}</td><td>{escape(r['note'] or '—')}</td></tr>" for r in upcoming) or '<tr><td colspan=5>No upcoming duty assigned.</td></tr>'
    saved_text={'shift':'Shift rules saved. These defaults stay active until you change them again.'}.get(saved,'Duty assigned successfully.')
    notice=f"<div class='notice'>{escape(saved_text)}</div>" if saved else (f"<div class='notice bad'>{escape(error)}</div>" if error else '')
    body=f"""{notice}<div class='hero'><div><div class='eyebrow'>Duty Management</div><h2>Manage Employee Duty</h2><div class='sub'>All employees, date range, regular schedule and separate Friday duty—সব এক page-এ। Saturday regular duty থাকবে।</div></div><div class='actions'><span class='pill'>Total {total}</span><span class='pill'>Assigned {assigned_n}</span><span class='pill'>Unassigned {unassigned}</span></div></div>{shift_card}{manage}<div class='section-gap'></div><div class='card' style='overflow:auto'><div class='card-head'><div><div class='eyebrow'>Assigned Duty</div><h2>Upcoming Duty List</h2></div><a class='btn secondary' href='/duty-schedules'>Advanced / Reminder Log</a></div><table><thead><tr><th>Date</th><th>Employee</th><th>Duty</th><th>Office</th><th>Note</th></tr></thead><tbody>{rows}</tbody></table></div>"""
    extra="""<style>
.duty-section{margin-top:16px}.shift-rule-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
@media(max-width:900px){.shift-rule-grid{grid-template-columns:1fr 1fr}}
@media(max-width:600px){.shift-rule-grid{grid-template-columns:1fr}}.duty-search{display:flex;gap:12px;align-items:center;margin:14px 0}.duty-search input{flex:1}.employee-pick-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;max-height:360px;overflow:auto;padding:2px}.employee-pick{display:flex;gap:10px;align-items:center;padding:12px;border:1px solid var(--line);border-radius:12px;background:var(--panel2);cursor:pointer}.employee-pick input{width:auto}.employee-pick span{display:grid;gap:2px}.employee-pick small{color:var(--muted)}.employee-pick:has(input:checked){border-color:var(--brand);background:rgba(8,127,91,.1)}.schedule-grid{display:grid;grid-template-columns:2fr 1fr;gap:14px}.schedule-card{padding:18px;border:1px solid var(--line);border-radius:14px;background:var(--panel2)}.schedule-card.friday{border-color:#7aa7ff;background:rgba(55,111,255,.08)}.schedule-days{font-size:12px;font-weight:850;color:var(--brand);text-transform:uppercase;letter-spacing:.06em}.friday .schedule-days{color:#376fff}.saturday-off{margin-top:12px}.duty-shortcuts{margin-top:12px}.duty-actionbar{position:sticky;bottom:12px;z-index:8;margin-top:16px;padding:14px 16px;border:1px solid var(--line);border-radius:14px;background:var(--panel);box-shadow:var(--shadow);display:flex;justify-content:space-between;align-items:center;gap:12px}
@media(max-width:900px){.employee-pick-grid{grid-template-columns:1fr 1fr}.schedule-grid{grid-template-columns:1fr}}
@media(max-width:600px){.employee-pick-grid{grid-template-columns:1fr}.duty-actionbar{align-items:stretch;flex-direction:column}.duty-actionbar .actions{width:100%}.duty-actionbar .btn{flex:1}}
</style><script>
document.addEventListener('DOMContentLoaded',()=>{
  const form=document.getElementById('bulkDutyForm');
  if(!form) return;
  const boxes=Array.from(form.querySelectorAll("input[name='employee_ids']"));
  const count=document.getElementById('selectedDutyCount');
  const summary=document.getElementById('dutySummary');
  const search=document.getElementById('employeeDutySearch');
  const start=document.getElementById('dutyStartDate');
  const end=document.getElementById('dutyEndDate');
  const defaults={start:start.value,end:end.value,regularStart:form.regular_start.value,regularEnd:form.regular_end.value,fridayStart:form.friday_start.value,fridayEnd:form.friday_end.value,office:form.office_name.value,regularNote:form.regular_note.value,fridayNote:form.friday_note.value};
  function update(){
    const n=boxes.filter(x=>x.checked).length;
    if(count) count.textContent=`${n} selected`;
    if(summary) summary.textContent=n?`${n} employee · ${start.value||'—'} to ${end.value||'—'}`:'Select employees and date range';
  }
  function showAll(){form.querySelectorAll('.employee-pick-grid .employee-pick').forEach(el=>el.style.display='flex');}
  boxes.forEach(x=>x.addEventListener('change',update));
  start.addEventListener('change',update); end.addEventListener('change',update);
  document.getElementById('selectAllBtn')?.addEventListener('click',()=>{boxes.forEach(x=>x.checked=true);update();});
  document.getElementById('clearAllBtn')?.addEventListener('click',()=>{
    boxes.forEach(x=>x.checked=false); search.value=''; showAll();
    start.value=defaults.start; end.value=defaults.end;
    form.regular_start.value=defaults.regularStart; form.regular_end.value=defaults.regularEnd;
    form.friday_start.value=defaults.fridayStart; form.friday_end.value=defaults.fridayEnd;
    form.office_name.value=defaults.office; form.regular_note.value=defaults.regularNote; form.friday_note.value=defaults.fridayNote;
    update();
  });
  search.addEventListener('input',()=>{const q=search.value.trim().toLowerCase();form.querySelectorAll('.employee-pick-grid .employee-pick').forEach(el=>{el.style.display=el.dataset.search.includes(q)?'flex':'none';});});
  document.querySelectorAll('[data-range]').forEach(b=>b.addEventListener('click',()=>{
    const now=new Date(),iso=d=>{const y=d.getFullYear(),m=String(d.getMonth()+1).padStart(2,'0'),day=String(d.getDate()).padStart(2,'0');return `${y}-${m}-${day}`};
    let a=new Date(now),z=new Date(now);
    if(b.dataset.range==='week'){const day=(now.getDay()+6)%7;a.setDate(now.getDate()-day);z=new Date(a);z.setDate(a.getDate()+6)}
    if(b.dataset.range==='month'){a=new Date(now.getFullYear(),now.getMonth(),1);z=new Date(now.getFullYear(),now.getMonth()+1,0)}
    start.value=iso(a);end.value=iso(z);update();
  }));
  document.getElementById('previewDutyBtn')?.addEventListener('click',()=>alert(`${summary.textContent}
Regular: Sunday–Thursday + Saturday
Friday: separate duty`));
  form.addEventListener('submit',e=>{
    if(!boxes.some(x=>x.checked)){e.preventDefault();alert('Select at least one employee.');return;}
    if(!start.value||!end.value||end.value<start.value){e.preventDefault();alert('End Date cannot be earlier than Start Date.');}
  });
  update();
});
</script>"""
    response=layout("Duty Management",body,request,"duty")
    return HTMLResponse(response.body.decode().replace('</body>',extra+'</body>'))

@app.post("/duty/shift-rules")
def save_shift_rules_route(request: Request, first_start: str=Form(...), first_end: str=Form(...),
                           second_start: str=Form(...), second_end: str=Form(...),
                           second_cutoff: str=Form(...), late_grace_minutes: int=Form(0)):
    """Persist the global shift rules. Employee duty always overrides them."""
    require_permission(request,"duty_manage")
    for value in (first_start,first_end,second_start,second_end,second_cutoff):
        if not re.fullmatch(r'\d{2}:\d{2}',str(value or '').strip()): raise HTTPException(400,'Invalid shift time')
    if late_grace_minutes<0 or late_grace_minutes>240: raise HTTPException(400,'Late grace must be between 0 and 240 minutes')
    try:
        rules=save_shift_rules(first_start,first_end,second_start,second_end,second_cutoff,late_grace_minutes)
    except ValueError as exc:
        return RedirectResponse(f"/duty?error={quote_plus(str(exc))}",303)
    audit(request,"save","shift_rules","global",
          f"First {rules[FIRST_START_KEY]}-{rules[FIRST_END_KEY]}; Second {rules[SECOND_START_KEY]}-{rules[SECOND_END_KEY]}; "
          f"cutoff {rules[CUTOFF_KEY]}; grace {rules[GRACE_KEY]}m; overtime manual-only")
    return RedirectResponse("/duty?saved=shift",303)

@app.post("/duty/bulk")
def save_bulk_duty(request: Request, employee_ids: list[int]=Form(...), start_date: str=Form(...), end_date: str=Form(...), regular_start: str=Form(...), regular_end: str=Form(...), friday_start: str=Form(...), friday_end: str=Form(...), regular_break_minutes: int=Form(0), friday_break_minutes: int=Form(0), office_name: str=Form("BURAQ Office"), regular_note: str=Form(""), friday_note: str=Form("Friday duty")):
    require_permission(request,"duty_manage")
    try:
        start=datetime.strptime(start_date,'%Y-%m-%d').date(); end=datetime.strptime(end_date,'%Y-%m-%d').date()
    except ValueError: raise HTTPException(400,'Invalid duty date')
    if end<start: raise HTTPException(400,'End Date cannot be earlier than Start Date')
    if (end-start).days>366: raise HTTPException(400,'Date range is too long')
    if not employee_ids: raise HTTPException(400,'Select at least one employee')
    for t in (regular_start,regular_end,friday_start,friday_end):
        if not re.fullmatch(r'\d{2}:\d{2}',t): raise HTTPException(400,'Invalid duty time')
    regular_break_minutes=_validated_break_minutes(regular_start,regular_end,regular_break_minutes)
    friday_break_minutes=_validated_break_minutes(friday_start,friday_end,friday_break_minutes)
    actor=str(request.session.get('hr_id') or 'super_admin'); created=0
    with get_db() as c:
        valid={int(x['id']) for x in c.execute("SELECT id FROM employees WHERE is_active").fetchall()}
        selected=sorted(set(employee_ids)&valid)
        if not selected: raise HTTPException(400,'No valid employee selected')
        current=start
        while current<=end:
            weekday=current.weekday()
            is_friday=weekday==4
            st,et=(friday_start,friday_end) if is_friday else (regular_start,regular_end)
            break_minutes=friday_break_minutes if is_friday else regular_break_minutes
            note=(friday_note if is_friday else regular_note).strip() or ('Friday duty' if is_friday else 'Regular duty')
            for employee_id in selected:
                c.execute("INSERT INTO custom_duties(employee_id,duty_date,start_time,end_time,break_minutes,office_name,note,created_by) VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(employee_id,duty_date) DO UPDATE SET start_time=excluded.start_time,end_time=excluded.end_time,break_minutes=excluded.break_minutes,office_name=excluded.office_name,note=excluded.note,is_active=excluded.is_active,updated_at=CURRENT_TIMESTAMP",(employee_id,current.isoformat(),st,et,break_minutes,office_name.strip() or 'BURAQ Office',note,actor)); created+=1
                c.execute("DELETE FROM duty_reminder_logs WHERE employee_id=? AND duty_date=?",(employee_id,current.isoformat()))
            current+=timedelta(days=1)
        audit(request,'bulk_assign','duty',f'{len(selected)} employees',f'{start_date} to {end_date}; {created} duty rows',db=c)
    return RedirectResponse('/duty?saved=1',303)

@app.get("/duty-schedules", response_class=HTMLResponse)
def duty_schedules_page(request: Request, saved: str=""):
    require_permission(request,"duty_view"); can_manage=has_permission(request,"duty_manage")
    with get_db() as c:
        employees=c.execute("SELECT id,staff_id,name FROM employees WHERE is_active ORDER BY staff_id").fetchall()
        rows=c.execute("SELECT d.*,e.staff_id,e.name FROM duty_schedules d JOIN employees e ON e.id=d.employee_id ORDER BY e.staff_id,d.weekday").fetchall()
        custom=c.execute("SELECT d.*,e.staff_id,e.name FROM custom_duties d JOIN employees e ON e.id=d.employee_id WHERE d.duty_date>=? ORDER BY d.duty_date,e.staff_id",(datetime.now(ZoneInfo(settings.timezone)).date().isoformat(),)).fetchall()
        logs=c.execute("SELECT l.*,e.staff_id,e.name FROM duty_reminder_logs l JOIN employees e ON e.id=l.employee_id ORDER BY l.id DESC LIMIT 50").fetchall()
    days=['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']; options=''.join(f"<option value='{e['id']}'>{escape(e['staff_id'])} - {escape(e['name'])}</option>" for e in employees)
    form=''
    if can_manage:
        day_options=''.join(f"<option value='{i}'>{d}</option>" for i,d in enumerate(days))
        form=f"""<div class='card'><h2>Assign Weekly Duty</h2><p class='sub'>একই employee ও weekday আবার save করলে schedule update হবে।</p><form method='post'><label>Employee</label><select name='employee_id'>{options}</select><label>Weekday</label><select name='weekday'>{day_options}</select><div class='two'><div><label>Start</label><input type='time' name='start_time' required></div><div><label>End</label><input type='time' name='end_time' required></div></div><label>Break (minutes)</label><input type='number' name='break_minutes' min='0' step='5' value='60'><label>Office</label><input name='office_name' value='BURAQ Office'><button class='btn'>Save Duty</button></form></div>"""
        custom_form=f"""<div class='card money-card'><div class='card-head'><div><div class='eyebrow'>Outside Weekly Roster</div><h2>Assign Custom Duty</h2></div><span class='tag'>One Day</span></div><p class='sub'>নির্দিষ্ট দিনের special duty weekly schedule-কে override করবে।</p><form method='post' action='/custom-duties'><label>Employee</label><select name='employee_id'>{options}</select><div class='two'><div><label>Duty Date</label><input type='date' name='duty_date' required></div><div><label>Office</label><input name='office_name' value='BURAQ Office'></div></div><div class='two'><div><label>Start</label><input type='time' name='start_time' required></div><div><label>End</label><input type='time' name='end_time' required></div></div><label>Break (minutes)</label><input type='number' name='break_minutes' min='0' step='5' value='60'><label>Note</label><input name='note' placeholder='Special duty reason (optional)'><button class='btn'>Save Custom Duty</button></form></div>"""
    else: custom_form=''
    roster=[]
    for r in rows:
        action=f"<form method='post' action='/duty-schedules/{r['id']}/delete' onsubmit=\"return confirm('Delete this duty?')\"><button class='btn danger'>Delete</button></form>" if can_manage else ''
        roster.append(f"<tr><td><b>{escape(r['staff_id'])}</b><div class='sub'>{escape(r['name'])}</div></td><td>{days[int(r['weekday'])]}</td><td>{escape(format_time_12h(r['start_time']))} - {escape(format_time_12h(r['end_time']))}</td><td>{escape(r['office_name'] or 'BURAQ Office')}</td><td><span class='status {'ok' if r['is_active'] else 'bad'}'>{'Active' if r['is_active'] else 'Off'}</span></td><td>{action}</td></tr>")
    log_rows=''.join(f"<tr><td>{escape(str(x['created_at']))}</td><td>{escape(x['staff_id'])} - {escape(x['name'])}</td><td>{escape(x['duty_date'])}</td><td>{escape(x['reminder_type'])}</td><td><span class='status ok'>{escape(x['status'])}</span></td></tr>" for x in logs) or '<tr><td colspan=5>No reminders sent yet.</td></tr>'
    custom_rows=[]
    for r in custom:
        action=f"<form method='post' action='/custom-duties/{r['id']}/delete' onsubmit=\"return confirm('Delete this custom duty?')\"><button class='btn danger'>Delete</button></form>" if can_manage else ''
        custom_rows.append(f"<tr><td><b>{escape(r['duty_date'])}</b></td><td>{escape(r['staff_id'])} - {escape(r['name'])}</td><td>{escape(format_time_12h(r['start_time']))} - {escape(format_time_12h(r['end_time']))}</td><td>{escape(r['office_name'] or 'BURAQ Office')}</td><td>{escape(r['note'] or '—')}</td><td>{action}</td></tr>")
    notice="<div class='notice'>Duty schedule saved.</div>" if saved else ''
    body=f"""{notice}<div class='hero'><div><div class='eyebrow'>Zero-Touch Workforce</div><h2>Duty Scheduler & Reminders</h2><div class='sub'>Weekly roster plus one-day custom duty with automatic reminders.</div></div><div class='actions'><span class='pill'>{len(rows)} weekly</span><span class='pill'>{len(custom)} custom</span></div></div><div class='two'>{form}<div class='card'><h2>Reminder Timing</h2><div class='salary-part'><span class='sub'>Before duty</span><b>30 minutes</b></div><div class='salary-part'><span class='sub'>Late alert</span><b>10 minutes after start</b></div><div class='salary-part'><span class='sub'>Checkout</span><b>10 minutes before end</b></div><p class='sub'>Custom duty থাকলে ওই দিনের weekly duty ও reminder override হবে।</p></div></div><div class='section-gap'></div>{custom_form}<div class='section-gap'></div><div class='card' style='overflow:auto'><h2>Upcoming Custom Duties</h2><table><thead><tr><th>Date</th><th>Employee</th><th>Duty</th><th>Office</th><th>Note</th><th></th></tr></thead><tbody>{''.join(custom_rows) or '<tr><td colspan=6>No upcoming custom duty.</td></tr>'}</tbody></table></div><div class='section-gap'></div><div class='card' style='overflow:auto'><h2>Weekly Roster</h2><table><thead><tr><th>Employee</th><th>Day</th><th>Duty</th><th>Office</th><th>Status</th><th></th></tr></thead><tbody>{''.join(roster) or '<tr><td colspan=6>No duty assigned.</td></tr>'}</tbody></table></div><div class='section-gap'></div><div class='card' style='overflow:auto'><h2>Recent Reminder Log</h2><table><thead><tr><th>Sent</th><th>Employee</th><th>Duty Date</th><th>Type</th><th>Status</th></tr></thead><tbody>{log_rows}</tbody></table></div>"""
    return layout("Duty Scheduler",body,request,"duty")

@app.post("/duty-schedules")
def save_duty_schedule(request: Request, employee_id: int=Form(...), weekday: int=Form(...), start_time: str=Form(...), end_time: str=Form(...), office_name: str=Form("BURAQ Office"), break_minutes: int=Form(0)):
    require_permission(request,"duty_manage")
    if weekday not in range(7) or not re.fullmatch(r"\d{2}:\d{2}",start_time) or not re.fullmatch(r"\d{2}:\d{2}",end_time): raise HTTPException(400,"Invalid duty schedule")
    break_minutes=_validated_break_minutes(start_time,end_time,break_minutes); actor=str(request.session.get('hr_id') or 'super_admin')
    with get_db() as c:
        c.execute("INSERT INTO duty_schedules(employee_id,weekday,start_time,end_time,break_minutes,office_name,created_by) VALUES(?,?,?,?,?,?,?) ON CONFLICT(employee_id,weekday) DO UPDATE SET start_time=excluded.start_time,end_time=excluded.end_time,break_minutes=excluded.break_minutes,office_name=excluded.office_name,is_active=excluded.is_active,updated_at=CURRENT_TIMESTAMP",(employee_id,weekday,start_time,end_time,break_minutes,office_name.strip() or 'BURAQ Office',actor))
        audit(request,'save','duty_schedule',f'{employee_id}:{weekday}',f'{start_time}-{end_time}',db=c)
    return RedirectResponse('/duty-schedules?saved=1',303)

@app.post("/duty-schedules/{schedule_id}/delete")
def delete_duty_schedule(request: Request, schedule_id: int):
    require_permission(request,"duty_manage")
    with get_db() as c:
        c.execute("DELETE FROM duty_schedules WHERE id=?",(schedule_id,)); audit(request,'delete','duty_schedule',str(schedule_id),db=c)
    return RedirectResponse('/duty-schedules',303)

@app.post("/custom-duties")
def save_custom_duty(request: Request, employee_id: int=Form(...), duty_date: str=Form(...), start_time: str=Form(...), end_time: str=Form(...), office_name: str=Form("BURAQ Office"), note: str=Form(""), break_minutes: int=Form(0)):
    require_permission(request,"duty_manage")
    try: datetime.strptime(duty_date,'%Y-%m-%d')
    except ValueError: raise HTTPException(400,'Invalid duty date')
    if not re.fullmatch(r"\d{2}:\d{2}",start_time) or not re.fullmatch(r"\d{2}:\d{2}",end_time): raise HTTPException(400,'Invalid duty time')
    break_minutes=_validated_break_minutes(start_time,end_time,break_minutes); actor=str(request.session.get('hr_id') or 'super_admin')
    with get_db() as c:
        c.execute("INSERT INTO custom_duties(employee_id,duty_date,start_time,end_time,break_minutes,office_name,note,created_by) VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(employee_id,duty_date) DO UPDATE SET start_time=excluded.start_time,end_time=excluded.end_time,break_minutes=excluded.break_minutes,office_name=excluded.office_name,note=excluded.note,is_active=excluded.is_active,updated_at=CURRENT_TIMESTAMP",(employee_id,duty_date,start_time,end_time,break_minutes,office_name.strip() or 'BURAQ Office',note.strip() or None,actor))
        c.execute("DELETE FROM duty_reminder_logs WHERE employee_id=? AND duty_date=?",(employee_id,duty_date))
        audit(request,'save','custom_duty',f'{employee_id}:{duty_date}',f'{start_time}-{end_time}',db=c)
    return RedirectResponse('/duty-schedules?saved=1',303)

@app.post("/custom-duties/{duty_id}/delete")
def delete_custom_duty(request: Request, duty_id: int):
    require_permission(request,"duty_manage")
    with get_db() as c: c.execute("DELETE FROM custom_duties WHERE id=?",(duty_id,)); audit(request,'delete','custom_duty',str(duty_id),db=c)
    return RedirectResponse('/duty-schedules',303)

@app.get("/duplicates")
def duplicate_analysis(request: Request, decision: str="", review: str="", scope: str="", saved: str="", error: str=""):
    require_permission(request, "approvals_view")
    # Land on the queue that needs a decision. `scope=all` opens the history.
    if not decision and not review and scope != "all":
        review = "pending"
    clauses, params = ["1=1"], []
    if decision in {"accept", "pending", "reject"}: clauses.append("f.decision=?"); params.append(decision)
    if review in {"none", "pending", "approved", "rejected"}: clauses.append("f.review_status=?"); params.append(review)
    image_field = "f.image_data" if review == "pending" and scope != "all" else "NULL AS image_data"
    row_limit = 60 if review == "pending" and scope != "all" else 300
    with get_db() as c:
        pending_order = "e.id, CASE WHEN f.action IN ('check_in','checkin','in') THEN 0 ELSE 1 END, f.created_at ASC" if review == "pending" and scope != "all" else "f.id DESC"
        rows = c.execute(f"""SELECT f.id,f.employee_id,f.action,f.media_id,{image_field},
            f.latitude,f.longitude,f.distance_meters,f.duplicate_score,f.hash_score,f.face_score,
            f.pose_score,f.landmark_score,f.matched_fingerprint_id,f.decision,f.review_status,
            f.attendance_applied,f.attendance_result,f.reviewed_by,f.reviewed_at,f.created_at,
            e.staff_id,e.name FROM attendance_fingerprints f JOIN employees e ON e.id=f.employee_id
            WHERE {" AND ".join(clauses)}
            ORDER BY CASE WHEN f.review_status='pending' THEN 0 ELSE 1 END, {pending_order} LIMIT {row_limit}""", tuple(params)).fetchall()
        waiting = int(c.execute("SELECT COUNT(*) c FROM attendance_fingerprints WHERE review_status='pending'").fetchone()["c"] or 0)
    can_manage=has_permission(request,"approvals_manage")
    thresholds=f"Accept &lt; {settings.duplicate_accept_below:.2f} • Review {settings.duplicate_accept_below:.2f}–{settings.duplicate_reject_at:.2f} • Reject ≥ {settings.duplicate_reject_at:.2f}"
    tabs=[("Needs review", "/duplicates?review=pending", review=="pending" and scope!="all", waiting),
          ("Approved", "/duplicates?review=approved", review=="approved", None),
          ("Rejected", "/duplicates?review=rejected", review=="rejected", None),
          ("All selfies", "/duplicates?scope=all", scope=="all", None)]
    tab_html="".join(f"<a class='tab{' active' if active else ''}' href='{url}'>{label}{f' ({count})' if count else ''}</a>" for label,url,active,count in tabs)
    empty="<div class='empty-state'>এই তালিকায় কিছু নেই। নতুন attendance selfie এলে এখানে দেখা যাবে।</div>"
    if review == "pending" and scope != "all":
        cards=[]
        for r in rows:
            risk_state="bad" if r["decision"]=="reject" else "warn" if r["decision"]=="pending" else "ok"
            risk_label="High duplicate risk" if r["decision"]=="reject" else "Needs review" if r["decision"]=="pending" else "Security clear"
            initials=escape("".join(part[:1] for part in str(r["name"]).split()[:2]).upper() or "E")
            photo=(f"<img src='data:image/jpeg;base64,{r['image_data']}' alt='{escape(str(r['name']))} attendance selfie' loading='lazy'>" if r["image_data"] else f"<div class='review-photo-fallback'>{initials}<span>Preview unavailable</span></div>")
            location=(f"Verified · {float(r['distance_meters']):.0f} m" if r["distance_meters"] is not None else "Verified · office radius not configured")
            action_label="Check in" if r["action"]=="check_in" else "Check out"
            submitted=escape(str(r["created_at"]).replace("T"," ")[:19])
            controls=""
            if can_manage:
                controls=f"""<div class='review-actions'>
                  <form method='post' action='/duplicates/{r['id']}/approve' onsubmit="const b=this.querySelector('button');b.disabled=true;b.textContent='Approving…'"><button class='btn' type='submit'>✓ Approve attendance</button></form>
                  <form method='post' action='/duplicates/{r['id']}/reject' onsubmit=\"return confirm('Reject this selfie and ask the employee to try again?')\"><button class='btn danger' type='submit'>✕ Reject &amp; ask again</button></form>
                </div>"""
            cards.append(f"""<article class='card selfie-review-card'>
              <div class='review-card-head'><div><div class='eyebrow'>Pending review · #{r['id']}</div><h3>{escape(str(r['name']))}</h3><div class='sub'>{escape(str(r['staff_id']))} · {action_label}</div></div><span class='status {risk_state}'>{risk_label}</span></div>
              <div class='review-card-grid'><div class='review-photo'>{photo}</div><div class='review-details'>
                <div class='review-facts'><div><span>Submitted</span><b>{submitted}</b></div><div><span>Face match</span><b>{float(r['face_score'] or 0)*100:.1f}%</b></div><div><span>Location</span><b>{location}</b></div><div><span>Duplicate score</span><b>{float(r['duplicate_score'] or 0)*100:.1f}%</b></div><div><span>Pose / Landmark</span><b>{float(r['pose_score'] or 0)*100:.0f}% / {float(r['landmark_score'] or 0)*100:.0f}%</b></div><div><span>Matched selfie</span><b>{'#'+str(r['matched_fingerprint_id']) if r['matched_fingerprint_id'] else 'None'}</b></div></div>
                {controls}<div class='sub'>Approve করলে submitted time অনুযায়ী attendance final হবে এবং employee WhatsApp confirmation পাবে।</div>
              </div></div></article>""")
        content=f"<div class='selfie-review-grid'>{''.join(cards)}</div>" if cards else empty
    else:
        history=[]
        for r in rows:
            state="bad" if r["review_status"]=="rejected" else "warn" if r["review_status"]=="pending" else "ok"
            applied="Final" if r["attendance_applied"] else ("Rejected" if r["review_status"]=="rejected" else "Not final")
            history.append(f"<tr><td>#{r['id']}</td><td><b>{escape(str(r['name']))}</b><br><span class='sub'>{escape(str(r['staff_id']))}</span></td><td>{'Check in' if r['action']=='check_in' else 'Check out'}</td><td><span class='status {state}'>{escape(str(r['review_status']).title())}</span><div class='sub'>{applied}</div></td><td>{float(r['face_score'] or 0)*100:.1f}%</td><td>{float(r['duplicate_score'] or 0)*100:.1f}%</td><td>{escape(str(r['attendance_result'] or '—'))}</td><td>{escape(str(r['created_at']))}</td></tr>")
        content=f"<div style='overflow:auto'><table><thead><tr><th>ID</th><th>Employee</th><th>Action</th><th>Status</th><th>Face</th><th>Duplicate</th><th>Attendance result</th><th>Submitted</th></tr></thead><tbody>{''.join(history)}</tbody></table></div>" if history else empty
    error_messages={"checkin-first":"আগে ওই employee-এর Check-in selfie approve করুন; এরপর Check-out approve হবে।","invalid-action":"এই পুরোনো selfie-এর action সঠিক নয়। Employee-কে নতুন attendance দিতে বলুন।","failed":"Approval সাময়িকভাবে সম্পন্ন হয়নি। আবার চেষ্টা করুন; সমস্যা থাকলে Deploy Logs দেখুন।"}
    notice = "<div class='notice'>Attendance selfie approved and finalized.</div>" if saved=="approved" else ("<div class='notice'>Selfie rejected; employee notification queued.</div>" if saved=="rejected" else (f"<div class='notice bad'>{error_messages.get(error,'Approval সম্পন্ন হয়নি।')}</div>" if error else ""))
    body=f"""<div class='hero'>
      <div><div class='eyebrow'>Attendance Approval</div><h2>Pending Selfies</h2><div class='sub'>প্রতিটি valid attendance selfie HR/Admin approve করার পর final হবে। · {thresholds}</div></div>
      <span class='pill'>{waiting} waiting</span>
    </div>
    {notice}<div class='tabs'>{tab_html}</div>{content}"""
    return layout("Selfie Review", body, request, "duplicates")

@app.post("/duplicates/{fingerprint_id}/{action}")
def review_duplicate(request: Request, fingerprint_id: int, action: str, background_tasks: BackgroundTasks):
    require_permission(request,"approvals_manage")
    if action not in {"approve","reject"}: raise HTTPException(400,"Invalid action")
    actor=str(request.session.get("hr_id") or "super_admin")
    notify=None
    if action == "approve":
        try:
            approved = approve_pending_attendance(fingerprint_id, actor)
        except ValueError as exc:
            code="checkin-first" if "Check-in approve" in str(exc) else "invalid-action"
            return RedirectResponse(f"/duplicates?review=pending&error={code}",303)
        except Exception:
            logger.exception("Selfie approval failed fingerprint=%s actor=%s",fingerprint_id,actor)
            return RedirectResponse("/duplicates?review=pending&error=failed",303)
        if not approved:
            with get_db() as c:
                existing=c.execute("SELECT review_status FROM attendance_fingerprints WHERE id=?",(fingerprint_id,)).fetchone()
            if existing and existing["review_status"]=="approved":
                return RedirectResponse("/duplicates?review=pending&saved=approved",303)
            if existing and existing["review_status"]=="rejected":
                return RedirectResponse("/duplicates?review=rejected",303)
            raise HTTPException(404,"Pending selfie not found")
        if approved["phone"]:
            notify=(approved["phone"],approved["name"],approved["action"],True,approved["score"],approved["result"])
        status="approved"
    else:
        status="rejected"
        with get_db() as c:
            row=c.execute("""SELECT f.id,f.action,f.duplicate_score,e.name,
                COALESCE(NULLIF(e.whatsapp_phone,''),NULLIF(e.phone,'')) notification_phone
                FROM attendance_fingerprints f JOIN employees e ON e.id=f.employee_id
                WHERE f.id=? AND f.review_status='pending'""",(fingerprint_id,)).fetchone()
            if not row: raise HTTPException(404,"Pending selfie not found or already reviewed")
            c.execute("UPDATE attendance_fingerprints SET review_status='rejected',reviewed_by=?,reviewed_at=CURRENT_TIMESTAMP,attendance_result='Rejected by HR/Admin' WHERE id=? AND review_status='pending'",(actor,fingerprint_id))
            if row["notification_phone"]:
                notify=(row["notification_phone"],row["name"],row["action"],False,float(row["duplicate_score"] or 0),"")
    audit(request,action,"attendance_fingerprint",str(fingerprint_id),status)
    if notify:
        background_tasks.add_task(send_selfie_review_result,*notify)
    else:
        logger.warning("Selfie review notification skipped: employee phone missing fingerprint=%s",fingerprint_id)
    return RedirectResponse(f"/duplicates?review=pending&saved={status}",303)

@app.get("/webhook/whatsapp", response_class=PlainTextResponse)
def verify(hub_mode: str | None = Query(None, alias="hub.mode"), hub_verify_token: str | None = Query(None, alias="hub.verify_token"), hub_challenge: str | None = Query(None, alias="hub.challenge")):
    if hub_mode == "subscribe" and hub_verify_token == get_setting("whatsapp_verify_token"):
        return hub_challenge or ""
    raise HTTPException(403, "Webhook verification failed")


def _location_link_employee(token: str):
    payload = verify_location_token(token)
    with get_db() as c:
        employee = c.execute(
            """SELECT id,name,COALESCE(NULLIF(whatsapp_phone,''),NULLIF(phone,'')) AS attendance_phone
               FROM employees WHERE id=? AND is_active""",
            (payload["employee_id"],),
        ).fetchone()
    if not employee or not employee["attendance_phone"]:
        raise ValueError("Employee is unavailable")
    expected_state = f"{payload['action']}_location"
    # Meta sends the sender as 8801..., while an employee profile is commonly
    # stored as 01... or +8801.... Find the actual conversation key instead of
    # assuming both strings use the same country-code format.
    with get_db() as c:
        pending_states = c.execute(
            "SELECT phone,state FROM conversation_states WHERE state=?",
            (expected_state,),
        ).fetchall()
    current_phone = next(
        (row["phone"] for row in pending_states if phones_match(row["phone"], employee["attendance_phone"])),
        "",
    )
    if not current_phone:
        raise ValueError("Location is no longer pending")
    return payload, employee, current_phone


@app.get("/attendance/location", response_class=HTMLResponse)
def attendance_location_page(t: str = Query(..., min_length=20)):
    try:
        _, employee, _ = _location_link_employee(t)
        employee_name = escape(str(employee["name"]))
        token_json = json.dumps(t)
        content = f"""
        <div class="card">
          <div class="pin">📍</div>
          <h1>Attendance Location</h1>
          <p class="hello">{employee_name}, আপনার বর্তমান location যাচাই করুন।</p>
          <div id="status" class="status">Location permission আসলে <b>Allow</b> দিন।</div>
          <button id="allow" type="button">Location Allow করুন</button>
          <p class="note">Location গ্রহণ হলে WhatsApp-এ ফিরে selfie পাঠাবেন।</p>
        </div>
        <script>
          const token = {token_json};
          const statusBox = document.getElementById('status');
          const allowButton = document.getElementById('allow');
          let requesting = false;
          function message(text, kind = '') {{
            statusBox.textContent = text;
            statusBox.className = 'status ' + kind;
          }}
          async function submitLocation(position) {{
            message('Location যাচাই হচ্ছে…', 'working');
            try {{
              const response = await fetch('/attendance/location/submit', {{
                method: 'POST',
                headers: {{'Content-Type': 'application/json'}},
                body: JSON.stringify({{
                  token,
                  latitude: position.coords.latitude,
                  longitude: position.coords.longitude,
                  accuracy: position.coords.accuracy
                }})
              }});
              const result = await response.json();
              if (!response.ok || !result.ok) throw new Error(result.message || 'Location যাচাই হয়নি।');
              message('✅ Location গ্রহণ হয়েছে। এখন WhatsApp-এ ফিরে selfie পাঠান।', 'success');
              allowButton.hidden = true;
            }} catch (error) {{
              message('⚠️ ' + error.message, 'error');
              allowButton.disabled = false;
              requesting = false;
            }}
          }}
          function requestLocation() {{
            if (requesting) return;
            if (!navigator.geolocation) {{
              message('এই browser Location support করে না। Chrome দিয়ে linkটি খুলুন।', 'error');
              return;
            }}
            requesting = true;
            allowButton.disabled = true;
            message('ফোনের Location permission-এ Allow দিন…', 'working');
            navigator.geolocation.getCurrentPosition(submitLocation, error => {{
              const help = error.code === 1
                ? 'Location permission বন্ধ আছে। Browser/App Settings থেকে Location Allow করে আবার চাপুন।'
                : 'Location পাওয়া যায়নি। GPS চালু করে আবার চেষ্টা করুন।';
              message('⚠️ ' + help, 'error');
              allowButton.disabled = false;
              requesting = false;
            }}, {{enableHighAccuracy: true, timeout: 20000, maximumAge: 0}});
          }}
          allowButton.addEventListener('click', requestLocation);
          window.addEventListener('load', requestLocation);
        </script>
        """
    except ValueError:
        content = """
        <div class="card">
          <div class="pin">⚠️</div>
          <h1>Linkটি আর কার্যকর নেই</h1>
          <p class="hello">Location ইতোমধ্যে গ্রহণ হয়েছে অথবা linkটির সময় শেষ হয়েছে।</p>
          <p class="note">WhatsApp menu থেকে Check In/Check Out আবার নির্বাচন করুন।</p>
        </div>
        """
    page = f"""<!doctype html><html lang="bn"><head>
      <meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1,viewport-fit=cover">
      <title>BURAQ Attendance Location</title>
      <style>
        *{{box-sizing:border-box}} body{{margin:0;min-height:100vh;display:grid;place-items:center;padding:22px;background:#f2f7f5;color:#10231d;font-family:system-ui,-apple-system,'Noto Sans Bengali',sans-serif}}
        .card{{width:min(100%,430px);background:#fff;border:1px solid #dce8e3;border-radius:24px;padding:30px 24px;text-align:center;box-shadow:0 18px 55px rgba(16,67,50,.12)}}
        .pin{{font-size:50px}} h1{{margin:8px 0;font-size:28px}} .hello{{font-size:17px;line-height:1.55;color:#465b54}}
        .status{{margin:22px 0;padding:16px;border-radius:14px;background:#eef7f3;line-height:1.5}} .status.success{{background:#dcfce7;color:#166534}} .status.error{{background:#fff1f2;color:#9f1239}} .status.working{{background:#eff6ff;color:#1e40af}}
        button{{width:100%;border:0;border-radius:14px;padding:16px;background:#07875f;color:#fff;font-size:18px;font-weight:750;cursor:pointer}} button:disabled{{opacity:.55}} .note{{margin:18px 0 0;color:#71827c;font-size:14px;line-height:1.5}}
      </style></head><body>{content}</body></html>"""
    return HTMLResponse(page, headers={"Cache-Control": "no-store, private"})


@app.post("/attendance/location/submit")
async def attendance_location_submit(request: Request):
    try:
        data = await request.json()
        token = str(data.get("token") or "")
        latitude = float(data.get("latitude"))
        longitude = float(data.get("longitude"))
        if not (-90 <= latitude <= 90 and -180 <= longitude <= 180):
            raise ValueError("Invalid coordinates")
        _, employee, conversation_phone = _location_link_employee(token)
        response = receive_location(conversation_phone, latitude, longitude)
        accepted = response.startswith("✅ Location গ্রহণ")
        await send_text(conversation_phone, response)
        if not accepted:
            return JSONResponse({"ok": False, "message": response.replace("\n", " ")}, status_code=422)
        return {"ok": True, "message": "Location accepted"}
    except (TypeError, ValueError, json.JSONDecodeError):
        return JSONResponse(
            {"ok": False, "message": "Linkটির সময় শেষ হয়েছে অথবা Location আর pending নেই।"},
            status_code=400,
        )

@app.post("/webhook/whatsapp")
async def webhook(request: Request, background_tasks: BackgroundTasks):
    payload = await request.json()
    # Meta needs an immediate 2xx. Download and Face AI continue only after the
    # acknowledgement, preventing Meta retries and lost selfie responses.
    background_tasks.add_task(handle, payload, settings.public_base_url or base_url(request))
    return {"status": "accepted"}
